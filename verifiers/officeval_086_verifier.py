#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动评估“数据1_复合饼图_Sheet2.xlsx”的复合饼图完成情况。

对外仅暴露 `evaluate(dir_path: str) -> dict`：接收脚本所在目录路径，
脚本自身负责在该目录里定位并打开被评估的文档，返回结构化评分字典。

评估原则：
1. 先检查“维度1：可用与可修改性”。任一关键项不满足，则直接判 0 分，不再评估维度2。
2. 维度1通过后，按维度2的得分点/扣分点逐项自动检测并累计分数。
3. 对 Excel 中较难精确还原的视觉要求（标签重叠、饼图内部真实圆心等），使用 OOXML 结构、图表锚点、图表配置、文本/样式等信息做启发式判断，并在 detail 中说明命中的点和未命中的原因。
"""

from __future__ import annotations

import json
import math
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field, replace as dataclass_replace
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

try:
    import openpyxl
except ImportError:  # 由 evaluate() 统一转换为 status="error"
    openpyxl = None

EMU_PER_CM = 360000
EMU_PER_PT = 12700
EMU_PER_PIXEL = 9525

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "officeRel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

EXPECTED_MAIN = {"人员薪酬": 0.55, "物料": 0.34, "其他": 0.11}
EXPECTED_DETAIL = {"交通": 0.15, "住宿": 0.65, "餐饮": 0.20}
ALLOWED_FONTS = {"微软雅黑", "Microsoft YaHei", "宋体", "SimSun", "Calibri"}
DEFAULT_OFFICE_PALETTE = [
    "4472C4",  # blue
    "ED7D31",  # orange
    "A5A5A5",  # gray
    "FFC000",  # yellow
    "5B9BD5",  # light blue
    "70AD47",  # green
    "7030A0",  # purple
    "00B0F0",  # cyan/teal-ish
]


@dataclass
class Anchor:
    left: float
    top: float
    width: float
    height: float
    xml_node: ET.Element
    rel_id: Optional[str] = None
    target: Optional[str] = None

    @property
    def right(self) -> float:
        return self.left + self.width

    @property
    def bottom(self) -> float:
        return self.top + self.height

    @property
    def cx(self) -> float:
        return self.left + self.width / 2

    @property
    def cy(self) -> float:
        return self.top + self.height / 2

    @property
    def diameter(self) -> float:
        return min(self.width, self.height)


@dataclass
class ChartInfo:
    target: str
    xml: ET.Element
    anchor: Optional[Anchor]
    chart_types: List[str]
    title: str
    categories: List[str]
    values: List[float]
    legend_pos: Optional[str]
    dlabel_flags: Dict[str, bool]
    colors: List[str]
    separator_lines: List[Tuple[Optional[str], Optional[float], Optional[str]]]
    font_info: Dict[str, Any]
    # 以下字段仅用于“单个 ofPieChart 复合饼图”被拆分出的虚拟饼图视图：
    # - of_pie：原始 ofPieChart 的拆分/连接线信息（主、拆分两个虚拟视图共享同一份）。
    # - pie_diameter：该虚拟饼图圆形本体的估算直径（EMU）；非虚拟视图为 None，按绘图区较小边估算。
    # - area_box：原始图表区域(left, top, right, bottom)，用于把饼图中心归一化到“图表区域内”而非两对象合并框。
    # - label_indices：本视图各分类在原始 series 里的数据点索引，用来根据 c:dLbl/c:idx 逐标签绑定；
    #   非虚拟视图为 None，按 range(len(categories)) 处理。
    of_pie: Optional[Dict[str, Any]] = None
    pie_diameter: Optional[float] = None
    area_box: Optional[Tuple[float, float, float, float]] = None
    label_indices: Optional[List[int]] = None

    def data_map(self) -> Dict[str, float]:
        return dict(zip(self.categories, self.values))


@dataclass
class LineInfo:
    left: float
    top: float
    width: float
    height: float
    color: Optional[str]
    width_pt: Optional[float]
    dash: Optional[str]
    # 连接线两端点（EMU 绝对坐标）。对普通直线，端点必为 bounding box 的两个对角，
    # 具体走向由 a:xfrm 的 flipH/flipV 决定：flipH XOR flipV → (left,bottom)-(right,top)，
    # 否则 → (left,top)-(right,bottom)。
    x1: float = 0.0
    y1: float = 0.0
    x2: float = 0.0
    y2: float = 0.0

    @property
    def mid_y(self) -> float:
        return self.top + self.height / 2


@dataclass
class Hit:
    score: int
    name: str
    passed: bool
    detail: str


@dataclass
class WorkbookModel:
    path: Path
    zip_names: List[str] = field(default_factory=list)
    workbook_xml: Optional[ET.Element] = None
    workbook_rels: Dict[str, str] = field(default_factory=dict)
    sheet_paths: Dict[str, str] = field(default_factory=dict)
    sheet_xmls: Dict[str, ET.Element] = field(default_factory=dict)
    drawing_paths: Dict[str, str] = field(default_factory=dict)
    charts: List[ChartInfo] = field(default_factory=list)
    lines: List[LineInfo] = field(default_factory=list)
    package_text: str = ""


# ----------------------------- 基础 XML / XLSX 工具 -----------------------------

def qname(ns_key: str, tag: str) -> str:
    return f"{{{NS[ns_key]}}}{tag}"


def parse_xml_from_zip(zf: zipfile.ZipFile, name: str) -> Optional[ET.Element]:
    try:
        return ET.fromstring(zf.read(name))
    except Exception:
        return None


def strip_leading_slash(path: str) -> str:
    return path[1:] if path.startswith("/") else path


def normalize_target(base_file: str, target: str) -> str:
    """Resolve an OOXML relationship target to a zip member path."""
    target = strip_leading_slash(target)
    if target.startswith("xl/") or target.startswith("docProps/") or target.startswith("_rels/"):
        return target
    base_dir = os.path.dirname(base_file).replace("\\", "/")
    joined = os.path.normpath(os.path.join(base_dir, target)).replace("\\", "/")
    return strip_leading_slash(joined)


def rels_path_for(part_name: str) -> str:
    part_name = strip_leading_slash(part_name)
    folder = os.path.dirname(part_name)
    base = os.path.basename(part_name)
    return f"{folder}/_rels/{base}.rels" if folder else f"_rels/{base}.rels"


def read_relationships(zf: zipfile.ZipFile, part_name: str) -> Dict[str, str]:
    rels_name = rels_path_for(part_name)
    if rels_name not in zf.namelist():
        return {}
    root = parse_xml_from_zip(zf, rels_name)
    if root is None:
        return {}
    rels: Dict[str, str] = {}
    for rel in root.findall("rel:Relationship", NS):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            rels[rid] = normalize_target(part_name, target)
    return rels


def cell_to_text(value: Any) -> str:
    return "" if value is None else str(value)


def almost_equal(a: float, b: float, tol: float = 0.015) -> bool:
    return abs(float(a) - float(b)) <= tol


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        number = float(text)
        return number / 100 if "%" in str(value) else number
    except ValueError:
        return None


def extract_text(elem: Optional[ET.Element]) -> str:
    if elem is None:
        return ""
    return "".join(t.text or "" for t in elem.iter() if t.tag.endswith("}t") or t.tag.endswith("}v"))


def get_bool_child(parent: Optional[ET.Element], child_name: str, default: bool = False) -> bool:
    if parent is None:
        return default
    child = parent.find(f"c:{child_name}", NS)
    if child is None:
        return default
    return child.attrib.get("val", "1") not in {"0", "false", "False"}


def hex_to_rgb(hex_color: str) -> Optional[Tuple[int, int, int]]:
    if not hex_color:
        return None
    s = hex_color.strip().replace("#", "")[-6:]
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", s):
        return None
    return tuple(int(s[i : i + 2], 16) for i in (0, 2, 4))


def color_family(hex_color: str) -> str:
    rgb = hex_to_rgb(hex_color)
    if rgb is None:
        return "unknown"
    r, g, b = rgb
    maxc, minc = max(rgb), min(rgb)
    if maxc - minc < 24:
        return "gray"
    # 简易 HSL hue 分类，足够用于评分中的颜色族判断。
    import colorsys

    h, l, s = colorsys.rgb_to_hls(r / 255, g / 255, b / 255)
    hue = h * 360
    if 190 <= hue <= 250:
        return "blue"
    if 15 <= hue <= 45:
        return "orange"
    if 155 <= hue <= 190:
        return "teal"
    if 255 <= hue <= 300:
        return "purple"
    if 75 <= hue <= 150:
        return "light_green" if l > 0.45 else "green"
    return "other"


def is_gray_or_light_gray(hex_color: Optional[str]) -> bool:
    rgb = hex_to_rgb(hex_color or "")
    if rgb is None:
        return False
    r, g, b = rgb
    return max(rgb) - min(rgb) <= 35 and sum(rgb) / 3 >= 120


def is_black_or_dark_gray(hex_color: Optional[str]) -> bool:
    if hex_color is None:
        # Excel 默认标题/标签一般为黑色；未显式设置时按默认黑色处理。
        return True
    rgb = hex_to_rgb(hex_color)
    if rgb is None:
        return False
    return sum(rgb) / 3 <= 90 and max(rgb) - min(rgb) <= 45


def line_width_pt(width_emu: Optional[str]) -> Optional[float]:
    if not width_emu:
        return None
    try:
        return int(width_emu) / EMU_PER_PT
    except ValueError:
        return None


def collect_srgb_colors(elem: ET.Element) -> List[str]:
    colors: List[str] = []
    for srgb in elem.findall(".//a:srgbClr", NS):
        val = srgb.attrib.get("val")
        if val:
            colors.append(val.upper())
    return colors


# ----------------------------- 包结构解析 -----------------------------

def load_model(path: Path) -> WorkbookModel:
    model = WorkbookModel(path=path)
    with zipfile.ZipFile(path) as zf:
        model.zip_names = zf.namelist()
        text_chunks: List[str] = []
        for name in model.zip_names:
            if name.endswith(".xml"):
                try:
                    text_chunks.append(zf.read(name).decode("utf-8", errors="ignore"))
                except Exception:
                    pass
        model.package_text = "\n".join(text_chunks)

        model.workbook_xml = parse_xml_from_zip(zf, "xl/workbook.xml")
        model.workbook_rels = read_relationships(zf, "xl/workbook.xml")
        if model.workbook_xml is not None:
            for sheet in model.workbook_xml.findall(".//main:sheet", NS):
                sheet_name = sheet.attrib.get("name")
                rid = sheet.attrib.get(qname("r", "id"))
                if sheet_name and rid and rid in model.workbook_rels:
                    model.sheet_paths[sheet_name] = model.workbook_rels[rid]

        for sheet_name, sheet_path in model.sheet_paths.items():
            root = parse_xml_from_zip(zf, sheet_path)
            if root is not None:
                model.sheet_xmls[sheet_name] = root
                drawing = root.find("main:drawing", NS)
                if drawing is not None:
                    rid = drawing.attrib.get(qname("r", "id"))
                    rels = read_relationships(zf, sheet_path)
                    if rid and rid in rels:
                        model.drawing_paths[sheet_name] = rels[rid]

        if "Sheet2" in model.drawing_paths:
            parse_sheet2_drawing_and_charts(zf, model)
    return model


def column_width_to_emu(width_chars: float) -> float:
    # Excel 的列宽单位不是像素。这里采用常见近似公式，足以用于图表锚点的相对位置判断。
    pixels = int(width_chars * 7 + 5) if width_chars >= 1 else int(width_chars * 12)
    return pixels * EMU_PER_PIXEL


def row_height_to_emu(height_pt: float) -> float:
    return height_pt * EMU_PER_PT


def sheet_dimension_maps(sheet_root: Optional[ET.Element]) -> Tuple[Dict[int, float], Dict[int, float]]:
    col_widths: Dict[int, float] = {}
    row_heights: Dict[int, float] = {}
    if sheet_root is None:
        return col_widths, row_heights

    for col_node in sheet_root.findall(".//main:col", NS):
        try:
            min_col = int(col_node.attrib.get("min", "1"))
            max_col = int(col_node.attrib.get("max", str(min_col)))
            width = float(col_node.attrib.get("width", "8.43"))
        except ValueError:
            continue
        for one_based_col in range(min_col, max_col + 1):
            col_widths[one_based_col - 1] = column_width_to_emu(width)

    default_row_height = 15.0
    sheet_format = sheet_root.find("main:sheetFormatPr", NS)
    if sheet_format is not None:
        try:
            default_row_height = float(sheet_format.attrib.get("defaultRowHeight", default_row_height))
        except ValueError:
            pass
    row_heights[-1] = row_height_to_emu(default_row_height)

    for row_node in sheet_root.findall(".//main:row", NS):
        try:
            row_idx = int(row_node.attrib.get("r", "1")) - 1
            height = float(row_node.attrib.get("ht", str(default_row_height)))
        except ValueError:
            continue
        row_heights[row_idx] = row_height_to_emu(height)

    return col_widths, row_heights


def anchor_from_xdr(anchor_node: ET.Element, col_widths: Optional[Dict[int, float]] = None, row_heights: Optional[Dict[int, float]] = None) -> Anchor:
    from_node = anchor_node.find("xdr:from", NS)
    ext = anchor_node.find("xdr:ext", NS)
    col = int(from_node.findtext("xdr:col", "0", NS)) if from_node is not None else 0
    row = int(from_node.findtext("xdr:row", "0", NS)) if from_node is not None else 0
    col_off = int(from_node.findtext("xdr:colOff", "0", NS)) if from_node is not None else 0
    row_off = int(from_node.findtext("xdr:rowOff", "0", NS)) if from_node is not None else 0

    col_widths = col_widths or {}
    row_heights = row_heights or {}
    default_col_width = column_width_to_emu(8.43)
    default_row_height = row_heights.get(-1, row_height_to_emu(15.0))
    left = sum(col_widths.get(i, default_col_width) for i in range(col)) + col_off
    top = sum(row_heights.get(i, default_row_height) for i in range(row)) + row_off
    width = int(ext.attrib.get("cx", "0")) if ext is not None else 0
    height = int(ext.attrib.get("cy", "0")) if ext is not None else 0
    return Anchor(left=left, top=top, width=width, height=height, xml_node=anchor_node)


def parse_sheet2_drawing_and_charts(zf: zipfile.ZipFile, model: WorkbookModel) -> None:
    drawing_path = model.drawing_paths["Sheet2"]
    drawing_root = parse_xml_from_zip(zf, drawing_path)
    if drawing_root is None:
        return
    rels = read_relationships(zf, drawing_path)
    chart_anchors: List[Anchor] = []
    col_widths, row_heights = sheet_dimension_maps(model.sheet_xmls.get("Sheet2"))

    for node in list(drawing_root):
        local = node.tag.split("}")[-1]
        if local not in {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}:
            continue
        anchor = anchor_from_xdr(node, col_widths, row_heights)
        chart_ref = node.find(".//c:chart", NS)
        if chart_ref is not None:
            rid = chart_ref.attrib.get(qname("r", "id"))
            anchor.rel_id = rid
            anchor.target = rels.get(rid or "")
            chart_anchors.append(anchor)
        if node.find(".//xdr:sp", NS) is not None or node.find(".//xdr:cxnSp", NS) is not None:
            model.lines.extend(parse_lines_from_anchor(node, anchor))

    for anchor in chart_anchors:
        if not anchor.target or anchor.target not in model.zip_names:
            continue
        chart_xml = parse_xml_from_zip(zf, anchor.target)
        if chart_xml is None:
            continue
        model.charts.append(parse_chart_info(anchor.target, chart_xml, anchor))


def parse_lines_from_anchor(node: ET.Element, anchor: Anchor) -> List[LineInfo]:
    lines: List[LineInfo] = []
    for shape in node.findall(".//xdr:sp", NS) + node.findall(".//xdr:cxnSp", NS):
        ln = shape.find(".//a:ln", NS)
        if ln is None:
            continue
        colors = collect_srgb_colors(ln)
        dash = None
        dash_node = ln.find("a:prstDash", NS)
        if dash_node is not None:
            dash = dash_node.attrib.get("val")
        # 直线端点：连接线是直线，端点是外框的一条对角线。
        # a:xfrm 的 flipH/flipV 决定走向：异或为真 → 从左下到右上，否则 → 从左上到右下。
        xfrm = shape.find(".//a:xfrm", NS)
        flip_h = xfrm is not None and xfrm.attrib.get("flipH") in {"1", "true"}
        flip_v = xfrm is not None and xfrm.attrib.get("flipV") in {"1", "true"}
        if flip_h ^ flip_v:
            x1, y1, x2, y2 = anchor.left, anchor.bottom, anchor.right, anchor.top
        else:
            x1, y1, x2, y2 = anchor.left, anchor.top, anchor.right, anchor.bottom
        lines.append(
            LineInfo(
                left=anchor.left,
                top=anchor.top,
                width=anchor.width,
                height=anchor.height,
                color=colors[0] if colors else None,
                width_pt=line_width_pt(ln.attrib.get("w")),
                dash=dash,
                x1=x1,
                y1=y1,
                x2=x2,
                y2=y2,
            )
        )
    return lines


def parse_chart_info(target: str, root: ET.Element, anchor: Optional[Anchor]) -> ChartInfo:
    chart_types = []
    for candidate in ["pieChart", "ofPieChart", "doughnutChart", "pie3DChart"]:
        if root.find(f".//c:{candidate}", NS) is not None:
            chart_types.append(candidate)

    title_node = root.find(".//c:chart/c:title", NS)
    title = extract_text(title_node)

    chart = root.find(".//c:ofPieChart", NS)
    if chart is None:
        chart = root.find(".//c:pieChart", NS)
    if chart is None:
        chart = root.find(".//c:pie3DChart", NS)
    ser = chart.find("c:ser", NS) if chart is not None else None
    categories = parse_categories(ser)
    values = parse_values(ser)

    legend_node = root.find(".//c:legend", NS)
    legend_pos_node = legend_node.find("c:legendPos", NS) if legend_node is not None else None
    legend_pos = legend_pos_node.attrib.get("val") if legend_pos_node is not None else None

    dlabels_node = root.find(".//c:dLbls", NS)
    dlabel_flags = {
        "showVal": get_bool_child(dlabels_node, "showVal"),
        "showCatName": get_bool_child(dlabels_node, "showCatName"),
        "showSerName": get_bool_child(dlabels_node, "showSerName"),
        "showPercent": get_bool_child(dlabels_node, "showPercent"),
        "showLegendKey": get_bool_child(dlabels_node, "showLegendKey"),
    }

    colors = parse_point_colors(ser, len(categories))
    separator_lines = parse_separator_lines(ser)
    font_info = parse_font_info(root)

    of_pie_meta = parse_of_pie_meta(chart) if chart is not None and chart.tag.endswith("ofPieChart") else None

    return ChartInfo(
        target=target,
        xml=root,
        anchor=anchor,
        chart_types=chart_types,
        title=title,
        categories=categories,
        values=values,
        legend_pos=legend_pos,
        dlabel_flags=dlabel_flags,
        colors=colors,
        separator_lines=separator_lines,
        font_info=font_info,
        of_pie=of_pie_meta,
    )


def parse_of_pie_meta(of_pie_chart: ET.Element) -> Dict[str, Any]:
    """解析单个 ofPieChart 的复合饼图元数据。

    ofPieChart 只有一个 ``c:ser``（首个 series 决定显示，后续被忽略），
    其数据点通过 ``c:splitType``+``c:splitPos``（或 ``c:custSplit``）被划分为
    “主饼图”和“右侧拆分饼图”两部分。这里读取足够的字段，供后续把单一 series
    切成主/拆分两个虚拟视图。
    """
    def _val(elem_name: str, default: Optional[str] = None) -> Optional[str]:
        node = of_pie_chart.find(f"c:{elem_name}", NS)
        if node is None:
            return default
        return node.attrib.get("val", default)

    def _int(elem_name: str, default: Optional[int] = None) -> Optional[int]:
        raw = _val(elem_name)
        if raw is None:
            return default
        try:
            return int(raw)
        except ValueError:
            return default

    # ofPieType: pie（Pie of Pie，默认）/ bar（Bar of Pie）
    of_pie_type = _val("ofPieType", "pie")
    # splitType 决定 splitPos/custSplit 的解释方式：pos/val/percent/cust
    split_type = _val("splitType", "auto")
    split_pos = None
    raw_pos = _val("splitPos")
    if raw_pos is not None:
        try:
            split_pos = float(raw_pos)
        except ValueError:
            split_pos = None
    # custSplit：显式列出被划入右侧拆分饼图的数据点索引
    cust_indices: List[int] = []
    cust_split = of_pie_chart.find("c:custSplit", NS)
    if cust_split is not None:
        for pt in cust_split.findall("c:secondPiePt", NS):
            try:
                cust_indices.append(int(pt.attrib.get("val", "-1")))
            except ValueError:
                continue
    # secondPieSize：右侧拆分饼图相对左侧主饼图直径的百分比，默认 75
    second_pie_size = _int("secondPieSize", 75)
    gap_width = _int("gapWidth", 100)
    # serLines：ofPieChart 自身携带的连接线元素（连接主饼图切分扇区与右侧拆分饼图）
    ser_lines_node = of_pie_chart.find("c:serLines", NS)
    ser_line_color = None
    ser_line_dash = None
    ser_line_width_pt = None
    if ser_lines_node is not None:
        ln = ser_lines_node.find(".//a:ln", NS)
        if ln is not None:
            colors_found = collect_srgb_colors(ln)
            ser_line_color = colors_found[0] if colors_found else None
            dash_node = ln.find("a:prstDash", NS)
            ser_line_dash = dash_node.attrib.get("val") if dash_node is not None else None
            ser_line_width_pt = line_width_pt(ln.attrib.get("w"))
    return {
        "of_pie_type": of_pie_type,
        "split_type": split_type,
        "split_pos": split_pos,
        "cust_indices": cust_indices,
        "second_pie_size": second_pie_size,
        "gap_width": gap_width,
        "has_ser_lines": ser_lines_node is not None,
        "ser_line_color": ser_line_color,
        "ser_line_dash": ser_line_dash,
        "ser_line_width_pt": ser_line_width_pt,
    }


def parse_categories(ser: Optional[ET.Element]) -> List[str]:
    if ser is None:
        return []
    cat = ser.find("c:cat", NS)
    if cat is None:
        return []
    pts = cat.findall(".//c:pt", NS)
    result: List[Tuple[int, str]] = []
    for pt in pts:
        idx = int(pt.attrib.get("idx", len(result)))
        v = pt.find("c:v", NS)
        result.append((idx, v.text if v is not None and v.text is not None else ""))
    return [v for _, v in sorted(result)]


def parse_values(ser: Optional[ET.Element]) -> List[float]:
    if ser is None:
        return []
    val = ser.find("c:val", NS)
    if val is None:
        return []
    pts = val.findall(".//c:pt", NS)
    result: List[Tuple[int, float]] = []
    for pt in pts:
        idx = int(pt.attrib.get("idx", len(result)))
        v = pt.find("c:v", NS)
        number = safe_float(v.text if v is not None else None)
        if number is not None:
            result.append((idx, number))
    return [v for _, v in sorted(result)]


def parse_point_colors(ser: Optional[ET.Element], count: int) -> List[str]:
    if ser is None:
        return []
    colors_by_idx: Dict[int, str] = {}
    for dpt in ser.findall("c:dPt", NS):
        idx_node = dpt.find("c:idx", NS)
        if idx_node is None:
            continue
        try:
            idx = int(idx_node.attrib.get("val", "0"))
        except ValueError:
            continue
        colors = collect_srgb_colors(dpt)
        if colors:
            colors_by_idx[idx] = colors[0]

    ser_sppr = ser.find("c:spPr", NS)
    ser_colors = collect_srgb_colors(ser_sppr) if ser_sppr is not None else []
    if colors_by_idx:
        return [colors_by_idx.get(i, ser_colors[0] if ser_colors else "") for i in range(count)]
    if ser_colors and count > 0:
        # 只有系列级颜色时，所有扇区会使用同一填充色，不能当作三种互相区分的扇区配色。
        return [ser_colors[0]] * count

    # 未显式写入扇区颜色时，OOXML 中无法证明使用了题目要求的三种指定色系；不要用 Office 默认调色板代替真实设置。
    return []


def parse_separator_lines(ser: Optional[ET.Element]) -> List[Tuple[Optional[str], Optional[float], Optional[str]]]:
    if ser is None:
        return []
    result = []
    for sppr in ser.findall(".//c:spPr", NS):
        ln = sppr.find("a:ln", NS)
        if ln is None:
            continue
        colors = collect_srgb_colors(ln)
        dash = ln.find("a:prstDash", NS)
        result.append((colors[0] if colors else None, line_width_pt(ln.attrib.get("w")), dash.attrib.get("val") if dash is not None else None))
    return result


def parse_font_info(root: ET.Element) -> Dict[str, Any]:
    typefaces: List[str] = []
    sizes_pt: List[float] = []
    colors: List[str] = []
    for latin in root.findall(".//a:latin", NS):
        face = latin.attrib.get("typeface")
        if face:
            typefaces.append(face)
    for ea in root.findall(".//a:ea", NS):
        face = ea.attrib.get("typeface")
        if face:
            typefaces.append(face)
    for rpr in root.findall(".//a:rPr", NS) + root.findall(".//a:defRPr", NS):
        if "sz" in rpr.attrib:
            try:
                sizes_pt.append(int(rpr.attrib["sz"]) / 100)
            except ValueError:
                pass
        colors.extend(collect_srgb_colors(rpr))
    return {"typefaces": typefaces, "sizes_pt": sizes_pt, "colors": colors}


def has_wordart_effects(root: ET.Element) -> bool:
    """已弃用：不再作为扣分判定使用，保留空实现以兼容旧引用。"""
    return False


# ----------------------------- 维度1 -----------------------------

def dimension1_checks(path: Path, model: Optional[WorkbookModel], wb_formula: Optional[Any], wb_values: Optional[Any]) -> Tuple[bool, List[str], List[str]]:
    passed: List[str] = []
    failed: List[str] = []

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        passed.append("交付文件为 .xlsx 或 .xlsm 格式")
    else:
        failed.append("交付文件不是 .xlsx/.xlsm 格式")

    if wb_formula is not None and model is not None:
        passed.append("文件可被 openpyxl 和 zip/xml 解析，视为可正常打开")
    else:
        failed.append("文件无法正常打开或无法解析为有效 Excel 工作簿")
        return False, passed, failed

    return not failed, passed, failed


def overlap_ratio(a: Anchor, b: Anchor) -> float:
    x_overlap = max(0.0, min(a.right, b.right) - max(a.left, b.left))
    y_overlap = max(0.0, min(a.bottom, b.bottom) - max(a.top, b.top))
    inter = x_overlap * y_overlap
    if inter <= 0:
        return 0.0
    return inter / max(min(a.width * a.height, b.width * b.height), 1.0)


# ----------------------------- 维度2评分 -----------------------------

def score_dimension2(model: WorkbookModel, wb_formula: Any, wb_values: Any) -> List[Hit]:
    hits: List[Hit] = []
    main_chart, detail_chart = identify_main_detail_charts(model.charts)
    all_bounds = chart_group_bounds([c for c in [main_chart, detail_chart] if c and c.anchor])

    hits.append(check_core_layout_and_data(model, main_chart, detail_chart, all_bounds, wb_values))
    hits.append(check_main_labels(main_chart))
    hits.append(check_detail_labels(detail_chart))
    hits.append(check_colors_and_separators(main_chart, detail_chart))
    hits.append(check_connector_lines(model, main_chart, detail_chart, all_bounds))
    hits.append(check_relative_position(main_chart, detail_chart))
    hits.append(check_legend(main_chart, detail_chart))
    hits.append(check_title_area(model, main_chart, detail_chart))

    return hits


def identify_main_detail_charts(charts: Sequence[ChartInfo]) -> Tuple[Optional[ChartInfo], Optional[ChartInfo]]:
    pie_charts = [c for c in charts if any(t in c.chart_types for t in ["pieChart", "ofPieChart", "pie3DChart"])]

    # 情况一：单个 ofPieChart 对象一次性承载了主/拆分共 6 个扇区，需按 splitPos/custSplit
    # 把单一 series 切成主饼图和右侧拆分饼图两个虚拟视图。
    for c in pie_charts:
        if "ofPieChart" not in c.chart_types:
            continue
        cats = set(c.categories)
        if set(EXPECTED_MAIN).issubset(cats) and set(EXPECTED_DETAIL).issubset(cats):
            main_view, detail_view = build_ofpie_virtual_views(c)
            if main_view is not None and detail_view is not None:
                return main_view, detail_view

    main = None
    detail = None
    for c in pie_charts:
        cats = set(c.categories)
        if set(EXPECTED_MAIN).issubset(cats):
            main = c
        if set(EXPECTED_DETAIL).issubset(cats):
            detail = c
    if main is None and pie_charts:
        main = sorted(pie_charts, key=lambda c: c.anchor.left if c.anchor else 0)[0]
    if detail is None and len(pie_charts) > 1:
        detail = sorted(pie_charts, key=lambda c: c.anchor.left if c.anchor else 0)[-1]
    return main, detail


def _split_indices_for_second_pie(chart: ChartInfo) -> List[int]:
    """按 ofPieChart 元数据判断哪些数据点落入右侧拆分饼图。

    - splitType=cust：由 custSplit/secondPiePt 显式列出。
    - splitType=pos：末尾 splitPos 个点划入拆分饼图（OOXML 定义）。
    - splitType=val：值不大于 splitPos 的点划入拆分饼图。
    - splitType=percent：百分比不大于 splitPos 的点划入拆分饼图。
    - splitType=auto 或缺省：优先按“拆分饼图应包含的三个明细分类名”判断，
      再退化为末尾若干个点。
    """
    meta = chart.of_pie or {}
    n = len(chart.categories)
    split_type = (meta.get("split_type") or "auto")
    split_pos = meta.get("split_pos")
    cust_indices = meta.get("cust_indices") or []
    if split_type == "cust" and cust_indices:
        return sorted(i for i in cust_indices if 0 <= i < n)
    if split_type == "pos" and split_pos is not None:
        k = max(0, min(n, int(split_pos)))
        return list(range(n - k, n))
    if split_type == "val" and split_pos is not None and chart.values:
        return [i for i, v in enumerate(chart.values) if v <= split_pos]
    if split_type == "percent" and split_pos is not None and chart.values:
        total = sum(v for v in chart.values if v is not None) or 1.0
        return [i for i, v in enumerate(chart.values) if (v / total) <= split_pos]
    # auto/未知：优先按分类名归组
    detail_by_name = [i for i, name in enumerate(chart.categories) if name in EXPECTED_DETAIL]
    if len(detail_by_name) == len(EXPECTED_DETAIL):
        return sorted(detail_by_name)
    # 最后兜底：末尾 len(EXPECTED_DETAIL) 个点
    return list(range(max(0, n - len(EXPECTED_DETAIL)), n))


def build_ofpie_virtual_views(chart: ChartInfo) -> Tuple[Optional[ChartInfo], Optional[ChartInfo]]:
    """把单个 ofPieChart 拆成 (主饼图视图, 拆分饼图视图) 两个虚拟 ChartInfo。

    位置估算采用“绘图区宽度 + splitPos 简化模型”：
      - 图表锚点向内留出上下左右一小段边距近似绘图区；
      - 按 1:secondPieSize 把绘图区宽度切成主/拆分两块，两饼图各自居中于所在块内，
        因此主饼图中心天然落在绘图区左侧、拆分饼图落在右侧；
      - 饼图直径受所在块宽度与绘图区高度共同约束（正圆），gapWidth 越大直径越紧凑；
      - y 方向两饼图中心均取绘图区垂直中点。
    这样得到的中心相对“图表区域”的归一化坐标即可直接与细则 25%–35% / 72%–82% 比对。
    """
    if chart.anchor is None or not chart.categories:
        return None, None
    meta = chart.of_pie or {}
    n = len(chart.categories)
    second_indices = set(_split_indices_for_second_pie(chart))
    main_indices = [i for i in range(n) if i not in second_indices]
    detail_indices = [i for i in range(n) if i in second_indices]
    if not main_indices or not detail_indices:
        return None, None

    a = chart.anchor
    # 绘图区近似：预留左右 5%、上部 15%（标题）、下部 10%（图例/边距）。
    plot_left = a.left + 0.05 * a.width
    plot_right = a.right - 0.05 * a.width
    plot_top = a.top + 0.15 * a.height
    plot_bottom = a.bottom - 0.10 * a.height
    plot_w = max(plot_right - plot_left, 1.0)
    plot_h = max(plot_bottom - plot_top, 1.0)
    plot_cy = (plot_top + plot_bottom) / 2

    second_ratio = max(0.25, min(2.0, (meta.get("second_pie_size") or 75) / 100.0))
    # 简化模型：把绘图区宽度按 1:secondPieSize 划成主/拆分两块，两饼图分别居中于各自块内；
    # gapWidth 隐含在“饼图不完全填满块”里，用一个直径缩放系数近似。gapWidth 只影响直径估计，不影响两饼图中心的相对水平位置。
    gap_ratio = max(0.0, min(5.0, (meta.get("gap_width") or 150) / 100.0))
    portion_m = 1.0 / (1.0 + second_ratio)
    portion_s = second_ratio / (1.0 + second_ratio)
    main_cx = plot_left + 0.5 * portion_m * plot_w
    detail_cx = plot_left + (portion_m + 0.5 * portion_s) * plot_w
    # 饼图直径受块宽与绘图区高度共同约束（正圆）；gapWidth 越大，饼图相对块宽度越紧凑。
    diameter_shrink = 1.0 / (1.0 + 0.15 * gap_ratio)
    d_main = min(portion_m * plot_w * diameter_shrink, plot_h * 0.85)
    d_second = d_main * second_ratio

    def _virtual_anchor(cx: float, cy: float, diameter: float) -> Anchor:
        # 用与饼图圆形本体等大的正方形做虚拟外框，overlap_ratio、水平间距等下游几何检查都能直接复用。
        return Anchor(
            left=cx - diameter / 2,
            top=cy - diameter / 2,
            width=diameter,
            height=diameter,
            xml_node=a.xml_node,
            rel_id=a.rel_id,
            target=a.target,
        )

    def _slice(indices: List[int]) -> Tuple[List[str], List[float], List[str]]:
        cats = [chart.categories[i] for i in indices]
        vals = [chart.values[i] if i < len(chart.values) else 0.0 for i in indices]
        cols = [chart.colors[i] if i < len(chart.colors) else "" for i in indices]
        return cats, vals, cols

    area_box = (a.left, a.top, a.right, a.bottom)

    m_cats, m_vals, m_cols = _slice(main_indices)
    d_cats, d_vals, d_cols = _slice(detail_indices)

    main_view = dataclass_replace(
        chart,
        anchor=_virtual_anchor(main_cx, plot_cy, d_main),
        categories=m_cats,
        values=m_vals,
        colors=m_cols,
        pie_diameter=d_main,
        area_box=area_box,
        label_indices=list(main_indices),
    )
    detail_view = dataclass_replace(
        chart,
        anchor=_virtual_anchor(detail_cx, plot_cy, d_second),
        categories=d_cats,
        values=d_vals,
        colors=d_cols,
        pie_diameter=d_second,
        area_box=area_box,
        label_indices=list(detail_indices),
    )
    return main_view, detail_view


def chart_group_bounds(charts: Sequence[ChartInfo]) -> Optional[Tuple[float, float, float, float]]:
    # 复合饼图（单个 ofPieChart 拆出的两个虚拟视图）应以原始图表区域为参考坐标系，而不是
    # 由虚拟饼图外框合并而成，以便与细则中“图表区域内 25%–35% / 72%–82%”的百分比一致。
    for c in charts:
        if c is not None and c.area_box is not None:
            return c.area_box
    anchors = [c.anchor for c in charts if c.anchor is not None]
    if not anchors:
        return None
    left = min(a.left for a in anchors)
    top = min(a.top for a in anchors)
    right = max(a.right for a in anchors)
    bottom = max(a.bottom for a in anchors)
    return left, top, right, bottom


def normalized_center(anchor: Anchor, bounds: Tuple[float, float, float, float]) -> Tuple[float, float]:
    left, top, right, bottom = bounds
    return (anchor.cx - left) / max(right - left, 1), (anchor.cy - top) / max(bottom - top, 1)


def data_matches(chart: Optional[ChartInfo], expected: Dict[str, float], tol: float = 0.015) -> Tuple[bool, str]:
    if chart is None:
        return False, "未找到对应图表"
    m = chart.data_map()
    missing = [k for k in expected if k not in m]
    wrong = [f"{k}={m.get(k)}≠{v}" for k, v in expected.items() if k in m and not almost_equal(m[k], v, tol)]
    if missing or wrong:
        return False, f"缺失 {missing}；比例不符 {wrong}"
    return True, "分类和值均匹配"


def read_sheet_category_values(
    wb_values: Any,
    labels: Sequence[str],
    preferred_sheets: Sequence[str] = ("Sheet1",),
) -> Dict[str, float]:
    """从工作表单元格中读取指定分类标签对应的数值。

    细则要求扇区面积比例与“sheet1表主分类数据”“源表其他明细数据”一致，因此这里
    从源单元格读取真实数值（取标签右侧或下方最近的数字单元格），而非使用写死的期望值。
    """
    result: Dict[str, float] = {}
    if wb_values is None:
        return result
    sheet_order = list(preferred_sheets) + [s for s in wb_values.sheetnames if s not in preferred_sheets]
    for label in labels:
        for sheet_name in sheet_order:
            if sheet_name not in wb_values.sheetnames:
                continue
            ws = wb_values[sheet_name]
            found: Optional[float] = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value != label:
                        continue
                    for offset in (1, 2):
                        ncol = cell.column + offset
                        if ncol <= ws.max_column:
                            num = safe_float(ws.cell(cell.row, ncol).value)
                            if num is not None:
                                found = num
                                break
                    if found is None:
                        nrow = cell.row + 1
                        if nrow <= ws.max_row:
                            num = safe_float(ws.cell(nrow, cell.column).value)
                            if num is not None:
                                found = num
                    if found is not None:
                        break
                if found is not None:
                    break
            if found is not None:
                result[label] = found
                break
    return result


def to_proportions(values: Dict[str, float]) -> Dict[str, float]:
    total = sum(v for v in values.values() if v is not None)
    if total <= 0:
        return {}
    return {k: v / total for k, v in values.items()}


def sector_ratio_matches_source(
    chart: Optional[ChartInfo],
    labels: Sequence[str],
    wb_values: Any,
    tol: float = 0.02,
) -> Tuple[bool, str]:
    """判断图表三个扇区的面积比例是否与源表对应数据的比例一致。

    饼图扇区面积占比 = 该扇区值 / 各扇区值之和，因此归一化后与源数据归一化比例比较。
    """
    if chart is None:
        return False, "图表不存在"
    source = read_sheet_category_values(wb_values, labels)
    missing_source = [l for l in labels if l not in source]
    if missing_source:
        return False, f"源表缺少分类数据：{missing_source}"
    chart_map = chart.data_map()
    missing_chart = [l for l in labels if l not in chart_map]
    if missing_chart:
        return False, f"图表缺少扇区：{missing_chart}"
    src_prop = to_proportions({l: source[l] for l in labels})
    chart_prop = to_proportions({l: chart_map[l] for l in labels})
    if not src_prop or not chart_prop:
        return False, "源数据或图表数据合计为 0，无法计算比例"
    wrong = [
        f"{l}:图{chart_prop[l]:.1%}≠源{src_prop[l]:.1%}"
        for l in labels
        if abs(chart_prop[l] - src_prop[l]) > tol
    ]
    if wrong:
        return False, f"面积比例与源数据不一致：{wrong}"
    return True, f"面积比例与源数据一致（源比例={ {l: round(src_prop[l], 3) for l in labels} }）"


def dashed_connector_between(
    model: WorkbookModel,
    main: Optional[ChartInfo],
    detail: Optional[ChartInfo],
) -> Tuple[bool, str]:
    """判断两个饼图之间是否存在虚线连接（细则仅要求“用虚线连接”，不约束颜色/条数/线宽）。"""
    # 单个 ofPieChart 的连接线由图表自身的 c:serLines 表达，不作为独立形状出现在 drawing 里。
    for chart in (main, detail):
        if chart is not None and chart.of_pie and chart.of_pie.get("has_ser_lines"):
            dash = chart.of_pie.get("ser_line_dash")
            is_dashed = (dash or "solid") != "solid"
            return True, f"ofPieChart 自带连接线(serLines)，线型={dash or '默认'}，{'虚线' if is_dashed else '按默认连接线处理'}"
    if not (main and detail and main.anchor and detail.anchor):
        return False, "缺少图表锚点，无法判断连接线位置"
    left_x = min(main.anchor.cx, detail.anchor.cx)
    right_x = max(main.anchor.cx, detail.anchor.cx)
    count = 0
    for line in model.lines:
        mid_x = line.left + line.width / 2
        is_dashed = (line.dash or "solid") != "solid"
        between = left_x <= mid_x <= right_x
        if is_dashed and between:
            count += 1
    return count >= 1, f"两饼图之间的虚线数量={count}"


def check_core_layout_and_data(
    model: WorkbookModel,
    main: Optional[ChartInfo],
    detail: Optional[ChartInfo],
    bounds: Optional[Tuple[float, float, float, float]],
    wb_values: Any,
) -> Hit:
    reasons = []
    main_sectors = list(EXPECTED_MAIN)   # 人员薪酬、物料、其他
    detail_sectors = list(EXPECTED_DETAIL)  # 交通、住宿、餐饮

    anchors_ok = bool(main and detail and main.anchor and detail.anchor and bounds)
    ok_main_pos = ok_detail_pos = ok_left_right = False
    if main and detail and main.anchor and detail.anchor and bounds:
        mcx, mcy = normalized_center(main.anchor, bounds)
        dcx, dcy = normalized_center(detail.anchor, bounds)
        ok_left_right = main.anchor.cx < detail.anchor.cx
        ok_main_pos = 0.25 <= mcx <= 0.35 and 0.45 <= mcy <= 0.55
        ok_detail_pos = 0.72 <= dcx <= 0.82 and 0.42 <= dcy <= 0.55
        reasons.append(f"主饼图位于左侧：{'通过' if ok_left_right else '不通过'}")
        reasons.append(f"主饼图中心=({mcx:.2%},{mcy:.2%})，要求宽25%-35%、高45%-55%：{'通过' if ok_main_pos else '不通过'}")
        reasons.append(f"拆分饼图中心=({dcx:.2%},{dcy:.2%})，要求宽72%-82%、高42%-55%：{'通过' if ok_detail_pos else '不通过'}")
    else:
        reasons.append("未能同时识别左右两个饼图及其锚点")

    main_has_sectors = bool(main) and set(main_sectors).issubset(set(main.categories))
    detail_has_sectors = bool(detail) and set(detail_sectors).issubset(set(detail.categories))
    reasons.append(f"主饼图包含三扇区{main_sectors}：{'通过' if main_has_sectors else '不通过'}")
    reasons.append(f"拆分饼图包含三扇区{detail_sectors}：{'通过' if detail_has_sectors else '不通过'}")

    dm, dm_reason = sector_ratio_matches_source(main, main_sectors, wb_values)
    dd, dd_reason = sector_ratio_matches_source(detail, detail_sectors, wb_values)
    reasons.append(f"主饼图面积比例与Sheet1主分类一致：{dm_reason}")
    reasons.append(f"拆分饼图面积比例与源表其他明细一致：{dd_reason}")

    connectors, conn_reason = dashed_connector_between(model, main, detail)
    reasons.append(f"两饼图之间虚线连接：{conn_reason}")

    ok = (
        anchors_ok
        and ok_left_right
        and ok_main_pos
        and ok_detail_pos
        and main_has_sectors
        and detail_has_sectors
        and dm
        and dd
        and connectors
    )
    return Hit(5, "+5 核心布局、数据比例与虚线连接", ok, "；".join(reasons))


def labels_have_category_and_percent(
    chart: Optional[ChartInfo],
    expected: Dict[str, float],
    *,
    allow_context_text: bool = False,
    allow_source_percent_format: bool = False,
) -> Tuple[bool, str]:
    if chart is None:
        return False, "图表不存在"
    dlabels_node = chart.xml.find(".//c:dLbls", NS)
    label_text = extract_text(dlabels_node)
    context_text = f"{chart.title} {' '.join(chart.categories)}" if allow_context_text else ""
    flags = chart.dlabel_flags

    # 主饼图标签得分点要求数据标签自身包含分类名，因此默认只看 dLbls/showCatName。
    # 右侧拆分饼图的扇区名常放在标题/分类缓存中，而标签只显示百分比；对该项允许用上下文文本证明分类名存在。
    category_visible = flags.get("showCatName", False) or bool(
        allow_context_text and all(name in context_text for name in expected)
    )
    percent_visible = flags.get("showPercent", False) or (
        flags.get("showVal", False)
        and (chart_has_percent_format(chart) or (allow_source_percent_format and chart_values_have_percent_format(chart)))
    )

    text_for_category = f"{label_text} {context_text}" if allow_context_text else label_text
    missing_text = []
    if not category_visible:
        missing_text = [name for name in expected if name not in text_for_category]

    text_for_percent = label_text
    missing_percent = []
    for _name, val in expected.items():
        percent_text = f"{round(val * 100):.0f}%"
        if percent_text not in text_for_percent and not percent_visible:
            missing_percent.append(percent_text)

    if missing_text or missing_percent or not percent_visible or not category_visible:
        return False, f"缺失标签文本={missing_text}，缺失标签百分比={missing_percent}，标签配置={flags}"
    source_note = "，使用源数据百分比格式" if allow_source_percent_format and chart_values_have_percent_format(chart) else ""
    return True, f"分类名称与百分比可显示{source_note}"


def chart_has_percent_format(chart: ChartInfo) -> bool:
    # 只检查数据标签自身的数字格式。c:val/c:numCache/c:formatCode 只是源数据缓存格式，
    # 不能证明数据标签实际按百分比显示。
    for fmt in chart.xml.findall(".//c:dLbls/c:numFmt", NS) + chart.xml.findall(".//c:dLbl/c:numFmt", NS):
        format_code = fmt.attrib.get("formatCode", "")
        if "%" in format_code:
            return True
    return False


def chart_values_have_percent_format(chart: ChartInfo) -> bool:
    for fmt in chart.xml.findall(".//c:val//c:formatCode", NS):
        if fmt.text and "%" in fmt.text:
            return True
    return False


def font_ok(chart: Optional[ChartInfo], min_size: float, max_size: float) -> Tuple[bool, str]:
    if chart is None:
        return False, "图表不存在"
    faces = chart.font_info.get("typefaces", [])
    sizes = chart.font_info.get("sizes_pt", [])
    colors = chart.font_info.get("colors", [])
    # 未显式写入字体时，Excel 图表常用 Calibri/黑色；按默认可接受。字号未写入则用 12pt 近似，避免把默认主题误判为人工不可评估。
    face_ok = not faces or any(face in ALLOWED_FONTS for face in faces)
    size_values = sizes or [12.0]
    size_ok = any(min_size <= s <= max_size for s in size_values)
    color_ok = not colors or any(is_black_or_dark_gray(c) for c in colors)
    return face_ok and size_ok and color_ok, f"字体={faces or ['默认Calibri']}，字号={sizes or ['默认约12']}，颜色={colors or ['默认黑色']}"


def _series_dlbls_nodes(chart: ChartInfo) -> List[ET.Element]:
    """图表 XML 中所有的 c:dLbls 容器（图表级 + series 级），用于回退查找。"""
    return chart.xml.findall(".//c:dLbls", NS)


def _resolve_dlbl_override(chart: ChartInfo, orig_idx: int) -> Optional[ET.Element]:
    """在原始 series 的 c:dLbls 里查找指定数据点索引对应的 c:dLbl 覆盖节点。"""
    for dlbls in _series_dlbls_nodes(chart):
        for dlbl in dlbls.findall("c:dLbl", NS):
            idx_node = dlbl.find("c:idx", NS)
            if idx_node is None:
                continue
            try:
                if int(idx_node.attrib.get("val", "-1")) == orig_idx:
                    return dlbl
            except ValueError:
                continue
    return None


def _label_flags_effective(chart: ChartInfo, dlbl_node: Optional[ET.Element]) -> Dict[str, bool]:
    """合并本 c:dLbl 与 series 级 c:dLbls 的显示旗标（本级 dLbl 覆盖 series 级）。"""
    flags = dict(chart.dlabel_flags)
    if dlbl_node is not None:
        for key in ("showVal", "showCatName", "showSerName", "showPercent", "showLegendKey"):
            child = dlbl_node.find(f"c:{key}", NS)
            if child is not None:
                flags[key] = child.attrib.get("val", "1") not in {"0", "false", "False"}
    return flags


def _label_num_format_is_percent(chart: ChartInfo, dlbl_node: Optional[ET.Element]) -> bool:
    """本数据标签是否显式使用百分比数字格式（先看 dLbl 内 numFmt，再看 series 级 dLbls/numFmt）。"""
    if dlbl_node is not None:
        fmt = dlbl_node.find("c:numFmt", NS)
        if fmt is not None:
            return "%" in fmt.attrib.get("formatCode", "")
    for dlbls in _series_dlbls_nodes(chart):
        fmt = dlbls.find("c:numFmt", NS)
        if fmt is not None:
            return "%" in fmt.attrib.get("formatCode", "")
    return False


def _label_rich_text(dlbl_node: Optional[ET.Element]) -> str:
    """如果该 c:dLbl 显式写死了 c:tx/c:rich 文本，返回其文本内容，否则空字符串。"""
    if dlbl_node is None:
        return ""
    tx = dlbl_node.find("c:tx", NS)
    if tx is None:
        return ""
    rich = tx.find("c:rich", NS)
    if rich is None:
        return ""
    return extract_text(rich)


def _collect_font_from(txpr: Optional[ET.Element]) -> Dict[str, List[Any]]:
    result: Dict[str, List[Any]] = {"typefaces": [], "sizes_pt": [], "colors": []}
    if txpr is None:
        return result
    for latin in txpr.findall(".//a:latin", NS):
        face = latin.attrib.get("typeface")
        if face:
            result["typefaces"].append(face)
    for ea in txpr.findall(".//a:ea", NS):
        face = ea.attrib.get("typeface")
        if face:
            result["typefaces"].append(face)
    for rpr in txpr.findall(".//a:rPr", NS) + txpr.findall(".//a:defRPr", NS) + txpr.findall(".//a:endParaRPr", NS):
        if "sz" in rpr.attrib:
            try:
                result["sizes_pt"].append(int(rpr.attrib["sz"]) / 100)
            except ValueError:
                pass
        result["colors"].extend(collect_srgb_colors(rpr))
    return result


def _label_font_info(chart: ChartInfo, dlbl_node: Optional[ET.Element]) -> Dict[str, List[Any]]:
    """获取该数据标签生效的字体名/字号/颜色。

    读取优先级：本 c:dLbl 的 c:txPr → 本 c:dLbl 的 c:tx/c:rich（写死富文本自带字体）→
    series 级 c:dLbls 的 c:txPr。不再从整张图任意文本节点取值，避免把标题/图例字体误当做数据标签字体。
    """
    if dlbl_node is not None:
        for tag in ("c:txPr", "c:tx"):
            for txpr in dlbl_node.findall(tag, NS):
                info = _collect_font_from(txpr)
                if info["typefaces"] or info["sizes_pt"] or info["colors"]:
                    return info
    for dlbls in _series_dlbls_nodes(chart):
        for txpr in dlbls.findall("c:txPr", NS):
            info = _collect_font_from(txpr)
            if info["typefaces"] or info["sizes_pt"] or info["colors"]:
                return info
    return {"typefaces": [], "sizes_pt": [], "colors": []}


def _label_manual_layout(dlbl_node: Optional[ET.Element]) -> Optional[Dict[str, Any]]:
    if dlbl_node is None:
        return None
    layout = dlbl_node.find("c:layout", NS)
    if layout is None:
        return None
    ml = layout.find("c:manualLayout", NS)
    if ml is None:
        return None

    def _f(name: str, default: float = 0.0) -> float:
        node = ml.find(f"c:{name}", NS)
        if node is None:
            return default
        try:
            return float(node.attrib.get("val", str(default)))
        except ValueError:
            return default

    def _mode(name: str) -> str:
        node = ml.find(f"c:{name}", NS)
        return node.attrib.get("val", "factor") if node is not None else "factor"

    return {
        "x": _f("x"), "y": _f("y"), "w": _f("w"), "h": _f("h"),
        "xMode": _mode("xMode"), "yMode": _mode("yMode"),
    }


def _label_in_chart_area(layout: Optional[Dict[str, Any]]) -> Tuple[bool, str]:
    """判断显式布局的数据标签是否仍留在图表区域内。

    - 未写 manualLayout：办公软件自动布局默认把数据标签保持在图表区域内，视为通过。
    - edge 模式：x/y 是相对图表区(或绘图区)的边距分数，必须 0<=x<=1、且 x+w<=1（y 同理）。
    - factor 模式：x/y 是相对自动位置的偏移分数，|x|、|y| 通常 <0.5；超过 0.5 视为有明显推出图表区的风险。
    """
    if layout is None:
        return True, "自动布局"
    reasons: List[str] = []
    ok = True
    if layout["xMode"] == "edge":
        if not (-0.001 <= layout["x"] <= 1.001):
            ok = False
            reasons.append(f"x={layout['x']:.2f}越界")
        elif layout["w"] and layout["x"] + layout["w"] > 1.001:
            ok = False
            reasons.append(f"x+w={layout['x'] + layout['w']:.2f}>1")
    elif abs(layout["x"]) > 0.5:
        ok = False
        reasons.append(f"factor偏移x={layout['x']:.2f}过大")
    if layout["yMode"] == "edge":
        if not (-0.001 <= layout["y"] <= 1.001):
            ok = False
            reasons.append(f"y={layout['y']:.2f}越界")
        elif layout["h"] and layout["y"] + layout["h"] > 1.001:
            ok = False
            reasons.append(f"y+h={layout['y'] + layout['h']:.2f}>1")
    elif abs(layout["y"]) > 0.5:
        ok = False
        reasons.append(f"factor偏移y={layout['y']:.2f}过大")
    return ok, "手动布局越界(" + ",".join(reasons) + ")" if not ok else "手动布局在图表区域内"


def _view_percent(values: Sequence[float], view_idx: int) -> Optional[float]:
    """按“本视图 total”计算该扇区显示百分比。

    办公软件里数据标签的 showPercent 对饼图/复合饼图各半的算法是：该切片值 / 该饼图内所有可见切片值之和。
    虚拟主/拆分视图的 values 已经是该饼图内的可见切片，直接以视图总和为基数即可。
    """
    total = sum(v for v in values if v is not None)
    if total <= 0 or not (0 <= view_idx < len(values)):
        return None
    return values[view_idx] / total * 100.0


def _pct_close(actual: Optional[float], target: int, tol_pp: float = 1.0) -> bool:
    """百分比匹配阈值：允许 ±1 个百分点，兼容 Excel 四舍五入到整数百分比后的抖动。"""
    if actual is None:
        return False
    return abs(actual - target) <= tol_pp


def _label_shows_category(text: str, name: str, flags: Dict[str, bool]) -> bool:
    return flags.get("showCatName", False) or (name in text)


def _label_shows_target_percent(
    text: str,
    target: int,
    flags: Dict[str, bool],
    percent_fmt: bool,
    actual_pct: Optional[float],
) -> Tuple[bool, str]:
    """判定该扇区数据标签是否显示了目标百分比。

    规则（细则要求“显示的百分比确为 55%/34%/11% 等目标值”）：
      1. 若 dLbl 有富文本（写死），文本包含目标百分比字符串（±1pp 也算匹配）即通过；
      2. 若 showPercent=True，则以 series 的实际值算出的百分比必须落在 target±1pp；
      3. 若 showVal=True 且数字格式为百分比，则按同样的规则用实际值验算。
    仅“开了 showPercent/showVal”而不核对实际数字，会漏掉“把 55% 手改成 88%”这类问题，因此在此处必须验算。
    """
    target_text = f"{target}%"
    # 富文本直接命中：允许 ±1pp（写死 "56%" 也接受，与自动 showPercent 的四舍五入行为一致）。
    if text:
        # 从富文本中抓出百分比数字，取任意一个在 target±1pp 内。
        for m in re.finditer(r"(-?\d+(?:\.\d+)?)\s*%", text):
            try:
                if abs(float(m.group(1)) - target) <= 1.0:
                    return True, f"标签富文本含 {m.group(0)}"
            except ValueError:
                continue
        if target_text in text:
            return True, f"标签富文本含 {target_text}"

    if flags.get("showPercent", False):
        if _pct_close(actual_pct, target):
            return True, f"showPercent=True，实际值={actual_pct:.1f}% 与 {target_text} 一致"
        return False, f"showPercent=True 但实际值={actual_pct if actual_pct is not None else '未知'} ≠ {target_text}"

    if flags.get("showVal", False) and percent_fmt:
        if _pct_close(actual_pct, target):
            return True, f"showVal 数字格式为百分比，实际值={actual_pct:.1f}% 与 {target_text} 一致"
        return False, f"showVal% 但实际值={actual_pct if actual_pct is not None else '未知'} ≠ {target_text}"

    if text:
        return False, f"标签富文本={text!r}，未见 {target_text}"
    return False, f"未启用 showPercent，也未在标签富文本中写明 {target_text}"


def main_pie_font_ok(chart: Optional[ChartInfo]) -> Tuple[bool, str]:
    """左侧主饼图“数据标签自身”的字体：微软雅黑/宋体/Calibri，字号 12-16 磅，颜色黑色。

    严格只读取每个 c:dLbl 的字体（或 series 级 c:dLbls/c:txPr），不从整张图任意文本节点取值。
    未显式写入时按办公软件默认（Calibri / 约 12 磅 / 黑色）处理。
    """
    if chart is None:
        return False, "图表不存在"
    faces_all: List[str] = []
    sizes_all: List[float] = []
    colors_all: List[str] = []
    for view_idx, _name in enumerate(chart.categories):
        orig_idx = chart.label_indices[view_idx] if chart.label_indices else view_idx
        dlbl = _resolve_dlbl_override(chart, orig_idx)
        info = _label_font_info(chart, dlbl)
        faces_all.extend(info["typefaces"])
        sizes_all.extend(info["sizes_pt"])
        colors_all.extend(info["colors"])

    face_ok = not faces_all or all(face in ALLOWED_FONTS for face in faces_all)
    size_values = sizes_all or [12.0]
    size_ok = all(12 <= s <= 16 for s in size_values)
    color_ok = not colors_all or all(is_black_or_dark_gray(c) for c in colors_all)
    ok = face_ok and size_ok and color_ok
    detail = (
        f"字体={faces_all or ['默认Calibri']}(要求微软雅黑/宋体/Calibri:{'通过' if face_ok else '不通过'})，"
        f"字号={sizes_all or ['默认约12']}(要求12-16磅:{'通过' if size_ok else '不通过'})，"
        f"颜色={colors_all or ['默认黑色']}(要求黑色:{'通过' if color_ok else '不通过'})"
    )
    return ok, detail


def main_pie_labels_ok(chart: Optional[ChartInfo]) -> Tuple[bool, str]:
    """左侧主饼图数据标签：三个扇区各自逐一绑定分类名与目标百分比。

    与旧实现不同的关键点：
      1. 按分类名在图表分类缓存里的位置找到数据点索引，再定位到 c:dLbl 覆盖节点；
      2. 判断“百分比可见”时不再只看 showPercent 旗标，而是把 series 实际值除以本视图总和验算，
         结果必须与细则期望值（55/34/11%）在 ±1pp 内一致；
      3. c:dLbl 内写死的 c:tx/c:rich 文本会被视为最终显示文本，同样按数字匹配。
    """
    if chart is None:
        return False, "图表不存在"
    sectors: Dict[str, int] = {"人员薪酬": 55, "物料": 34, "其他": 11}
    categories = chart.categories
    reasons: List[str] = []
    all_ok = True
    for name, target in sectors.items():
        if name not in categories:
            reasons.append(f"缺少扇区{name}")
            all_ok = False
            continue
        view_idx = categories.index(name)
        orig_idx = chart.label_indices[view_idx] if chart.label_indices else view_idx
        dlbl = _resolve_dlbl_override(chart, orig_idx)
        flags = _label_flags_effective(chart, dlbl)
        percent_fmt = _label_num_format_is_percent(chart, dlbl)
        text = _label_rich_text(dlbl)
        actual_pct = _view_percent(chart.values, view_idx)
        cat_visible = _label_shows_category(text, name, flags)
        pct_ok, pct_reason = _label_shows_target_percent(text, target, flags, percent_fmt, actual_pct)
        if not cat_visible:
            all_ok = False
            reasons.append(f"{name}: 分类名不可见(flags.showCatName={flags.get('showCatName')} 富文本={text!r})")
            continue
        if not pct_ok:
            all_ok = False
            reasons.append(f"{name}: {pct_reason}")
            continue
        reasons.append(f"{name}✓ {pct_reason}")
    return all_ok, "；".join(reasons)


def check_main_labels(main: Optional[ChartInfo]) -> Hit:
    if main is None:
        return Hit(1, "+1 左侧主饼图标签文本、百分比、字体与区域", False, "图表不存在")
    label_ok, label_reason = main_pie_labels_ok(main)
    font_pass, font_reason = main_pie_font_ok(main)
    # 逐标签检查显式手动布局是否越出图表区域；未显式设定即认为自动布局在区域内。
    area_reasons: List[str] = []
    area_ok = True
    for view_idx, name in enumerate(main.categories):
        orig_idx = main.label_indices[view_idx] if main.label_indices else view_idx
        dlbl = _resolve_dlbl_override(main, orig_idx)
        layout = _label_manual_layout(dlbl)
        sub_ok, sub_reason = _label_in_chart_area(layout)
        area_reasons.append(f"{name}={sub_reason}")
        if not sub_ok:
            area_ok = False
    ok = label_ok and font_pass and area_ok
    return Hit(
        1,
        "+1 左侧主饼图标签文本、百分比、字体与区域",
        ok,
        f"{label_reason}；{font_reason}；标签区域检查：{'/'.join(area_reasons)}",
    )


def detail_pie_font_ok(chart: Optional[ChartInfo]) -> Tuple[bool, str]:
    """右侧拆分饼图“数据标签自身”的字体：微软雅黑/宋体/Calibri，字号 10-14 磅，颜色黑色。

    与主饼图一致，严格只读取每个 c:dLbl 的字体（或 series 级 c:dLbls/c:txPr），
    不再从整张图任意文本节点（标题/图例等）取值。未显式写入时按办公软件默认
    （Calibri / 约 12 磅 / 黑色）处理。
    """
    if chart is None:
        return False, "图表不存在"
    faces_all: List[str] = []
    sizes_all: List[float] = []
    colors_all: List[str] = []
    for view_idx, _name in enumerate(chart.categories):
        orig_idx = chart.label_indices[view_idx] if chart.label_indices else view_idx
        dlbl = _resolve_dlbl_override(chart, orig_idx)
        info = _label_font_info(chart, dlbl)
        faces_all.extend(info["typefaces"])
        sizes_all.extend(info["sizes_pt"])
        colors_all.extend(info["colors"])

    face_ok = not faces_all or all(face in ALLOWED_FONTS for face in faces_all)
    size_values = sizes_all or [12.0]
    size_ok = all(10 <= s <= 14 for s in size_values)
    color_ok = not colors_all or all(is_black_or_dark_gray(c) for c in colors_all)
    ok = face_ok and size_ok and color_ok
    detail = (
        f"字体={faces_all or ['默认Calibri']}(要求微软雅黑/宋体/Calibri:{'通过' if face_ok else '不通过'})，"
        f"字号={sizes_all or ['默认约12']}(要求10-14磅:{'通过' if size_ok else '不通过'})，"
        f"颜色={colors_all or ['默认黑色']}(要求黑色:{'通过' if color_ok else '不通过'})"
    )
    return ok, detail


def detail_pie_labels_ok(chart: Optional[ChartInfo]) -> Tuple[bool, str]:
    """右侧拆分饼图数据标签：交通/住宿/餐饮逐扇区绑定分类名与目标百分比。

    与主饼图检查同源：
      1. 按分类名→数据点索引→c:dLbl 覆盖节点逐一定位；
      2. “百分比可见”不再只看 showPercent 旗标，而是用 series 实际值除以本视图总和验算，
         结果须与细则期望值（交通15/住宿65/餐饮20%）在 ±1pp 内一致；
      3. c:dLbl 内写死的 c:tx/c:rich 富文本按数字匹配。
    """
    if chart is None:
        return False, "图表不存在"
    sectors: Dict[str, int] = {"交通": 15, "住宿": 65, "餐饮": 20}
    categories = chart.categories
    reasons: List[str] = []
    all_ok = True
    for name, target in sectors.items():
        if name not in categories:
            reasons.append(f"缺少扇区{name}")
            all_ok = False
            continue
        view_idx = categories.index(name)
        orig_idx = chart.label_indices[view_idx] if chart.label_indices else view_idx
        dlbl = _resolve_dlbl_override(chart, orig_idx)
        flags = _label_flags_effective(chart, dlbl)
        percent_fmt = _label_num_format_is_percent(chart, dlbl)
        text = _label_rich_text(dlbl)
        actual_pct = _view_percent(chart.values, view_idx)
        cat_visible = _label_shows_category(text, name, flags)
        pct_ok, pct_reason = _label_shows_target_percent(text, target, flags, percent_fmt, actual_pct)
        if not cat_visible:
            all_ok = False
            reasons.append(f"{name}: 分类名不可见(flags.showCatName={flags.get('showCatName')} 富文本={text!r})")
            continue
        if not pct_ok:
            all_ok = False
            reasons.append(f"{name}: {pct_reason}")
            continue
        reasons.append(f"{name}✓ {pct_reason}")
    return all_ok, "；".join(reasons)


def check_detail_labels(detail: Optional[ChartInfo]) -> Hit:
    label_ok, label_reason = detail_pie_labels_ok(detail)
    font_pass, font_reason = detail_pie_font_ok(detail)
    ok = label_ok and font_pass
    return Hit(1, "+1 右侧拆分饼图标签文本、百分比与字体", ok, f"{label_reason}；{font_reason}")


def is_light_or_medium_saturation(hex_color: str) -> bool:
    """判断填充色是否为“浅色或中等饱和度”（非接近黑色、非极暗的深色）。

    办公软件里扇区填充色取自 dPt/spPr 的 srgbClr。细则要求浅色或中等饱和度，
    因此排除接近黑色/过暗的颜色；亮度中等及以上即视为满足。
    """
    rgb = hex_to_rgb(hex_color)
    if rgb is None:
        return False
    import colorsys

    r, g, b = rgb
    _h, l, _s = colorsys.rgb_to_hls(r / 255, g / 255, b / 255)
    return l >= 0.30


SPECIFIED_COLOR_FAMILIES = {"blue", "orange", "teal", "purple", "light_green"}


def check_colors_and_separators(main: Optional[ChartInfo], detail: Optional[ChartInfo]) -> Hit:
    if not main or not detail:
        return Hit(1, "+1 扇区配色与分隔线", False, "缺少主饼图或拆分饼图")

    main_colors = [c for c in main.colors if c]
    detail_colors = [c for c in detail.colors if c]

    # 点1：左侧主饼图三个扇区使用互相区分的浅色或中等饱和度填充色。
    three_present = len(main_colors) >= 3
    three_distinct = len(set(main_colors[:3])) >= 3
    three_light_medium = three_present and all(is_light_or_medium_saturation(c) for c in main_colors[:3])
    distinct_ok = three_present and three_distinct and three_light_medium

    # 点2：至少包含 蓝色、橙色、青绿色、紫色、浅绿色 中的三种不同颜色。
    main_families = {color_family(c) for c in main_colors}
    main_specified = main_families & SPECIFIED_COLOR_FAMILIES
    specified_ok = len(main_specified) >= 3

    # 点3：右侧拆分饼图扇区颜色与左侧主饼图风格一致（同属指定色系或与左图相同的色族）。
    detail_families = {color_family(c) for c in detail_colors}
    style_ok = bool(detail_colors) and all(
        color_family(c) in (main_families | SPECIFIED_COLOR_FAMILIES) for c in detail_colors
    )

    # 点4：扇区之间有白色或浅灰色分隔线，线宽 0.75–1.5 磅。
    sep_ok = separator_ok(main) and separator_ok(detail)

    ok = distinct_ok and specified_ok and style_ok and sep_ok
    return Hit(
        1,
        "+1 扇区配色与分隔线",
        ok,
        f"主饼图颜色={main_colors}（三扇区互相区分:{'是' if three_distinct else '否'}，浅色/中饱和:{'是' if three_light_medium else '否'}），"
        f"指定色系={sorted(main_specified)}（{len(main_specified)}/3，要求≥3:{'通过' if specified_ok else '不通过'}），"
        f"右饼图颜色={detail_colors}（{sorted(detail_families)}，风格一致:{'通过' if style_ok else '不通过'}），"
        f"分隔线={'白/浅灰且0.75-1.5磅' if sep_ok else '未检测到白/浅灰0.75-1.5磅分隔线'}",
    )


def separator_ok(chart: ChartInfo) -> bool:
    if not chart.separator_lines:
        # 细则明确要求白色/浅灰色分隔线和 0.75–1.5 磅线宽；未显式写入时不能证明命中。
        return False
    for color, width, _dash in chart.separator_lines:
        color_ok = color is not None and (color.upper() in {"FFFFFF", "F2F2F2", "D9D9D9"} or is_gray_or_light_gray(color))
        width_ok = width is not None and 0.75 <= width <= 1.5
        if color_ok and width_ok:
            return True
    return False


def is_gray_or_light_gray_line(hex_color: Optional[str]) -> bool:
    """连接线颜色是否为灰色或浅灰色（细则要求）。

    灰色 = 低饱和度（R/G/B 接近），排除接近纯黑与纯白。中等灰到浅灰都接受。
    """
    rgb = hex_to_rgb(hex_color or "")
    if rgb is None:
        return False
    brightness = sum(rgb) / 3
    return max(rgb) - min(rgb) <= 40 and 64 <= brightness <= 235


# 允许的“虚线/点状虚线”线型集合：细则只承认虚线或点状虚线，其他非 solid（如 dashDot、lgDash 等
# 长划线/长划点、solidDash 混合线）都不算符合。
ALLOWED_CONNECTOR_DASH = {"dash", "sysDash", "dot", "sysDot"}


def _dash_is_dashed_or_dotted(dash: Optional[str]) -> bool:
    return (dash or "solid") in ALLOWED_CONNECTOR_DASH


def connector_lines_ok(model: WorkbookModel, main: Optional[ChartInfo], detail: Optional[ChartInfo], bounds: Optional[Tuple[float, float, float, float]]) -> Tuple[bool, str]:
    # 单个 ofPieChart 携带 serLines 时，两条连接线是图表原生特性，无法从 drawing 里再拿到独立线条。
    # 只要 serLines 存在且颜色为灰/浅灰、线型为虚线/点状虚线、线宽 0.75–1.25 磅，就视作达标。
    # ofPieChart 的 serLines 本身就是“其他扇区→拆分饼图”的原生连接线，端点连通性由渲染器保证。
    for chart in (main, detail):
        if chart is not None and chart.of_pie and chart.of_pie.get("has_ser_lines"):
            meta = chart.of_pie
            color = meta.get("ser_line_color")
            dash_raw = meta.get("ser_line_dash")
            dash: Optional[str] = dash_raw if isinstance(dash_raw, str) else None
            wpt = meta.get("ser_line_width_pt")
            color_ok = color is None or is_gray_or_light_gray_line(color)
            dash_ok = dash is None or _dash_is_dashed_or_dotted(dash)
            width_ok = wpt is None or 0.75 <= wpt <= 1.25
            ok = color_ok and dash_ok and width_ok
            return ok, (
                f"ofPieChart serLines：颜色={color or '默认'}({'灰/浅灰' if color_ok else '不合规'})"
                f"，线型={dash or '默认'}(允许dash/sysDash/dot/sysDot:{'通过' if dash_ok else '不通过'})"
                f"，线宽={wpt if wpt is not None else '默认'}pt({'0.75-1.25' if width_ok else '不合规'})"
            )

    if not bounds:
        return False, "无图表边界"
    _left, top, _right, bottom = bounds
    height = bottom - top

    # 连接端点几何：左端点必须靠近主饼图“其他”扇区所在的圆周（主圆心右侧、主圆内），
    # 右端点必须靠近右侧拆分饼图圆周（拆分圆心左侧、拆分圆内）。
    main_anchor = main.anchor if main is not None else None
    detail_anchor = detail.anchor if detail is not None else None
    have_geom = main is not None and detail is not None and main_anchor is not None and detail_anchor is not None
    if have_geom:
        assert main is not None and detail is not None and main_anchor is not None and detail_anchor is not None
        main_cx = main_anchor.cx
        main_cy = main_anchor.cy
        main_r = estimated_visible_pie_diameter(main) / 2
        detail_cx = detail_anchor.cx
        detail_cy = detail_anchor.cy
        detail_r = estimated_visible_pie_diameter(detail) / 2
    else:
        main_cx = main_cy = main_r = detail_cx = detail_cy = detail_r = 0.0
    # 允许 15% 的半径松弛（考虑估算圆与实际饼图差异、连接线端点略在圆外的情况）
    slack = 0.15

    def _endpoint_reasons(line: LineInfo) -> Tuple[bool, str]:
        if not have_geom:
            return True, "缺少两饼图锚点，不做端点检查"
        # 判断左右端点：line.x1<line.x2 时 (x1,y1) 就是左端；否则交换。
        if line.x1 <= line.x2:
            lx, ly, rx, ry = line.x1, line.y1, line.x2, line.y2
        else:
            lx, ly, rx, ry = line.x2, line.y2, line.x1, line.y1
        # 左端点：落在主饼图右半圆附近（其他扇区就位于主饼图右侧半圆的一段）
        left_from_main = (
            main_cx <= lx <= main_cx + main_r * (1 + slack)
            and abs(ly - main_cy) <= main_r * (1 + slack)
            and math.hypot(lx - main_cx, ly - main_cy) <= main_r * (1 + slack)
        )
        # 右端点：落在拆分饼图左半圆附近
        right_to_detail = (
            detail_cx - detail_r * (1 + slack) <= rx <= detail_cx
            and abs(ry - detail_cy) <= detail_r * (1 + slack)
            and math.hypot(rx - detail_cx, ry - detail_cy) <= detail_r * (1 + slack)
        )
        return left_from_main and right_to_detail, (
            f"左端({lx:.0f},{ly:.0f})→主饼图其他扇区圆周:{'是' if left_from_main else '否'}，"
            f"右端({rx:.0f},{ry:.0f})→拆分饼图圆周:{'是' if right_to_detail else '否'}"
        )

    candidates: list[Tuple[float, LineInfo, str]] = []
    rejected: list[str] = []
    for line in model.lines:
        y = (line.mid_y - top) / max(height, 1)
        # 点：灰色或浅灰色
        color_ok = is_gray_or_light_gray_line(line.color)
        # 点：线宽 0.75–1.25 磅
        width_ok = line.width_pt is not None and 0.75 <= line.width_pt <= 1.25
        # 点：线型限定为虚线或点状虚线
        dash_ok = _dash_is_dashed_or_dotted(line.dash)
        endpoints_ok, ep_reason = _endpoint_reasons(line)
        if color_ok and width_ok and dash_ok and endpoints_ok:
            candidates.append((y, line, ep_reason))
        elif color_ok or width_ok or dash_ok:
            rejected.append(
                (
                    f"色={'✓' if color_ok else '✗'}/宽={'✓' if width_ok else '✗'}({line.width_pt})"
                    f"/型={'✓' if dash_ok else '✗'}({line.dash})/端点={'✓' if endpoints_ok else '✗'}({ep_reason})"
                )
            )

    has_top = any(0.15 <= y <= 0.45 for y, _, _ in candidates)
    has_bottom = any(0.55 <= y <= 0.85 for y, _, _ in candidates)
    two_lines = has_top and has_bottom
    reject_note = f"，被淘汰候选={rejected[:3]}" if rejected and not two_lines else ""
    return two_lines, (
        f"符合条件线条数量={len(candidates)}，上连接线(15%-45%)={has_top}，下连接线(55%-85%)={has_bottom}"
        f"{reject_note}"
    )


def check_connector_lines(model: WorkbookModel, main: Optional[ChartInfo], detail: Optional[ChartInfo], bounds: Optional[Tuple[float, float, float, float]]) -> Hit:
    ok, reason = connector_lines_ok(model, main, detail, bounds)
    return Hit(1, "+1 左右饼图两条灰色虚线连接线", ok, reason)


def _read_plot_area_manual_layout(chart: ChartInfo) -> Optional[Dict[str, float]]:
    """读取 c:plotArea/c:layout/c:manualLayout，将其转成相对图表锚点的分数矩形。

    OOXML 里 manualLayout 的 x/y/w/h 是 0-1 之间的分数，xMode/yMode/wMode/hMode 决定
    是相对“图表区(edge)”还是相对“默认自动布局(factor)”。factor 模式下的偏移我们保守只解释
    x/y 分量的边距压缩，w/h 仍按 edge 处理，因为常见工作簿在这里几乎都用 edge。
    """
    if chart is None or chart.xml is None:
        return None
    ml = chart.xml.find(".//c:plotArea/c:layout/c:manualLayout", NS)
    if ml is None:
        return None

    def _mode(name: str) -> str:
        node = ml.find(f"c:{name}", NS)
        return node.attrib.get("val", "factor") if node is not None else "factor"

    def _f(name: str, default: float) -> float:
        node = ml.find(f"c:{name}", NS)
        if node is None:
            return default
        try:
            return float(node.attrib.get("val", str(default)))
        except ValueError:
            return default

    x = _f("x", 0.0); y = _f("y", 0.0); w = _f("w", 1.0); h = _f("h", 1.0)
    x_mode = _mode("xMode"); y_mode = _mode("yMode")
    # factor 模式：以自动布局为基准的偏移分数。这里以“边距 x”的方向近似压缩宽度，不完全等价但足够启发。
    if x_mode != "edge":
        x = max(0.0, min(1.0, 0.1 + x))  # 默认自动布局起点约 10%
    if y_mode != "edge":
        y = max(0.0, min(1.0, 0.1 + y))
    x = max(0.0, min(1.0, x))
    y = max(0.0, min(1.0, y))
    w = max(0.05, min(1.0 - x, w))
    h = max(0.05, min(1.0 - y, h))
    return {"x": x, "y": y, "w": w, "h": h}


def _estimated_plot_area_fraction(chart: ChartInfo) -> Dict[str, float]:
    """当没有显式 manualLayout 时，根据是否有标题/图例、图例位置估算绘图区在图表内的分数矩形。

    办公软件默认布局的经验值：
      - 标题占顶部约 12–18% 高度；
      - 图例在底部约 12% / 右侧约 18% / 左侧约 18% / 顶部约 10%；
      - 绘图区两侧再各留 3% 边距。
    这些数值不追求像素级还原，只用来把外框拆掉“标题/图例/空白”那一圈，得到接近真实饼图直径。
    """
    x = 0.03
    y = 0.03
    w = 0.94
    h = 0.94

    title_node = chart.xml.find(".//c:chart/c:title", NS) if chart.xml is not None else None
    if title_node is not None:
        # 只有 c:autoTitleDeleted val=1 表示标题被明确删除
        deleted = chart.xml.find(".//c:autoTitleDeleted", NS) if chart.xml is not None else None
        if not (deleted is not None and deleted.attrib.get("val", "0") in {"1", "true"}):
            y += 0.15
            h -= 0.15

    legend_pos = chart.legend_pos
    if legend_pos == "b":
        h -= 0.14
    elif legend_pos == "t":
        y += 0.10
        h -= 0.10
    elif legend_pos == "r":
        w -= 0.18
    elif legend_pos == "l":
        x += 0.18
        w -= 0.18
    elif legend_pos == "tr":
        w -= 0.15

    return {"x": x, "y": y, "w": max(0.1, w), "h": max(0.1, h)}


# 饼图圆形本体占绘图区较小边的比例。Excel/WPS 里饼图默认占绘图区 88–95%，取 0.9。
PIE_DIAMETER_TO_PLOT_MIN_SIDE = 0.9


def estimated_visible_pie_diameter(chart: ChartInfo) -> float:
    """估算饼图圆形本体的实际直径（EMU）。

    改用“绘图区较小边 × PIE_DIAMETER_TO_PLOT_MIN_SIDE”，而不是锚点较小边固定 0.5 倍：
      1. 单个 ofPieChart 拆出的虚拟视图已在解析时按 splitPos+绘图区宽度算好直径；
      2. 有显式 c:plotArea/c:layout/c:manualLayout 时按该分数矩形折算；
      3. 否则按标题/图例的存在情况经验估算绘图区分数矩形。
    这样能把外框上的标题、图例、空白边距剔除，随图表实际渲染变化，而非固定 0.5 系数。
    """
    if chart.pie_diameter is not None:
        return chart.pie_diameter
    if not chart.anchor:
        return 0.0
    frac = _read_plot_area_manual_layout(chart) or _estimated_plot_area_fraction(chart)
    plot_w = frac["w"] * chart.anchor.width
    plot_h = frac["h"] * chart.anchor.height
    return min(plot_w, plot_h) * PIE_DIAMETER_TO_PLOT_MIN_SIDE


def _pie_center(chart: ChartInfo) -> Tuple[float, float]:
    """饼图圆心：虚拟视图直接用锚点中心；实际图表用绘图区中心近似（比锚点中心更贴近真实圆心）。"""
    assert chart.anchor is not None
    if chart.pie_diameter is not None:
        return chart.anchor.cx, chart.anchor.cy
    frac = _read_plot_area_manual_layout(chart) or _estimated_plot_area_fraction(chart)
    a = chart.anchor
    cx = a.left + (frac["x"] + frac["w"] / 2) * a.width
    cy = a.top + (frac["y"] + frac["h"] / 2) * a.height
    return cx, cy


def visible_pie_horizontal_gap_cm(main: ChartInfo, detail: ChartInfo) -> float:
    assert main.anchor is not None and detail.anchor is not None
    main_cx, _ = _pie_center(main)
    detail_cx, _ = _pie_center(detail)
    main_d = estimated_visible_pie_diameter(main)
    detail_d = estimated_visible_pie_diameter(detail)
    return (detail_cx - main_cx - (main_d + detail_d) / 2) / EMU_PER_CM


def visible_pie_overlap_ratio(main: ChartInfo, detail: ChartInfo) -> float:
    assert main.anchor is not None and detail.anchor is not None
    main_r = estimated_visible_pie_diameter(main) / 2
    detail_r = estimated_visible_pie_diameter(detail) / 2
    main_cx, main_cy = _pie_center(main)
    detail_cx, detail_cy = _pie_center(detail)
    center_distance = math.hypot(detail_cx - main_cx, detail_cy - main_cy)
    if center_distance >= main_r + detail_r:
        return 0.0
    # 评分只需要判断是否明显重叠；返回相对较小圆直径的近似侵入比例即可。
    return (main_r + detail_r - center_distance) / max(min(main_r, detail_r) * 2, 1.0)


def check_relative_position(main: Optional[ChartInfo], detail: Optional[ChartInfo]) -> Hit:
    if not main or not detail or not main.anchor or not detail.anchor:
        return Hit(1, "+1 左右饼图相对位置、大小和间距", False, "缺少图表锚点")

    # 点1：左侧主饼图直径大于右侧拆分饼图（“饼图宽度”按圆形本体直径衡量，避免受外框空白影响）。
    main_d = estimated_visible_pie_diameter(main)
    detail_d = estimated_visible_pie_diameter(detail)
    width_larger = main_d > detail_d

    # 点2：左侧主饼图直径约为右侧拆分饼图直径的 1.3–1.8 倍。
    ratio = main_d / max(detail_d, 1)
    ratio_ok = 1.3 <= ratio <= 1.8

    # 点3：两饼图之间水平间距约 3cm–7cm。
    gap_cm = visible_pie_horizontal_gap_cm(main, detail)
    gap_ok = 3 <= gap_cm <= 7

    # 点4：互不重叠。
    visual_overlap = visible_pie_overlap_ratio(main, detail)
    not_overlap = visual_overlap == 0

    ok = width_larger and ratio_ok and gap_ok and not_overlap
    return Hit(
        1,
        "+1 左右饼图相对位置、大小和间距",
        ok,
        f"主饼图直径={main_d / EMU_PER_CM:.2f}cm > 拆分饼图直径={detail_d / EMU_PER_CM:.2f}cm:{'通过' if width_larger else '不通过'}，"
        f"直径比={ratio:.2f}(要求1.3-1.8:{'通过' if ratio_ok else '不通过'})，"
        f"水平间距={gap_cm:.2f}cm(要求3-7cm:{'通过' if gap_ok else '不通过'})，"
        f"重叠率={visual_overlap:.2%}(要求互不重叠:{'通过' if not_overlap else '不通过'})",
    )


def legend_overlay(chart: ChartInfo) -> bool:
    """图例是否叠加在绘图区之上（overlay=1 会遮挡饼图/数据标签）。

    办公软件里“显示图例”默认放在图表四周的空白处（overlay=0）；只有勾选“覆盖图表显示图例”
    或手动把图例拖到饼图上时 overlay=1，才可能遮挡数据标签。
    """
    legend = chart.xml.find(".//c:legend", NS)
    if legend is None:
        return False
    overlay = legend.find("c:overlay", NS)
    if overlay is None:
        return False
    return overlay.attrib.get("val", "1") not in {"0", "false", "False"}


def _legend_manual_layout(chart: ChartInfo) -> Optional[Dict[str, Any]]:
    """读取 c:legend/c:layout/c:manualLayout → {x,y,w,h,xMode,yMode} 分数矩形。

    办公软件默认自动放图例（无 manualLayout）时返回 None；只有用户手动拖动图例才会写这段布局。
    """
    if chart.xml is None:
        return None
    ml = chart.xml.find(".//c:legend/c:layout/c:manualLayout", NS)
    if ml is None:
        return None

    def _f(name: str, default: float = 0.0) -> float:
        node = ml.find(f"c:{name}", NS)
        if node is None:
            return default
        try:
            return float(node.attrib.get("val", str(default)))
        except ValueError:
            return default

    def _mode(name: str) -> str:
        node = ml.find(f"c:{name}", NS)
        return node.attrib.get("val", "factor") if node is not None else "factor"

    return {
        "x": _f("x"), "y": _f("y"), "w": _f("w"), "h": _f("h"),
        "xMode": _mode("xMode"), "yMode": _mode("yMode"),
    }


def _deleted_legend_indices(chart: ChartInfo) -> set:
    """收集 c:legend/c:legendEntry 中被 c:delete=1 显式删除的原始数据点索引。

    图例的分类名/颜色都由 series/dPt 自动填充；只有 c:legendEntry 里显式 delete 才会把某个分类
    从图例中隐藏，等价于该分类“图例条目名称/颜色缺失”。
    """
    if chart.xml is None:
        return set()
    result: set = set()
    for entry in chart.xml.findall(".//c:legend/c:legendEntry", NS):
        idx_node = entry.find("c:idx", NS)
        delete = entry.find("c:delete", NS)
        if idx_node is None or delete is None:
            continue
        if delete.attrib.get("val", "1") in {"0", "false", "False"}:
            continue
        try:
            result.add(int(idx_node.attrib.get("val", "-1")))
        except ValueError:
            pass
    return result


def _rects_intersect(a: Dict[str, float], b: Dict[str, float]) -> bool:
    """判断两个 [0,1] 分数矩形（x,y,w,h）是否相交。"""
    if a["w"] <= 0 or a["h"] <= 0 or b["w"] <= 0 or b["h"] <= 0:
        return False
    return not (
        a["x"] + a["w"] <= b["x"] or b["x"] + b["w"] <= a["x"]
        or a["y"] + a["h"] <= b["y"] or b["y"] + b["h"] <= a["y"]
    )


def _legend_overlaps_labels(chart: ChartInfo) -> Tuple[bool, str]:
    """图例手动布局是否与任一数据标签手动布局相交。

    办公软件里图例默认布局的坐标不写入 OOXML，因此这里只能对显式 manualLayout 场景生效：
      - 图例有 c:legend/c:layout/c:manualLayout（edge 模式）；
      - 数据标签有 c:dLbl/c:layout/c:manualLayout（edge 模式）。
    这两者都是相对图表区域的 [0,1] 分数矩形，直接做矩形相交即可判定“手动移动导致的遮挡”。
    factor 模式的偏移无法脱离“默认位置”做几何相交，因此跳过（overlay=1 分支已单独覆盖）。
    """
    lg = _legend_manual_layout(chart)
    if lg is None or lg["xMode"] != "edge" or lg["yMode"] != "edge":
        return False, ""
    if lg["w"] <= 0 or lg["h"] <= 0:
        return False, ""
    hit_labels: List[str] = []
    for lbl in chart.xml.findall(".//c:dLbl", NS):
        ll = _label_manual_layout(lbl)
        if ll is None or ll["xMode"] != "edge" or ll["yMode"] != "edge":
            continue
        # 数据标签 manualLayout 常见只写 x/y 不写 w/h；给一个保守的可视框（15%×8%）避免漏判。
        rect = {"x": ll["x"], "y": ll["y"], "w": ll["w"] or 0.15, "h": ll["h"] or 0.08}
        if _rects_intersect(lg, rect):
            idx_node = lbl.find("c:idx", NS)
            hit_labels.append(idx_node.attrib.get("val", "?") if idx_node is not None else "?")
    if hit_labels:
        return True, f"图例手动布局与数据标签相交(idx={hit_labels})"
    return False, ""


def _legend_entries_match_sectors(chart: ChartInfo, expected_cats: Iterable[str]) -> Tuple[bool, str]:
    """校验期望分类在图例中：图例条目未被删除、且对应扇区取到 srgbClr 填充色。

    图例条目 c:legendEntry 只承载 idx/delete/txPr，色块直接沿用 series/dPt 的 spPr 填充色。
    因此“图例颜色与扇区颜色一致”等价于：
      1) 对应扇区在 spPr 里读到 srgbClr（chart.colors[i] 非空）；
      2) 该 idx 未在 legendEntry 里被 c:delete=1 隐藏。
    ofPieChart 虚拟视图的 categories/colors 已切片，但 legend/legendEntry 仍指向原始 series idx，
    因此优先用 label_indices 把视图内位置映射回原始点索引。
    """
    if chart is None:
        return True, "无图表"
    expected = set(expected_cats)
    if chart.label_indices is not None:
        view_pairs = list(zip(range(len(chart.categories)), chart.label_indices))
    else:
        view_pairs = [(i, i) for i in range(len(chart.categories))]

    deleted_idx = _deleted_legend_indices(chart)
    missing_color: List[str] = []
    hidden_entry: List[str] = []
    for view_i, orig_idx in view_pairs:
        name = chart.categories[view_i] if view_i < len(chart.categories) else ""
        if name not in expected:
            continue
        color = chart.colors[view_i] if view_i < len(chart.colors) else ""
        if not color:
            missing_color.append(name)
        if orig_idx in deleted_idx:
            hidden_entry.append(name)

    reasons: List[str] = []
    if missing_color:
        reasons.append(f"扇区未取到填充色:{missing_color}")
    if hidden_entry:
        reasons.append(f"图例条目被删除:{hidden_entry}")
    if reasons:
        return False, "，".join(reasons)
    return True, "图例条目与扇区色对应齐全"


def check_legend(main: Optional[ChartInfo], detail: Optional[ChartInfo]) -> Hit:
    charts = [c for c in [main, detail] if c]
    if not charts:
        return Hit(1, "+1 图例位置与分类颜色一致", False, "没有图表")

    # 点1：图表没有遮挡饼图的数据标签。
    # overlay=1 属于软件自身覆盖；显式 manualLayout 与数据标签相交则是手动拖动造成的遮挡。
    obscure_reasons: List[str] = []
    seen_targets: set = set()
    for c in charts:
        if c.target in seen_targets:
            continue  # ofPieChart 主/拆分视图共享同一份 XML，避免重复计入。
        seen_targets.add(c.target)
        if legend_overlay(c):
            obscure_reasons.append("图例 overlay=1")
        hit, reason = _legend_overlaps_labels(c)
        if hit:
            obscure_reasons.append(reason)
    not_obscuring = not obscure_reasons

    # 点2：若出现图例，则位于图表底部或右侧空白处；未出现图例（legend_pos 为 None）时该点不适用。
    pos_ok = all(c.legend_pos in {None, "b", "r"} for c in charts)

    # 点3：分类名称与“人员薪酬、物料、其他、交通、住宿、餐饮”一致。
    cats_ok = set(EXPECTED_MAIN).issubset(set(main.categories if main else [])) and set(
        EXPECTED_DETAIL
    ).issubset(set(detail.categories if detail else []))

    # 点4：图例条目对应系列/点颜色与扇区颜色匹配（图例未删除、扇区已从 spPr 读到 srgbClr）。
    color_reasons: List[str] = []
    ok_m, reason_m = _legend_entries_match_sectors(main, EXPECTED_MAIN) if main else (True, "无主饼图")
    if not ok_m:
        color_reasons.append(f"主饼图:{reason_m}")
    ok_d, reason_d = _legend_entries_match_sectors(detail, EXPECTED_DETAIL) if detail else (True, "无拆分饼图")
    if not ok_d:
        color_reasons.append(f"拆分饼图:{reason_d}")
    colors_ok = not color_reasons

    ok = not_obscuring and pos_ok and cats_ok and colors_ok
    return Hit(
        1,
        "+1 图例不遮挡且位置/分类/颜色一致",
        ok,
        f"图例位置={[c.legend_pos for c in charts]}(底部/右侧/无:{'通过' if pos_ok else '不通过'})，"
        f"未遮挡数据标签={'通过' if not_obscuring else '不通过(' + ';'.join(obscure_reasons) + ')'}，"
        f"分类名称齐全={'通过' if cats_ok else '不通过'}，"
        f"图例颜色与扇区一致={'通过' if colors_ok else '不通过(' + ';'.join(color_reasons) + ')'}",
    )


def title_color_ok(chart: Optional[ChartInfo]) -> bool:
    """标题颜色为黑色或深灰色。

    读图表标题文本 run 属性里的 srgbClr；办公软件中未显式设置标题颜色时默认黑色，按默认可接受。
    """
    if chart is None:
        return True
    title_node = chart.xml.find(".//c:chart/c:title", NS)
    if title_node is None:
        return True
    colors: List[str] = []
    for rpr in title_node.findall(".//a:rPr", NS) + title_node.findall(".//a:defRPr", NS):
        colors.extend(collect_srgb_colors(rpr))
    if not colors:
        return True  # 未显式写入颜色，按 Excel 默认黑色处理
    return any(is_black_or_dark_gray(c) for c in colors)


def title_top_center_ok(chart: Optional[ChartInfo]) -> bool:
    """标题位于图表顶部居中。

    办公软件中图表标题默认就在顶部居中；只有存在手动布局(manualLayout)把标题拖走时才可能偏离。
    未设置标题或无手动布局，均视为顶部居中。
    """
    if chart is None:
        return True
    title_node = chart.xml.find(".//c:chart/c:title", NS)
    if title_node is None:
        return True
    manual = title_node.find(".//c:manualLayout", NS)
    if manual is None:
        return True  # 无手动布局，采用默认顶部居中
    # 有手动布局时，若纵向偏移量很小(靠顶部)且未大幅横向偏移，仍视为顶部居中。
    x_node = manual.find("c:x", NS)
    y_node = manual.find("c:y", NS)
    try:
        x = abs(float(x_node.attrib.get("val", "0"))) if x_node is not None else 0.0
        y = abs(float(y_node.attrib.get("val", "0"))) if y_node is not None else 0.0
    except ValueError:
        return True
    return y <= 0.15 and x <= 0.25


def check_title_area(model: WorkbookModel, main: Optional[ChartInfo], detail: Optional[ChartInfo]) -> Hit:
    titles = [c.title for c in [main, detail] if c]
    all_titles = " ".join(titles)

    # 点1：图表顶部没有出现默认标题“图表标题”。
    no_default = "图表标题" not in all_titles

    has_title = bool(all_titles.strip())
    # 以下三点均为“若设置标题”才要求；未设置标题时不适用（视为通过）。
    # 点2：标题含“费用构成”或“复合饼图”含义。
    meaningful = (not has_title) or ("费用构成" in all_titles or "复合饼图" in all_titles)
    # 点3：颜色为黑色或深灰色。
    color_ok = (not has_title) or all(title_color_ok(c) for c in [main, detail] if c and c.title)
    # 点4：位于图表顶部居中。
    pos_ok = (not has_title) or all(title_top_center_ok(c) for c in [main, detail] if c and c.title)

    ok = no_default and meaningful and color_ok and pos_ok
    return Hit(
        1,
        "+1 图表标题区域",
        ok,
        f"标题={titles or ['无标题']}，无默认“图表标题”={'通过' if no_default else '不通过'}，"
        f"含“费用构成/复合饼图”含义={'通过' if meaningful else '不通过'}，"
        f"颜色黑/深灰={'通过' if color_ok else '不通过'}，顶部居中={'通过' if pos_ok else '不通过'}",
    )


def check_penalty_irrelevant_sheets(wb_formula: Any, model: WorkbookModel) -> Hit:
    # 保留占位以兼容旧引用；当前评分链路不再调用该函数。
    extras = [s for s in wb_formula.sheetnames if s not in {"Sheet1", "Sheet2"}]
    return Hit(0, "", False, f"额外工作表={extras or '无'}")


# ----------------------------- 输出 -----------------------------

SCRIPT_ID = "086"


def _strip_score_prefix(name: str) -> str:
    return re.sub(r"^[+\-]\d+\s+", "", name)


def _hit_to_item(hit: Hit) -> dict[str, object]:
    """把内部 Hit 结构映射为统一约定里的 dim2_items 单元。

    - gain 项：score 为正，命中时 delta = score，未命中 delta = 0
    - 扣分项：score 为负（Hit.passed=True 表示扣分被触发），命中时 delta = score，未命中 delta = 0
    """
    delta = hit.score if hit.passed else 0
    # 保留内部 hit.detail 的计算与调试用途，但对外输出的 detail 字段留空。
    return {
        "rule": _strip_score_prefix(hit.name),
        "max_delta": hit.score,
        "delta": delta,
        "hit": hit.passed,
        "detail": "",
    }


def _build_report(
    file_name: str,
    dim1_ok: bool,
    dim1_failed: list[str],
    hits: list[Hit] | None,
    status: str = "ok",
    error: str | None = None,
) -> dict[str, object]:
    dim1_reason = "" if dim1_ok else "；".join(dim1_failed)
    hit_list: list[Hit] = list(hits) if hits else []
    dim2_items = [_hit_to_item(h) for h in hit_list]
    # total = 所有命中项 delta 之和（正向命中加分、扣分命中减分）；维度一未通过时置 0。
    total_score = sum((h.score if h.passed else 0) for h in hit_list) if dim1_ok else 0
    # 满分 = 所有正向得分项 max_delta 之和；扣分项不计入满分。
    max_score = sum(h.score for h in hit_list if h.score > 0)

    return {
        "id": SCRIPT_ID,
        "file_name": file_name,
        "status": status,
        "error": error,
        "dim1_pass": bool(dim1_ok),
        "dim1_reason": dim1_reason,
        "dim2_items": dim2_items,
        "total_score": total_score,
        "max_score": max_score,
    }


def _locate_target_file(dir_path: Path) -> Path | None:
    """在脚本所在目录里定位被评估的 Excel 文档。

    优先使用与本题匹配的固定文件名；若不存在则回退到目录下任一 .xlsx/.xlsm 文件。
    """
    preferred = dir_path / "数据1_复合饼图_Sheet2.xlsx"
    if preferred.exists():
        return preferred
    for pattern in ("*.xlsx", "*.xlsm"):
        for candidate in sorted(dir_path.glob(pattern)):
            # 跳过 Office 打开时可能残留的临时文件
            if candidate.name.startswith("~$"):
                continue
            return candidate
    return None


def evaluate(dir_path: str) -> dict[str, object]:
    """统一入口：接收脚本所在目录路径，返回结构化评分字典。"""
    try:
        dir_p = Path(dir_path)
        if not dir_p.exists() or not dir_p.is_dir():
            return _build_report(
                file_name="",
                dim1_ok=False,
                dim1_failed=[f"目录不存在或不是目录：{dir_path}"],
                hits=None,
                status="error",
                error=f"目录不存在或不是目录：{dir_path}",
            )

        target = _locate_target_file(dir_p)
        if target is None:
            return _build_report(
                file_name="",
                dim1_ok=False,
                dim1_failed=["目录下未找到 .xlsx/.xlsm 文件"],
                hits=None,
                status="error",
                error="目录下未找到 .xlsx/.xlsm 文件",
            )

        file_name = target.name
        model: WorkbookModel | None = None
        wb_formula = None
        wb_values = None
        open_errors: list[str] = []

        try:
            wb_formula = openpyxl.load_workbook(target, data_only=False)
            wb_values = openpyxl.load_workbook(target, data_only=True)
        except Exception as exc:
            open_errors.append(str(exc))

        try:
            if target.exists() and zipfile.is_zipfile(target):
                model = load_model(target)
            else:
                open_errors.append("文件不存在或不是有效的 xlsx/xlsm zip 包")
        except Exception as exc:
            open_errors.append(f"XML解析失败：{exc}")

        dim1_ok, _dim1_passed, dim1_failed = dimension1_checks(target, model, wb_formula, wb_values)
        if open_errors and not dim1_ok:
            dim1_failed.extend(open_errors)

        hits = (
            score_dimension2(model, wb_formula, wb_values)
            if dim1_ok and model and wb_formula and wb_values
            else None
        )
        return _build_report(file_name, dim1_ok, dim1_failed, hits)
    except Exception as exc:  # 兜底：脚本自身异常一律标记为 error
        return _build_report(
            file_name="",
            dim1_ok=False,
            dim1_failed=[f"脚本异常：{exc}"],
            hits=None,
            status="error",
            error=str(exc),
        )


if __name__ == "__main__":
    target_dir = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).parent)
    print(json.dumps(evaluate(target_dir), ensure_ascii=False, indent=2))

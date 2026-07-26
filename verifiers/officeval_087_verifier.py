#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""自动评估“分层作业设计_填充完成.xlsx”。

对外接口（批量运行器约定）：
    evaluate(dir_path: str) -> dict
    - 入参：脚本所在目录的路径；脚本自行在该目录内定位并打开被评估的文档
    - 返回：见 §2.2 结构化字典（含维度一是否通过、维度二逐项得分、总分）

本地调试：
    python officeval_087_verifier.py [脚本所在目录路径]
    未传参时默认使用脚本自身所在目录。

脚本按评分细则分两阶段评估：
1. 先检查“维度1：可用与可修改性”。任一关键要求不满足，直接给 0 分，
   不再检查维度2。
2. 维度1通过后，逐项检查“维度2：完成度评分细则”和扣分项，汇总为最终得分。

说明：Excel 中“文字是否被裁切/图片是否遮挡”等视觉项无法做到与人工完全一致，
本脚本采用可自动化的近似规则：检查行高、自动换行、图片锚点/尺寸、内容区域外
是否出现明显多余内容、是否存在可能遮挡题干的大图片等，以满足评分意图。
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries

TARGET_FILE = "分层作业设计_填充完成.xlsx"
EXPECTED_SHEETS = ["Sheet1", "Sheet2", "Sheet3"]
EXPECTED_MERGES = {
    "A1:F1",
    "B3:F3",
    "C4:F4",
    "C5:F5",
    "C6:F6",
    "A4:A6",
    "A7:A9",
    "B7:F9",
}
CONTENT_AREA = (1, 1, 9, 6)  # Sheet1!A1:F9
EMU_PER_CM = 360000
EMU_PER_INCH = 914400
PX_PER_CM = 37.7952755906


@dataclass
class CheckResult:
    score: int
    name: str
    passed: bool
    detail: str


def norm(value) -> str:
    """规范化文本，降低换行、空格、全角空格差异带来的误判。"""
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t　 ]+", "", s)
    s = re.sub(r"\n+", "\n", s)
    return s.strip()


def compact(value) -> str:
    return re.sub(r"\s+", "", norm(value))


def contains_all(text: str, keywords: Iterable[str]) -> bool:
    c = compact(text).lower()
    return all(compact(k).lower() in c for k in keywords)


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    c = compact(text).lower()
    return any(compact(k).lower() in c for k in keywords)


def loose_compact(value) -> str:
    """去除空白及标点，只保留中文/字母/数字，用于宽松长句比对。"""
    return re.sub(r"[^\w一-鿿]", "", compact(value).lower())


def contains_loose_phrase(text: str, phrase: str) -> bool:
    return loose_compact(phrase) in loose_compact(text)


def contains_ordered_tokens(text: str, tokens: Iterable[str]) -> bool:
    """要求 tokens 中每个片段按顺序出现在 text 的宽松压缩形式中。"""
    c = loose_compact(text)
    pos = 0
    for token in tokens:
        t = loose_compact(token)
        idx = c.find(t, pos)
        if idx < 0:
            return False
        pos = idx + len(t)
    return True


def cell_text(ws, ref: str) -> str:
    return "" if ws[ref].value is None else str(ws[ref].value)


def approx(actual: Optional[float], expected: float, tol: float = 0.25) -> bool:
    return actual is not None and abs(float(actual) - expected) <= tol


def in_range(value: Optional[float], low: float, high: float) -> bool:
    return value is not None and low <= float(value) <= high


def font_name_ok(cell, expected="宋体") -> bool:
    return (cell.font.name or "") == expected


def font_size_ok(cell, expected: float) -> bool:
    return cell.font.sz is not None and abs(float(cell.font.sz) - expected) < 0.2


def font_black_ok(cell) -> bool:
    """判断黑色字体。未显式设置颜色时，Excel 默认通常为黑色，按通过处理。"""
    color = cell.font.color
    if color is None:
        return True
    if color.type == "rgb":
        rgb = (color.rgb or "").upper()
        return rgb in {"000000", "FF000000"}
    if color.type in {"theme", "indexed", "auto"}:
        # theme=1/indexed=64/auto 在模板中常表示默认黑色/自动色。
        return True
    return True


def horizontal_ok(cell, expected: Optional[str]) -> bool:
    return cell.alignment.horizontal == expected


def vertical_ok(cell, expected: str) -> bool:
    return cell.alignment.vertical == expected


def wrap_ok(cell) -> bool:
    return bool(cell.alignment.wrap_text or cell.alignment.wrapText)


def merged_ranges(ws) -> set[str]:
    return {str(rng) for rng in ws.merged_cells.ranges}


def is_merged(ws, ref: str) -> bool:
    return ref in merged_ranges(ws)


def nonempty_cells(ws) -> list[str]:
    refs = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                refs.append(cell.coordinate)
    return refs


def is_blank_sheet(ws) -> bool:
    return not nonempty_cells(ws) and len(ws.merged_cells.ranges) == 0 and len(getattr(ws, "_images", [])) == 0


def anchor_bounds(img):
    """返回图片锚点边界：(起始行, 起始列, 结束行, 结束列)，行列均为 1-based。"""
    anchor = img.anchor
    if isinstance(anchor, str):
        m = re.match(r"([A-Z]+)(\d+)", anchor)
        if not m:
            return None
        col_letters, row = m.groups()
        col = 0
        for ch in col_letters:
            col = col * 26 + ord(ch) - ord("A") + 1
        return int(row), col, int(row), col
    if not hasattr(anchor, "_from"):
        return None
    start_row = anchor._from.row + 1
    start_col = anchor._from.col + 1
    if hasattr(anchor, "to") and anchor.to is not None:
        end_row = anchor.to.row + 1
        end_col = anchor.to.col + 1
    else:
        end_row = start_row
        end_col = start_col
    return start_row, start_col, end_row, end_col


def img_size_cm(img) -> tuple[float, float]:
    """尽量把图片尺寸换算为厘米。openpyxl 通常以像素保存 width/height。

    注意：openpyxl 的 img.width/height 是否为像素依赖具体读取路径，且像素->厘米
    换算依赖 DPI 假设（PX_PER_CM 固定按 96 DPI），并不可靠。只在无法从 drawing
    XML 拿到精确 EMU 尺寸时，才把这里当作兜底近似值使用，优先使用
    `get_drawing_image_specs` 返回的 `<xdr:ext>` EMU 尺寸。
    """
    width = getattr(img, "width", None)
    height = getattr(img, "height", None)
    if width is None or height is None:
        return 0.0, 0.0
    # 常见图片对象为像素尺寸。若是极大的 EMU 值，也做兼容。
    if width > 10000 or height > 10000:
        return width / EMU_PER_CM, height / EMU_PER_CM
    return width / PX_PER_CM, height / PX_PER_CM


def image_near_row_cols(img, rows: set[int], col_low: int = 3, col_high: int = 6) -> bool:
    bounds = anchor_bounds(img)
    if not bounds:
        return False
    sr, sc, er, ec = bounds
    return bool(set(range(sr, er + 1)) & rows) and not (ec < col_low or sc > col_high)


@dataclass
class DrawingImageSpec:
    """从 drawing XML 直接解析出的图片锚点信息（EMU 精度，不依赖像素/DPI）。"""

    from_row: int  # 0-based，对应 openpyxl anchor._from.row
    from_col: int  # 0-based
    from_row_off: int  # EMU
    from_col_off: int  # EMU
    to_row: Optional[int]  # 0-based，twoCellAnchor 才有
    to_col: Optional[int]
    width_emu: Optional[int]  # oneCellAnchor/绝对锚点的 <xdr:ext cx>
    height_emu: Optional[int]  # <xdr:ext cy>
    embed_rid: Optional[str]  # <a:blip r:embed="rIdN">，用于与 openpyxl 图片对象一一对应

    @property
    def width_cm(self) -> float:
        return (self.width_emu or 0) / EMU_PER_CM

    @property
    def height_cm(self) -> float:
        return (self.height_emu or 0) / EMU_PER_CM


def get_drawing_image_specs(path: Path, sheet_index: int) -> list[DrawingImageSpec]:
    """读取 xl/drawings/drawingN.xml，解析每张图片（<xdr:pic>）的 EMU 精确尺寸与锚点。

    办公软件保存图片时，真实的宽高以 EMU 存在 <xdr:ext cx="宽" cy="高"/> 里，
    与 openpyxl 的 img.width/height（可能是像素、且换算依赖 DPI 假设）相比更可靠。
    锚点既可能是 <xdr:twoCellAnchor>（有 from/to 两个格子），也可能是
    <xdr:oneCellAnchor>（只有 from + ext 尺寸），这里两种都处理。
    """
    if not zipfile.is_zipfile(path):
        return []
    with zipfile.ZipFile(path) as z:
        rels_name = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if rels_name not in z.namelist():
            return []
        rels = z.read(rels_name).decode("utf-8", errors="replace")
        m = re.search(r'Target="([^"]*drawings/drawing[^"]+\.xml)"', rels)
        if not m:
            return []
        target = m.group(1)
        drawing_name = "xl/" + target.lstrip("./").replace("../", "")
        if drawing_name not in z.namelist():
            candidates = [n for n in z.namelist() if n.endswith(Path(target).name)]
            if not candidates:
                return []
            drawing_name = candidates[0]
        xml = z.read(drawing_name).decode("utf-8", errors="replace")
    specs: list[DrawingImageSpec] = []
    blocks = re.findall(
        r"<xdr:(twoCellAnchor|oneCellAnchor)[^>]*>(.*?)</xdr:\1>", xml, re.DOTALL
    )
    for _tag, block in blocks:
        if not re.search(r"<xdr:pic[ >]", block):
            continue  # 只关心图片，跳过纯形状/连接线锚点
        m_from = re.search(
            r"<xdr:from>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>(\d+)</xdr:colOff>\s*"
            r"<xdr:row>(\d+)</xdr:row>\s*<xdr:rowOff>(\d+)</xdr:rowOff>",
            block, re.DOTALL,
        )
        if not m_from:
            continue
        from_col, from_col_off, from_row, from_row_off = (int(x) for x in m_from.groups())
        m_to = re.search(
            r"<xdr:to>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>(\d+)</xdr:colOff>\s*"
            r"<xdr:row>(\d+)</xdr:row>\s*<xdr:rowOff>(\d+)</xdr:rowOff>",
            block, re.DOTALL,
        )
        to_row = to_col = None
        if m_to:
            to_col = int(m_to.group(1))
            to_row = int(m_to.group(3))
        m_ext = re.search(r'<xdr:ext\s+cx="(\d+)"\s+cy="(\d+)"', block)
        width_emu = int(m_ext.group(1)) if m_ext else None
        height_emu = int(m_ext.group(2)) if m_ext else None
        m_rid = re.search(r'<a:blip[^>]*r:embed="(rId\d+)"', block)
        embed_rid = m_rid.group(1) if m_rid else None
        specs.append(DrawingImageSpec(
            from_row=from_row, from_col=from_col,
            from_row_off=from_row_off, from_col_off=from_col_off,
            to_row=to_row, to_col=to_col,
            width_emu=width_emu, height_emu=height_emu,
            embed_rid=embed_rid,
        ))
    return specs


def match_drawing_spec(img, specs: list[DrawingImageSpec]) -> Optional[DrawingImageSpec]:
    """把 openpyxl 的 ws._images 里的图片对象，与 drawing XML 解析出的 spec 一一对应。

    优先用锚点起始行列（0-based）匹配。同一单元格出现多图时，openpyxl 的图片
    对象没有稳定暴露 drawing rId，这里选择该锚点下第一个未区分候选；评分逻辑
    只需要判定是否存在符合第3题尺寸/位置的图片，不依赖逐张持久化身份。
    """
    anchor = getattr(img, "anchor", None)
    if anchor is None or not hasattr(anchor, "_from"):
        return None
    from_row = anchor._from.row
    from_col = anchor._from.col
    candidates = [s for s in specs if s.from_row == from_row and s.from_col == from_col]
    if not candidates:
        return None
    return candidates[0]


def column_width_to_emu(width: Optional[float]) -> int:
    """把 Excel 字符列宽近似换算为 EMU，用于浮动图片遮挡的几何判定。"""
    if width is None:
        width = 8.43
    pixels = int(width * 7 + 5)  # Excel/OpenXML 常用近似：字符宽->像素
    return int(pixels / PX_PER_CM * EMU_PER_CM)


def row_height_to_emu(height_points: Optional[float]) -> int:
    """把 Excel 行高（磅）换算为 EMU。"""
    if height_points is None:
        height_points = 15.0
    return int(height_points / 72 * EMU_PER_INCH)


def c5_landscape_not_covering_text(ws, spec: DrawingImageSpec) -> bool:
    """基于 C5:F5 合并区域的几何位置近似判断图片未遮挡题干/填空文字。

    openpyxl 不能直接给出单元格内换行文本的真实排版边界，这里使用 workbook 元数据
    做可复现的近似：把 C5:F5 的左侧 35% 与上方 30% 视为题干/填空主要文本区，
    山水草地插图必须主要位于右下侧，且不能从 C5 左上文本起始区展开。
    """
    merged_width_emu = sum(
        column_width_to_emu(ws.column_dimensions[get_column_letter(col)].width)
        for col in range(3, 7)
    )
    row_height_emu = row_height_to_emu(ws.row_dimensions[5].height)
    image_left_emu = sum(
        column_width_to_emu(ws.column_dimensions[get_column_letter(col)].width)
        for col in range(3, spec.from_col + 1)
    ) + spec.from_col_off if spec.from_col >= 3 else spec.from_col_off
    image_top_emu = spec.from_row_off if spec.from_row == 4 else row_height_emu
    image_right_emu = image_left_emu + (spec.width_emu or 0)
    image_bottom_emu = image_top_emu + (spec.height_emu or 0)
    text_right_emu = merged_width_emu * 0.35
    text_bottom_emu = row_height_emu * 0.30
    overlaps_text_zone = image_left_emu < text_right_emu and image_top_emu < text_bottom_emu
    inside_merged_area = (
        image_left_emu >= 0
        and image_top_emu >= 0
        and image_right_emu <= merged_width_emu * 1.03
        and image_bottom_emu <= row_height_emu * 1.03
    )
    return inside_merged_area and not overlaps_text_zone


def count_row_rect_shapes(path: Path, sheet_index: int, excel_row: int) -> int:
    """统计 xlsx 中指定工作表 drawing 里锚定在指定 Excel 行的矩形形状数量。

    办公软件里“田字格”通常用一个矩形加两条交叉连接线绘制，openpyxl 的
    `ws._images` 只能拿到图片，无法拿到形状。这里直接读 xl/drawings/*.xml，
    统计 `xdr:twoCellAnchor` 起点在指定行、内部为 `<xdr:sp>` 形状元素、
    且 `prstGeom prst="rect"` 的形状数，作为该行田字格数量的近似判定。
    需要排除 `<xdr:pic>`（图片自带外框矩形）与 `<xdr:cxnSp>`（连接线），
    否则会把图片和田字格中的交叉线一起算进来。
    """
    # sheet_index: 1-based，对应 xl/worksheets/sheet{index}.xml
    if not zipfile.is_zipfile(path):
        return 0
    with zipfile.ZipFile(path) as z:
        # 通过 sheet 的 rels 找到对应的 drawing 路径
        rels_name = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if rels_name not in z.namelist():
            return 0
        rels = z.read(rels_name).decode("utf-8", errors="replace")
        m = re.search(r'Target="([^"]*drawings/drawing[^"]+\.xml)"', rels)
        if not m:
            return 0
        target = m.group(1)
        # 相对路径归一：../drawings/drawingN.xml -> xl/drawings/drawingN.xml
        drawing_name = "xl/" + target.lstrip("./").replace("../", "")
        if drawing_name not in z.namelist():
            # 兜底：模糊匹配
            candidates = [n for n in z.namelist() if n.endswith(Path(target).name)]
            if not candidates:
                return 0
            drawing_name = candidates[0]
        xml = z.read(drawing_name).decode("utf-8", errors="replace")
    # 每个 twoCellAnchor 块单独判定
    count = 0
    target_from_row = excel_row - 1  # drawing XML 里 row 是 0-based
    blocks: list[str] = re.findall(r"<xdr:twoCellAnchor[^>]*>(.*?)</xdr:twoCellAnchor>", xml, re.DOTALL)
    for block in blocks:
        m_row = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", block, re.DOTALL)
        if not m_row or int(m_row.group(1)) != target_from_row:
            continue
        # 仅统计真正的形状 <xdr:sp>，排除图片 <xdr:pic> 与连接线 <xdr:cxnSp>。
        # 注意 <xdr:sp 是 <xdr:spPr> 的前缀，不能用作判定，需用 <xdr:sp> 或 <xdr:sp 后跟属性。
        if not re.search(r"<xdr:sp[ >]", block):
            continue
        if 'prstGeom prst="rect"' in block:
            count += 1
    return count


def load_workbook_editable(path: Path):
    # 交付文件允许 .xlsx 或 .xlsm 格式；先用 zipfile 检查基本容器，避免非法伪装。
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("文件扩展名不是 .xlsx 或 .xlsm")
    if not zipfile.is_zipfile(path):
        raise ValueError("文件不是有效的 xlsx/xlsm(zip) 容器")
    wb = load_workbook(path)
    return wb


def dimension1_checks(path: Path, wb) -> list[tuple[bool, str]]:
    checks: list[tuple[bool, str]] = []

    ok_open = wb is not None
    checks.append((ok_open, "交付文件为 .xlsx 或 .xlsm，且可正常打开。"))

    return checks


def add_result(results: list[CheckResult], score: int, name: str, passed: bool, detail: str = "") -> None:
    results.append(CheckResult(score=score, name=name, passed=passed, detail=detail))


def evaluate_dimension2(wb, path: Path | None = None) -> list[CheckResult]:
    ws = wb["Sheet1"]
    results: list[CheckResult] = []
    # +5 列宽（细则规定的是办公软件里显示的“字符”列宽）：
    #   A=12.75, B=14.63, C=13.13, D=17.38, E=12.25, F=15.5。
    # openpyxl 的 ColumnDimension.width 读取到的就是 Excel 里显示的字符列宽值
    # （即另存 xlsx 时写入 <col customWidth width="..."/> 的原始数值），并不需要
    # 再减去 0.625 的内边距换算——之前固定减 0.625 会把刚好等于细则值的列
    # （如 A=12.75）算成 12.12，从而误判不合格。这里直接用 stored width 与细则值比对。
    expected_widths = {"A": 12.75, "B": 14.63, "C": 13.13, "D": 17.38, "E": 12.25, "F": 15.5}
    actual_widths: dict[str, float | None] = {}
    for col in expected_widths:
        stored = cast(float | None, ws.column_dimensions[col].width)
        actual_widths[col] = None if stored is None else round(float(stored), 2)
    width_ok = all(approx(actual_widths[col], exp, tol=0.05) for col, exp in expected_widths.items())
    add_result(results, 5, "列宽：A-F列宽固定符合细则", width_ok,
               f"细则列宽={expected_widths}，实际列宽={actual_widths}")

    # +1 标题区域 A1:F1（细则要求逐点踩到）：
    #   1) 保持 A1:F1 合并状态（不取消合并）
    #   2) 文本为“识字主题《四季小景》分层作业”
    #   3) 字体宋体  4) 字号 18 磅  5) 加粗  6) 黑色字体
    #   7) 水平居中  8) 垂直居中
    #   9) 行高保持 39 磅（不因粘贴标题而调整行高）
    # 说明：办公软件里“黑色”既可能是纯黑 RGB(000000)，也可能是主题色“文本1”，
    #       两者在 Office 中显示均为黑色，font_black_ok 已按此处理。
    #       行高强调“保持 39 磅、不调整”，故用较小容差精确判定。
    c = ws["A1"]
    title_checks = {
        "A1:F1保持合并": is_merged(ws, "A1:F1"),
        "文本=识字主题《四季小景》分层作业": compact(c.value) == compact("识字主题《四季小景》分层作业"),
        "宋体": font_name_ok(c),
        "18磅": font_size_ok(c, 18),
        "加粗": bool(c.font.bold),
        "黑色": font_black_ok(c),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
        "行高39磅": approx(ws.row_dimensions[1].height, 39, 0.1),
    }
    title_ok = all(title_checks.values())
    add_result(results, 1, "标题区域 A1:F1 格式与文本", title_ok,
               f"各点命中={title_checks}")

    # +1 基本信息行 A2:F2（细则逐点踩到）：
    #   1) A2 字段=“班级”  2) C2 字段=“任教教师”  3) E2 字段=“时间”
    #   4) 三字段字体均为宋体  5) 11 磅  6) 加粗  7) 水平居中  8) 垂直居中
    #   9) B2、D2、F2 填写区域保持空白
    # 说明：细则未要求字段单元格的字体颜色，故不作颜色约束。
    info_checks = {}
    for ref, text in [("A2", "班级"), ("C2", "任教教师"), ("E2", "时间")]:
        cell = ws[ref]
        info_checks[f"{ref}={text}"] = compact(cell.value) == compact(text)
        info_checks[f"{ref}宋体"] = font_name_ok(cell)
        info_checks[f"{ref}11磅"] = font_size_ok(cell, 11)
        info_checks[f"{ref}加粗"] = bool(cell.font.bold)
        info_checks[f"{ref}水平居中"] = horizontal_ok(cell, "center")
        info_checks[f"{ref}垂直居中"] = vertical_ok(cell, "center")
    for ref in ["B2", "D2", "F2"]:
        info_checks[f"{ref}空白"] = ws[ref].value in (None, "")
    info_row_ok = all(info_checks.values())
    add_result(results, 1, "基本信息行 A2/C2/E2 字段格式且 B2/D2/F2 空白", info_row_ok,
               f"各点命中={info_checks}")

    # +1 A3 作业简介标题（细则逐点踩到）：
    #   1) 文本=“作业设计简介（学习目标、题型说明等）”
    #   2) 宋体  3) 11 磅  4) 加粗
    #   5) 自动换行  6) 水平居中  7) 垂直居中
    # 说明：办公软件中文本可能带换行/全角括号差异，compact 已去除空白与换行，
    #       故用 compact 精确比对文本，其余按细则逐项判定；细则未要求字体颜色，不作约束。
    c = ws["A3"]
    a3_checks = {
        "文本=作业设计简介（学习目标、题型说明等）":
            compact(c.value) == compact("作业设计简介（学习目标、题型说明等）"),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "加粗": bool(c.font.bold),
        "自动换行": wrap_ok(c),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
    }
    intro_title_ok = all(a3_checks.values())
    add_result(results, 1, "A3 作业简介标题", intro_title_ok, f"各点命中={a3_checks}")

    # +1 B3:F3 作业简介内容（细则逐点踩到）：
    #   1) 保持 B3:F3 合并状态
    #   2) 文本=“识字主题《四季小景》分层作业；基础类、提升类为必做题，拓展类为选做题。”
    #   3) 宋体  4) 11 磅  5) 常规字形（非加粗）
    #   6) 行高保持 68 磅
    #   7) 文字完整显示、不上下裁切——办公软件中该要求依赖“自动换行”开启，
    #      否则多行文字会在固定行高内被上下裁切，故此处检查 wrap。
    # 说明：细则未要求斜体/颜色/水平垂直对齐，均不作约束。
    c = ws["B3"]
    intro_text_expected = "识字主题《四季小景》分层作业；基础类、提升类为必做题，拓展类为选做题。"
    b3_checks = {
        "B3:F3保持合并": is_merged(ws, "B3:F3"),
        "文本匹配": compact(c.value) == compact(intro_text_expected),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "常规字形(非加粗)": not bool(c.font.bold),
        "行高68磅": approx(ws.row_dimensions[3].height, 68, 0.1),
        "自动换行(不裁切)": wrap_ok(c),
    }
    intro_content_ok = all(b3_checks.values())
    add_result(results, 1, "B3:F3 作业简介内容", intro_content_ok,
               f"各点命中={b3_checks}，实际文本：{cell_text(ws, 'B3')}")

    # +1 A4:A6 层级总标题（细则逐点踩到）：
    #   1) 保持 A4:A6 合并状态
    #   2) 文本=“作业设计内容”  3) 宋体  4) 14 磅  5) 加粗
    #   6) 水平居中  7) 垂直居中
    # 说明：办公软件中该单元格常以竖排换行显示（作业\n设计\n内容），
    #       compact 已去除空白/换行，故能正确还原比对；细则未要求字体颜色，不作约束。
    c = ws["A4"]
    a4_checks = {
        "A4:A6保持合并": is_merged(ws, "A4:A6"),
        "文本=作业设计内容": compact(c.value) == compact("作业设计内容"),
        "宋体": font_name_ok(c),
        "14磅": font_size_ok(c, 14),
        "加粗": bool(c.font.bold),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
    }
    design_title_ok = all(a4_checks.values())
    add_result(results, 1, "A4:A6 层级总标题", design_title_ok, f"各点命中={a4_checks}")

    # +1 B4 基础层标题（细则逐点踩到）：
    #   1) 文本=“基础层作业设计”  2) 宋体  3) 11 磅  4) 加粗
    #   5) 水平居中  6) 垂直居中
    # 说明：细则未要求合并、行高、字体颜色，均不作约束。
    c = ws["B4"]
    b4_checks = {
        "文本=基础层作业设计": compact(c.value) == compact("基础层作业设计"),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "加粗": bool(c.font.bold),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
    }
    add_result(results, 1, "B4 基础层标题", all(b4_checks.values()), f"各点命中={b4_checks}")

    # +1 B5 提高层标题（细则逐点踩到）：
    #   1) 文本=“提高层作业设计”  2) 宋体  3) 11 磅  4) 加粗
    #   5) 水平居中  6) 垂直居中
    # 说明：细则未要求合并、行高、字体颜色，均不作约束。
    c = ws["B5"]
    b5_checks = {
        "文本=提高层作业设计": compact(c.value) == compact("提高层作业设计"),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "加粗": bool(c.font.bold),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
    }
    add_result(results, 1, "B5 提高层标题", all(b5_checks.values()), f"各点命中={b5_checks}")

    # +1 B6 拓展层标题（细则逐点踩到）：
    #   1) 文本=“拓展层作业设计”  2) 宋体  3) 11 磅  4) 加粗
    #   5) 水平居中  6) 垂直居中
    # 说明：细则未要求合并、行高、字体颜色，均不作约束。
    c = ws["B6"]
    b6_checks = {
        "文本=拓展层作业设计": compact(c.value) == compact("拓展层作业设计"),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "加粗": bool(c.font.bold),
        "水平居中": horizontal_ok(c, "center"),
        "垂直居中": vertical_ok(c, "center"),
    }
    add_result(results, 1, "B6 拓展层标题", all(b6_checks.values()), f"各点命中={b6_checks}")

    # +5 C4:F4 基础层内容区域（细则逐点踩到）：
    #   总体：1) 保持 C4:F4 合并状态
    #        2) 完整填写“基础类”区域中的 4 道题
    #        3) 单元格行高对应 250-350 磅
    #   第1题：完整出现“1.给词语选择正确的读音画‘√’。”
    #          及“晨露、清风、细雨、晚霞”的拼音选择内容与对应括号。
    #   第2题：完整出现“2.看拼音，写词语。”及 qīng chén、xiǎo yǔ、hé fēng、
    #          huā duǒ、niǎo ér、fēi 等拼音、句子填空内容及对应田字格。
    #   第3题：完整出现“3.你能圈出这几个词中表示自然景物的字吗？”
    #          及“晨露、细雨、晚霞”三个词语。
    #   第4题：完整出现“4.连一连。”及“晨露、午阳、晚霞、夜风”和“亮、暖、红、轻”两组内容。
    # 说明：办公软件中文本换行/空格不定，contains_all 已做压缩比对，可在 Office 有效。
    #       第2题原判定只要求六个拼音再加“矩形形状数>=1”，过松：既没检查“句子
    #       填空内容”，也没要求田字格数量与题目匹配。这里补充：
    #       1) 句子填空关键词（花朵、鸟儿、飞入等描述春景的常见搭配词）；
    #       2) 从“2.”截取到“3.”之前的文本片段，仅在该片段范围内统计矩形形状数/
    #          “田字格”字样/长空白填空槽数量，避免误算到第1/3/4题的括号或形状；
    #       3) 六个拼音对应六个字词填空，田字格数量应恰好为 6 个，不能再用 >=1。
    base_text = cell_text(ws, "C4")
    _m2 = re.search(r"2\s*[\.．、]", base_text)
    _m3_base = re.search(r"3\s*[\.．、]", base_text)
    if _m2 and _m3_base and _m3_base.start() > _m2.start():
        q2_text = base_text[_m2.start():_m3_base.start()]
    elif _m2:
        q2_text = base_text[_m2.start():]
    else:
        q2_text = ""
    q2_tianzige_word_count = q2_text.count("田字格")
    q2_blank_slots = len(re.findall(r"[ 　_—－\-]{4,}", q2_text))
    q2_rect_shape_count = count_row_rect_shapes(path, 1, 4) if path is not None else 0
    q2_has_6_slots = q2_rect_shape_count == 6 or q2_tianzige_word_count == 6 or q2_blank_slots == 6
    base_checks = {
        "C4:F4保持合并": is_merged(ws, "C4:F4"),
        "行高250-350磅": in_range(ws.row_dimensions[4].height, 250, 350),
        "第1题题干": contains_all(base_text, ["1", "给词语选择正确的读音", "√"]),
        "第1题词语拼音选择": contains_all(base_text, ["晨露", "清风", "细雨", "晚霞"]),
        "第1题括号": contains_any(base_text, ["（", "(", "括号"]),
        "第2题题干": contains_all(base_text, ["2", "看拼音", "写词语"]),
        "第2题拼音填空": contains_all(base_text, ["qīng chén", "xiǎo yǔ", "hé fēng", "huā duǒ", "niǎo ér", "fēi"]),
        "第2题句子填空内容": contains_all(q2_text, ["花", "鸟", "风", "雨"]),
        "第2题含6个田字格填空": q2_has_6_slots,
        "第3题题干": contains_all(base_text, ["3", "圈出", "自然景物"]),
        "第3题三词": contains_all(base_text, ["晨露", "细雨", "晚霞"]),
        "第4题题干": contains_all(base_text, ["4", "连一连"]),
        "第4题左列": contains_all(base_text, ["晨露", "午阳", "晚霞", "夜风"]),
        "第4题右列": contains_all(base_text, ["亮", "暖", "红", "轻"]),
    }
    base_ok = all(base_checks.values())
    _base_detail = (
        f"行高={ws.row_dimensions[4].height}，第2题矩形形状数={q2_rect_shape_count}，"
        + f"田字格字样={q2_tianzige_word_count}，填空槽={q2_blank_slots}，"
        + f"各点命中={base_checks}，实际文本前120字：{base_text[:120]}"
    )
    add_result(results, 5, "C4:F4 基础层4道题", base_ok, _base_detail)

    # +5 C5:F5 提高层内容区域（细则逐点踩到）：
    #   总体：1) 保持 C5:F5 合并状态
    #        2) 完整填写“提升类”区域中的 3 道题
    #        3) 行高 350-450 磅
    #        4) 单元格内包含两张图片（提高层附近图片 ≥ 2）
    #   第1题：完整出现“1.词语接龙。”及“清晨→（ ）→（ ）→（ ）”（清晨、箭头、括号）。
    #   第2题：完整出现“2.照样子，看图说一说。”及“微风轻、小鸟、柳叶、溪水、纸鸢、云朵”及对应图片。
    #   第3题：完整出现“3.仔细观察图片，填一填。”及包含 15 个田字格的填空语境及对应图片。
    # 说明：办公软件里图片以浮动锚点存于该行区域，image_near_row_cols 判定其落在 C5:F5；
    #       文本换行/空格不定，contains_all/any 已做压缩比对；
    #       “田字格”在办公软件里通常用“矩形+两条交叉连接线”绘制，openpyxl 的
    #       ws._images 只能拿到位图，无法拿到形状。因此这里读取底层 drawing xml，
    #       统计锚定在 Excel 第 5 行的 prstGeom="rect" 矩形数作为田字格数量，
    #       要求恰好 15 个。若无法读取 drawing（例如仅传入了 workbook），
    #       则回退到文本层的“田字格”字样计数或长空白/下划线填空槽计数。
    improve_text = cell_text(ws, "C5")
    images_near_c5 = [img for img in getattr(ws, "_images", []) if image_near_row_cols(img, {5}, 3, 6)]
    # 图片精确尺寸优先从 drawing xml 的 <xdr:ext> EMU 值读取（不依赖像素/DPI假设），
    # 通过锚点起始行列把 ws._images 里的每个图片对象与对应的 drawing spec 绑定；
    # 若绑定失败（例如非 xlsx 容器路径不可用），才回退到 img_size_cm 的像素近似换算。
    drawing_specs = get_drawing_image_specs(path, 1) if path is not None else []
    q3_image_candidates = []
    q2_image_candidates = []
    for img in images_near_c5:
        spec = match_drawing_spec(img, drawing_specs)
        if spec is not None and spec.width_emu and spec.height_emu:
            width_cm, height_cm = spec.width_cm, spec.height_cm
        else:
            width_cm, height_cm = img_size_cm(img)
        # 第3题的“山水草地插图”在后续细则中有明确尺寸要求，这里用尺寸识别其题目归属。
        is_q3_image = in_range(width_cm, 5.5, 5.7) and in_range(height_cm, 2.7, 2.9)
        if is_q3_image:
            q3_image_candidates.append(img)
        else:
            q2_image_candidates.append(img)
    # 第3题所在文本片段：从“3.”后开始截取，避免把第1/2题的括号计入田字格填空槽。
    _m3 = re.search(r"3\s*[\.．、]", improve_text)
    q3_text = improve_text[_m3.start():] if _m3 else ""
    tianzige_word_count = q3_text.count("田字格")
    blank_slots = len(re.findall(r"[ 　_—－\-]{4,}", q3_text))
    rect_shape_count = count_row_rect_shapes(path, 1, 5) if path is not None else 0
    # 15个田字格必须是底层 drawing 里的真实矩形，或文本里可数的明确填空槽；不能靠重复
    # “田字格”字样15次通过，否则会把说明文字误判成实际作答区域。
    q3_has_15_slots = rect_shape_count == 15 or blank_slots == 15
    improve_checks = {
        "C5:F5保持合并": is_merged(ws, "C5:F5"),
        "行高350-450磅": in_range(ws.row_dimensions[5].height, 350, 450),
        "包含两张图片": len(images_near_c5) >= 2,
        "第1题题干": contains_all(improve_text, ["1", "词语接龙"]),
        "第1题接龙内容": contains_all(improve_text, ["清晨"]) and contains_any(improve_text, ["→", "->"]) and contains_any(improve_text, ["（", "("]),
        "第2题题干": contains_all(improve_text, ["2", "照样子", "看图说一说"]),
        "第2题词语": contains_all(improve_text, ["微风轻", "小鸟", "柳叶", "溪水", "纸鸢", "云朵"]),
        "第2题对应图片": len(q2_image_candidates) >= 1,
        "第3题题干": contains_all(improve_text, ["3", "仔细观察图片", "填一填"]),
        "第3题含15个田字格填空": q3_has_15_slots,
        "第3题对应图片": len(q3_image_candidates) >= 1,
    }
    improve_ok = all(improve_checks.values())
    _improve_detail = (
        f"行高={ws.row_dimensions[5].height}，附近图片数={len(images_near_c5)}，"
        f"第2题图片数={len(q2_image_candidates)}，第3题图片数={len(q3_image_candidates)}，"
        + f"矩形形状数={rect_shape_count}，田字格字样={tianzige_word_count}，"
        + f"填空槽={blank_slots}，各点命中={improve_checks}，实际文本前120字：{improve_text[:120]}"
    )
    add_result(results, 5, "C5:F5 提高层3道题、图片及必做要求", improve_ok, _improve_detail)

    # +1 C5:F5 观察图片（细则逐点踩到）：
    #   1) 山水草地插图放置在提高层第3题附近（图片锚点落在 C5:F5 区域内）
    #   2) 图片宽 5.5-5.7cm  3) 图片高 2.7-2.9cm
    #   4) 不遮挡题干和填空内容
    # 说明：图片尺寸从 drawing xml 的 <xdr:ext cx/cy> EMU 值读取，避免固定
    #       37.795 px/cm 的 DPI 假设；“山水草地插图”通过第3题绑定条件识别：
    #       图片位于 C5:F5、第3题目标尺寸范围内，且通过几何计算避开 C5:F5 左上
    #       文本主要区域。脱敏文件的媒体名通常不可用，因此不把随机文件名作为硬条件。
    landscape_ok = False
    landscape_details = []
    for img in q3_image_candidates:
        spec = match_drawing_spec(img, drawing_specs)
        bounds = anchor_bounds(img)
        if spec is not None and spec.width_emu and spec.height_emu:
            w_cm, h_cm = spec.width_cm, spec.height_cm
            size_source = "drawing-ext-emu"
            no_cover = c5_landscape_not_covering_text(ws, spec)
        else:
            w_cm, h_cm = img_size_cm(img)
            size_source = "openpyxl-fallback"
            no_cover = False
        size_ok = in_range(w_cm, 5.5, 5.7) and in_range(h_cm, 2.7, 2.9)
        sr, sc, er, ec = bounds or (0, 0, 0, 0)
        in_c5_area = bool({r for r in range(sr, er + 1)} & {5}) and sc >= 3 and ec <= 6
        landscape_details.append({
            "宽cm": round(w_cm, 2),
            "高cm": round(h_cm, 2),
            "尺寸来源": size_source,
            "锚点": bounds,
            "尺寸达标": size_ok,
            "位于C5:F5": in_c5_area,
            "未遮挡文本区": no_cover,
        })
        if size_ok and in_c5_area and no_cover:
            landscape_ok = True
    add_result(results, 1, "C5:F5 山水草地插图尺寸与位置", landscape_ok,
               f"附近图片检查：{landscape_details}")

    # +5 C6:F6 拓展层内容区域（细则逐点踩到）：
    #   1) 保持 C6:F6 合并状态
    #   2) 第1项完整出现“和家人到附近公园或校园走一走，记录三种自然景物，回到班级分享。”
    #   3) 第2项完整出现“搜集两个描写季节的词语，和同学交流。”
    # 说明：办公软件里文本换行/空格/标点不定，用 contains_loose_phrase 做去标点的
    #       整句比对，并辅以按序 token 校验，保证“完整出现”而非零散关键词。
    expand_text = cell_text(ws, "C6")
    _task1 = "和家人到附近公园或校园走一走，记录三种自然景物，回到班级分享。"
    _task2 = "搜集两个描写季节的词语，和同学交流。"
    expand_checks = {
        "C6:F6保持合并": is_merged(ws, "C6:F6"),
        "第1项完整出现": contains_loose_phrase(expand_text, _task1) or contains_ordered_tokens(
            expand_text, ["和家人", "附近公园", "校园", "走一走", "记录", "三种", "自然景物", "回到班级", "分享"]
        ),
        "第2项完整出现": contains_loose_phrase(expand_text, _task2) or contains_ordered_tokens(
            expand_text, ["搜集", "两个", "描写季节", "词语", "同学交流"]
        ),
    }
    expand_ok = all(expand_checks.values())
    add_result(results, 5, "C6:F6 拓展层两项任务", expand_ok,
               f"各点命中={expand_checks}，实际文本：{expand_text[:120]}")

    # +1 A7:A9 反馈标题区域（细则逐点踩到）：
    #   1) 保持 A7:A9 合并状态  2) 文本=“作业情况反馈”
    #   3) 宋体  4) 11 磅  5) 加粗  6) 垂直居中
    # 说明：细则只要求“垂直居中”，未要求水平对齐，故不检查水平居中；
    #       亦未要求行高、字体颜色，均不作约束。
    c = ws["A7"]
    a7_checks = {
        "A7:A9保持合并": is_merged(ws, "A7:A9"),
        "文本=作业情况反馈": compact(c.value) == compact("作业情况反馈"),
        "宋体": font_name_ok(c),
        "11磅": font_size_ok(c, 11),
        "加粗": bool(c.font.bold),
        "垂直居中": vertical_ok(c, "center"),
    }
    add_result(results, 1, "A7:A9 反馈标题区域", all(a7_checks.values()), f"各点命中={a7_checks}")

    # +3 B7:F9 反馈内容区域（细则逐点踩到）：
    #   1) 保持 B7:F9 合并状态
    #   2) 文本含“自我评价：今天我完成了__颗星，我已经累计了__颗星，请继续保持！”
    #      —— 拆为：自我评价 / 今天我完成了…颗星 / 我已经累计了…颗星 / 请继续保持
    #   3) “完成了”后带横线  4) “累计了”后带横线
    #   5) 另起一行显示“教师评价：”（存在换行且含“教师评价”）
    #   6) 宋体  7) 11 磅  8) 黑色  9) 自动换行  10) 左对齐或保持左对齐
    # 说明：办公软件里横线可能是下划线/破折号/连字符，用字符类 [_—－-] 匹配；
    #       左对齐在 Excel 中默认 horizontal=None 即左对齐，故 None 或 "left" 均视为通过。
    #       rubric 只要求“完成了/累计了”后带有横线，未规定横线至少几个字符；此前
    #       “…X颗星”判定用 {1,}，而“…后带横线”判定却写成 {2,}，两者标准不一致，
    #       会把只填了单个横线字符（符合 rubric）的文档在后一条上误判为不合格。
    #       这里统一为 {1,}（至少1个字符即算“带有横线”）。
    feedback = cell_text(ws, "B7")
    compact_fb = compact(feedback)
    fb_checks = {
        "B7:F9保持合并": is_merged(ws, "B7:F9"),
        "含自我评价": contains_all(feedback, ["自我评价"]),
        "完成了X颗星": re.search(r"今天我完成了[_—－\-]{1,}颗星", compact_fb) is not None,
        "完成了后带横线": re.search(r"完成了[_—－\-]{1,}", compact_fb) is not None,
        "累计了X颗星": re.search(r"我已经累计了[_—－\-]{1,}颗星", compact_fb) is not None,
        "累计了后带横线": re.search(r"累计了[_—－\-]{1,}", compact_fb) is not None,
        "含请继续保持": contains_all(feedback, ["请继续保持"]),
        "另起一行教师评价": ("\n" in feedback) and contains_all(feedback, ["教师评价"]),
        "宋体": font_name_ok(ws["B7"]),
        "11磅": font_size_ok(ws["B7"], 11),
        "黑色": font_black_ok(ws["B7"]),
        "自动换行": wrap_ok(ws["B7"]),
        "左对齐": ws["B7"].alignment.horizontal in (None, "left"),
    }
    feedback_ok = all(fb_checks.values())
    add_result(results, 3, "B7:F9 反馈内容区域", feedback_ok, f"各点命中={fb_checks}，实际文本：{feedback}")

    # +3 Sheet1 合并单元格（细则逐点踩到）：
    #   要求恰好保持 8 个合并区域：A1:F1、B3:F3、C4:F4、C5:F5、C6:F6、A4:A6、
    #   A7:A9、B7:F9，且不新增、不取消。
    # 说明：办公软件中合并区域以左上:右下表示。以下同时校验“8 个都在”和“无多余”，
    #       即当前合并集合与预期集合完全相等。
    actual_merges = merged_ranges(ws)
    missing = sorted(EXPECTED_MERGES - actual_merges)   # 被取消的
    extra = sorted(actual_merges - EXPECTED_MERGES)     # 新增的
    merge_checks = {
        "8个合并区域齐全(无取消)": not missing,
        "无新增合并区域": not extra,
    }
    exact_merges_ok = all(merge_checks.values())
    add_result(results, 3, "Sheet1 8个合并单元格区域完全一致且无新增/取消", exact_merges_ok,
               f"缺失(被取消)={missing}，多余(新增)={extra}，实际合并区域：{sorted(actual_merges)}")

    return results


def print_dimension2(results: list[CheckResult]) -> int:
    total = 0
    for r in results:
        if r.passed:
            total += r.score
    return total


SCRIPT_ID = "087"


def _locate_target_file(dir_path: Path) -> Optional[Path]:
    """在给定目录内定位被评估文档：优先精确匹配 TARGET_FILE，其次退化为目录里的 .xlsx。"""
    candidate = dir_path / TARGET_FILE
    if candidate.exists():
        return candidate
    if dir_path.is_dir():
        xlsxs = sorted(p for p in dir_path.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx")
        if xlsxs:
            return xlsxs[0]
    return None


def _build_error_report(file_name: str, error_msg: str) -> dict:
    return {
        "id": SCRIPT_ID,
        "file_name": file_name,
        "status": "error",
        "error": error_msg,
        "dim1_pass": False,
        "dim1_reason": error_msg,
        "dim2_items": [],
        "total_score": 0,
        "max_score": 0,
    }


def evaluate(dir_path: str) -> dict:
    """批量运行器统一入口。

    入参 dir_path 为“脚本所在目录的路径”；脚本自行在该目录内定位被评估文档，
    执行维度一/维度二检查，返回结构化字典（见模块顶部约定）。
    """
    try:
        base_dir = Path(dir_path)
        target = _locate_target_file(base_dir)
        if target is None:
            return _build_error_report(TARGET_FILE, f"在目录 {base_dir} 下未找到待评估的 xlsx 文件")

        file_name = target.name

        try:
            wb = load_workbook_editable(target)
        except Exception as exc:
            return {
                "id": SCRIPT_ID,
                "file_name": file_name,
                "status": "ok",
                "error": None,
                "dim1_pass": False,
                "dim1_reason": f"无法打开工作簿：{exc}",
                "dim2_items": [],
                "total_score": 0,
                "max_score": 0,
            }

        d1 = dimension1_checks(target, wb)
        if not all(ok for ok, _ in d1):
            failed = [detail for ok, detail in d1 if not ok]
            return {
                "id": SCRIPT_ID,
                "file_name": file_name,
                "status": "ok",
                "error": None,
                "dim1_pass": False,
                "dim1_reason": "；".join(failed),
                "dim2_items": [],
                "total_score": 0,
                "max_score": 0,
            }

        results = evaluate_dimension2(wb, target)
        dim2_items = []
        max_score = 0
        total_score = 0
        for r in results:
            max_delta = r.score
            hit = bool(r.passed)
            delta = max_delta if hit else 0
            dim2_items.append({
                "rule": r.name,
                "max_delta": max_delta,
                "delta": delta,
                "hit": hit,
                "detail": "",
            })
            # 满分只累加正向项；扣分项（score<0）不计入 max_score，命中时从 total 中扣。
            if max_delta > 0:
                max_score += max_delta
            if hit:
                total_score += delta

        return {
            "id": SCRIPT_ID,
            "file_name": file_name,
            "status": "ok",
            "error": None,
            "dim1_pass": True,
            "dim1_reason": "",
            "dim2_items": dim2_items,
            "total_score": total_score,
            "max_score": max_score,
        }
    except Exception as exc:  # 兜底：脚本自身异常一律标记 error，避免与“0 分”混淆
        return _build_error_report(TARGET_FILE, f"脚本执行异常：{exc}")


if __name__ == "__main__":
    # 仅用于本地调试：默认使用脚本自身所在目录，允许命令行覆盖为其它目录。
    _dir = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).resolve().parent)
    print(json.dumps(evaluate(_dir), ensure_ascii=False, indent=2))

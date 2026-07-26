#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 自动评分脚本：子表数据提取到汇总表。

对外接口：
    evaluate(dir_path: str) -> dict
        参数 dir_path 为脚本所在目录路径，脚本自己在该目录内定位并打开被评估的
        xlsx/xlsm 文档，返回结构化评分字典。

评分逻辑：
1. 先检查维度1（可用与可修改性），任意失败则最终 0 分并停止。
2. 维度1全部通过后，继续检查维度2，命中得分点/扣分点后累计分数。
3. 结果通过返回值以结构化 dict 交付，不通过 print 输出主结果。
"""

from __future__ import annotations

import json
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


ERROR_VALUES = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
EXPECTED_HEADERS = [
    "序号",
    "构件分区",
    "子表",
    "新构件编号",
    "楼层/部位",
    "砼体积(m³)",
    "总重量合计(kg)",
    "含量(kg/m³)",
    "备注",
]
EXPECTED_CHILD_SHEETS = [f"Sheet{i}" for i in list(range(1, 11)) + list(range(12, 62))]
EXPECTED_SHEETS = ["汇总"] + EXPECTED_CHILD_SHEETS
STRUCT_A_SHEETS = ["Sheet3", *[f"Sheet{i}" for i in range(6, 11)], *[f"Sheet{i}" for i in range(12, 40)], "Sheet41"]
STRUCT_B_SHEETS = ["Sheet5", *[f"Sheet{i}" for i in range(42, 62)]]


@dataclass
class CheckResult:
    passed: bool
    label: str
    detail: str = ""
    score: int = 0


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "")


def normalize_formula(value: Any) -> str:
    return normalize_text(value).upper().replace("'", "")


def values_equal(a: Any, b: Any, tolerance: float = 1e-6) -> bool:
    if is_number(a) and is_number(b):
        return abs(float(a) - float(b)) <= tolerance
    return normalize_text(a) == normalize_text(b)


def column_index_to_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def cell_has_formula_error(value: Any) -> bool:
    return isinstance(value, str) and value.strip().upper() in ERROR_VALUES


def find_cell_containing(ws: Worksheet, keyword: str, max_row: int = 80, max_col: int = 20) -> Optional[Cell]:
    target = normalize_text(keyword)
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row), min_col=1, max_col=min(max_col, ws.max_column)):
        for cell in row:
            if target and target in normalize_text(cell.value):
                return cell
    return None


def sheet_has_label(ws: Worksheet, keyword: str) -> bool:
    return find_cell_containing(ws, keyword) is not None


def formula_reference(formula: Any) -> Optional[tuple[str, str]]:
    """从类似 ='Sheet3'!D20 或 =Sheet3!D20 的公式中提取 (sheet, cell)。"""
    if not is_formula(formula):
        return None
    text = formula.strip()
    # 仅提取直接引用。对于 INDIRECT/INDEX 等等效公式，返回 None，由其他检查处理。
    match = re.fullmatch(r"=\s*'?([^'=!]+)'?!\$?([A-Z]{1,3})\$?(\d+)\s*", text, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1), f"{match.group(2).upper()}{match.group(3)}"


_TW_CELL_CACHE: dict[str, set[str]] = {}
_KG_CELL_CACHE: dict[str, set[str]] = {}


def _next_row_coord(coord: str) -> str:
    match = re.match(r"([A-Z]+)(\d+)", coord)
    if not match:
        return coord
    return f"{match.group(1)}{int(match.group(2)) + 1}"


def _scan_child_labels(ws: Worksheet) -> dict[str, str]:
    """返回子表中关键标签所在单元格坐标：砼重量(t)、重量(kg)、总重量合计。"""
    found: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(40, ws.max_row), min_col=1, max_col=min(16, ws.max_column)):
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value.strip()
                if text == "砼重量(t)" and "tw" not in found:
                    found["tw"] = cell.coordinate
                elif text == "重量(kg)" and "kg" not in found:
                    found["kg"] = cell.coordinate
                elif text == "总重量合计" and "tot" not in found:
                    found["tot"] = cell.coordinate
    return found


def forbidden_tw_cells(wb_formula: Any, sheet: str) -> set[str]:
    """“砼重量(t)”列的数据单元格（细则禁止 F 列从此取数）。"""
    if sheet in _TW_CELL_CACHE:
        return _TW_CELL_CACHE[sheet]
    cells: set[str] = set()
    if sheet in wb_formula.sheetnames:
        labels = _scan_child_labels(wb_formula[sheet])
        tw_label = labels.get("tw")
        if tw_label:
            # 标签正下方为该列数值。
            cells.add(_next_row_coord(tw_label).upper())
    _TW_CELL_CACHE[sheet] = cells
    return cells


def forbidden_kg_item_cells(wb_formula: Any, sheet: str) -> set[str]:
    """配筋明细中每一条“重量(kg)”单项数据单元格（细则禁止 G 列用单条代替合计）。"""
    if sheet in _KG_CELL_CACHE:
        return _KG_CELL_CACHE[sheet]
    cells: set[str] = set()
    if sheet in wb_formula.sheetnames:
        labels = _scan_child_labels(wb_formula[sheet])
        kg_label = labels.get("kg")
        if kg_label:
            match = re.match(r"([A-Z]+)(\d+)", kg_label)
            if match:
                col = match.group(1)
                start = int(match.group(2)) + 1
                # 单项数据从标签下一行起，到“总重量合计”行之前；未识别合计行时回退 4 行。
                end = start + 4
                tot_label = labels.get("tot")
                if tot_label:
                    tot_match = re.match(r"([A-Z]+)(\d+)", tot_label)
                    if tot_match and tot_match.group(1) == col:
                        end = int(tot_match.group(2))
                cells = {f"{col}{r}".upper() for r in range(start, end)}
    _KG_CELL_CACHE[sheet] = cells
    return cells


def child_source_cell(wb_formula: Any, sheet: str) -> tuple[Optional[str], Optional[str]]:
    """按结构A/B返回该子表砼体积、总重量合计的目标单元格坐标(vol_cell, tot_cell)。

    结构A（Sheet3、Sheet6~10、Sheet12~39、Sheet41）：砼体积在D20，总重量合计在I10。
    结构B（Sheet5、Sheet42~61）：砼体积在E21，总重量合计在I11。
    其余子表（Sheet1/2/4等未归入两种结构）用标签定位“总重量合计”“砼体积”所在单元格。
    """
    if sheet in STRUCT_A_SHEETS:
        return "D20", "I10"
    if sheet in STRUCT_B_SHEETS:
        return "E21", "I11"
    if sheet in wb_formula.sheetnames:
        labels = _scan_child_labels(wb_formula[sheet])
        tot_label = labels.get("tot")
        vol_cell = find_cell_containing(wb_formula[sheet], "砼体积")
        vol_coord = None
        if vol_cell is not None:
            match = re.match(r"([A-Z]+)(\d+)", vol_cell.coordinate)
            if match:
                # 砼体积标签通常与数值同行、右侧或下一行，按同行下一列取值失败时回退到下一行。
                vol_coord = _next_row_coord(vol_cell.coordinate).upper()
        return vol_coord, (tot_label.upper() if tot_label else None)
    return None, None


def child_cell_value(wb_values: Any, sheet: str, coord: Optional[str]) -> Any:
    if not coord or sheet not in wb_values.sheetnames:
        return None
    try:
        return wb_values[sheet][coord].value
    except Exception:
        return None


def summary_value_matches_child_source(
    wb_values: Any, sheet: str, summary_value: Any, source_coord: Optional[str]
) -> bool:
    """校验汇总单元格计算值是否与子表目标单元格计算值一致（容差1e-6）。"""
    if source_coord is None:
        return False
    source_value = child_cell_value(wb_values, sheet, source_coord)
    if not is_number(source_value):
        return False
    return values_equal(summary_value, source_value)


def sheet_for_summary_row(row: int) -> Optional[str]:
    index = row - 3
    if 0 <= index < len(EXPECTED_CHILD_SHEETS):
        return EXPECTED_CHILD_SHEETS[index]
    return None


def summary_row_for_sheet(sheet_name: str) -> Optional[int]:
    try:
        return EXPECTED_CHILD_SHEETS.index(sheet_name) + 3
    except ValueError:
        return None


def _cell_ref_pattern(row: int, col_letter: str) -> str:
    # 支持带 $ 的绝对/混合引用（如 $G3、G$3、$G$3），列字母大小写已由 normalize_formula 统一为大写。
    return r"\$?" + col_letter + r"\$?" + str(row)


def evaluate_h_formula_equivalent(formula: Any, row: int) -> bool:
    """结构化校验 H 列公式：必须是“G{row} 除以 F{row}”（分子为G、分母为F），可选套 IFERROR。

    不再使用“同一行同时出现 G、F、'/'”这种宽松子串判断——那会把 =F{row}/G{row}（分子分母颠倒）
    误判为等效公式。这里显式解析除法表达式两侧的操作数，确认左侧（分子）引用的是 G{row}、
    右侧（分母）引用的是 F{row}，且两者之间恰好是一个除号（不允许中间夹杂其他单元格或函数）。
    """
    if not is_formula(formula):
        return False
    f = normalize_formula(formula)

    g_ref = _cell_ref_pattern(row, "G")
    f_ref = _cell_ref_pattern(row, "F")

    # 核心除法表达式：G{row}/F{row}，允许两侧各自包一层括号，如 (G3)/(F3)。
    div_pattern = re.compile(r"\(*" + g_ref + r"\)*\s*/\s*\(*" + f_ref + r"\)*")

    # 标准写法：=IFERROR(G{row}/F{row},"")，允许第二个参数是任意字符串/空串，允许空格差异。
    iferror_pattern = re.compile(
        r"^=IFERROR\(\s*" + div_pattern.pattern + r"\s*,\s*\"[^\"]*\"\s*\)$"
    )
    if iferror_pattern.match(f):
        return True

    # 等效写法：去掉最外层 IFERROR 包装后，公式本身就是（且仅是）G{row}/F{row} 这一个除法表达式。
    bare_pattern = re.compile(r"^=\(*" + g_ref + r"\)*\s*/\s*\(*" + f_ref + r"\)*\)*$")
    if bare_pattern.match(f):
        return True

    return False


def load_workbooks(path: str) -> tuple[Any, Any]:
    """加载工作簿。返回 (wb_formula, wb_values)。"""
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    return wb_formula, wb_values


def dimension1_checks(path: str, wb_formula: Any, wb_values: Any) -> list[CheckResult]:
    results: list[CheckResult] = []

    ext = os.path.splitext(path)[1].lower()
    results.append(CheckResult(ext in {".xls", ".xlsx", ".xlsm"}, "D1-1 交付文件为xls、xlsx或.xlsm格式，文件可正常打开", f"当前扩展名：{ext}"))

    return results


def dimension2_checks(wb_formula: Any, wb_values: Any) -> list[CheckResult]:
    results: list[CheckResult] = []
    ws_sum_f = wb_formula["汇总"]
    ws_sum_v = wb_values["汇总"]

    # +5：严格对应细则。
    #   F列（砼体积）：F3:F62 全部为数值或引用公式，不出现空白/文本单位/错误值，
    #                 且不能从“砼重量(t)”列提取数据。
    #   G列（总重量）：G3:G62 全部为数值或引用公式，
    #                 且不能用配筋明细中任意一条“重量(kg)”代替总重量合计。
    f_bad: list[str] = []
    g_bad: list[str] = []
    for row in range(3, 63):
        sheet = sheet_for_summary_row(row)
        vol_coord, tot_coord = child_source_cell(wb_formula, sheet) if sheet is not None else (None, None)
        # ---- F 列：砼体积 ----
        cf = ws_sum_f.cell(row=row, column=6)
        vf = ws_sum_v.cell(row=row, column=6).value
        if is_blank(vf):
            f_bad.append(f"F{row} 空白")
        elif cell_has_formula_error(vf):
            f_bad.append(f"F{row} 错误值={vf!r}")
        elif isinstance(vf, str) and any(u in vf for u in ("m³", "m3")):
            f_bad.append(f"F{row} 文本单位={vf!r}")
        elif not is_number(vf):
            f_bad.append(f"F{row} 非数值={vf!r}")
        elif not (is_formula(cf.value) or is_number(cf.value)):
            f_bad.append(f"F{row} 非数值或公式={cf.value!r}")
        else:
            # 不能从“砼重量(t)”列提取数据；同时数值/取数必须确实来自对应子表的砼体积单元格。
            f_ref = formula_reference(cf.value)
            if sheet is not None and f_ref is not None and f_ref[0] == sheet:
                forbidden = forbidden_tw_cells(wb_formula, sheet)
                if f_ref[1] in forbidden:
                    f_bad.append(f"F{row} 取自砼重量(t)列={sheet}!{f_ref[1]}")
                elif f_ref[1] != vol_coord:
                    # 直接引用能解析，但落点不是该子表的砼体积单元格：仍需按取值比对兜底判断。
                    if not summary_value_matches_child_source(wb_values, sheet, vf, vol_coord):
                        f_bad.append(f"F{row} 取数单元格={sheet}!{f_ref[1]}，与{sheet}!{vol_coord}(砼体积)不一致")
            elif is_formula(cf.value):
                # 公式无法直接解析出(sheet, cell)（如 INDIRECT/INDEX、跨表间接引用等），
                # 不能默认合格：必须比对计算结果与该子表砼体积单元格的值是否一致。
                if sheet is None or not summary_value_matches_child_source(wb_values, sheet, vf, vol_coord):
                    f_bad.append(f"F{row} 公式={cf.value!r} 无法核实取自{sheet}!{vol_coord}(砼体积)，计算值={vf!r}")
        # ---- G 列：总重量合计 ----
        cg = ws_sum_f.cell(row=row, column=7)
        vg = ws_sum_v.cell(row=row, column=7).value
        if is_blank(vg):
            g_bad.append(f"G{row} 空白")
        elif cell_has_formula_error(vg):
            g_bad.append(f"G{row} 错误值={vg!r}")
        elif not is_number(vg):
            g_bad.append(f"G{row} 非数值={vg!r}")
        elif not (is_formula(cg.value) or is_number(cg.value)):
            g_bad.append(f"G{row} 非数值或公式={cg.value!r}")
        else:
            # 不能用配筋明细中任意一条“重量(kg)”代替总重量合计；同时必须确实取自该子表的总重量合计单元格。
            g_ref = formula_reference(cg.value)
            if sheet is not None and g_ref is not None and g_ref[0] == sheet:
                forbidden = forbidden_kg_item_cells(wb_formula, sheet)
                if g_ref[1] in forbidden:
                    g_bad.append(f"G{row} 取自配筋明细单条重量(kg)={sheet}!{g_ref[1]}")
                elif g_ref[1] != tot_coord:
                    if not summary_value_matches_child_source(wb_values, sheet, vg, tot_coord):
                        g_bad.append(f"G{row} 取数单元格={sheet}!{g_ref[1]}，与{sheet}!{tot_coord}(总重量合计)不一致")
            elif is_formula(cg.value):
                # 同上：INDIRECT/INDEX 等无法解析引用的公式，不能默认合格，必须比对取值。
                if sheet is None or not summary_value_matches_child_source(wb_values, sheet, vg, tot_coord):
                    g_bad.append(f"G{row} 公式={cg.value!r} 无法核实取自{sheet}!{tot_coord}(总重量合计)，计算值={vg!r}")
    ok = not f_bad and not g_bad
    detail_parts = []
    if f_bad:
        detail_parts.append("F列异常：" + "; ".join(f_bad[:8]) + (" ..." if len(f_bad) > 8 else ""))
    if g_bad:
        detail_parts.append("G列异常：" + "; ".join(g_bad[:8]) + (" ..." if len(g_bad) > 8 else ""))
    results.append(CheckResult(ok, "+5 F列填砼体积、G列填总重量合计，均为数值或公式，F不取砼重量(t)列、G不取配筋单条重量(kg)", "；".join(detail_parts) or "F3:F62 砼体积、G3:G62 总重量合计均合规", 5))

    # +5：严格对应细则的两条格式要求。
    #   砼体积格式：F3:F62 使用数值格式“0.000”，单元格中不重复显示“m³”文字。
    #   总重量格式：G3:G62 使用数值格式“0.000”，单元格中不重复显示“kg”文字。
    f_fmt_bad: list[str] = []
    g_fmt_bad: list[str] = []
    for row in range(3, 63):
        # F 列：砼体积。
        cf = ws_sum_f.cell(row=row, column=6)
        vf = ws_sum_v.cell(row=row, column=6).value
        if cf.number_format != "0.000":
            f_fmt_bad.append(f"F{row} 格式={cf.number_format!r}")
        if isinstance(vf, str) and "m³" in vf:
            f_fmt_bad.append(f"F{row} 含单位文字={vf!r}")
        # G 列：总重量。
        cg = ws_sum_f.cell(row=row, column=7)
        vg = ws_sum_v.cell(row=row, column=7).value
        if cg.number_format != "0.000":
            g_fmt_bad.append(f"G{row} 格式={cg.number_format!r}")
        if isinstance(vg, str) and "kg" in vg:
            g_fmt_bad.append(f"G{row} 含单位文字={vg!r}")
    ok = not f_fmt_bad and not g_fmt_bad
    fmt_parts = []
    if f_fmt_bad:
        fmt_parts.append("砼体积F列：" + "; ".join(f_fmt_bad[:8]) + (" ..." if len(f_fmt_bad) > 8 else ""))
    if g_fmt_bad:
        fmt_parts.append("总重量G列：" + "; ".join(g_fmt_bad[:8]) + (" ..." if len(g_fmt_bad) > 8 else ""))
    results.append(CheckResult(ok, "+5 F3:F62砼体积、G3:G62总重量均用数值格式0.000，F列不显示“m³”、G列不显示“kg”文字", "；".join(fmt_parts) or "F/G 列格式均为0.000，且无重复单位文字", 5))

    # +5：严格对应细则的两条含量列要求。
    #   含量公式：H3:H62 用公式计算总重量合计除以砼体积，
    #            H3 为 =IFERROR(G3/F3,"") 或等效公式，并向下填充至 H62。
    #   含量格式：H3:H62 用数值格式“0.00”，保持浅橙色填充，
    #            计算结果与对应行 G 列除以 F 列的结果一致。
    # +5：严格对应细则的含量列两条要求（不含填充色约束）。
    #   含量公式：H3:H62 用公式计算总重量合计除以砼体积，
    #            H3 为 =IFERROR(G3/F3,"") 或等效公式，并向下填充至 H62。
    #   含量格式：H3:H62 用数值格式“0.00”，计算结果与对应行 G 列除以 F 列的结果一致。
    h_formula_bad: list[str] = []
    h_format_bad: list[str] = []
    h_result_bad: list[str] = []
    for row in range(3, 63):
        c_formula = ws_sum_f.cell(row=row, column=8)
        h_value = ws_sum_v.cell(row=row, column=8).value
        f_value = ws_sum_v.cell(row=row, column=6).value
        g_value = ws_sum_v.cell(row=row, column=7).value
        # 公式：=IFERROR(G{row}/F{row},"") 或等效，向下填充至 H62。
        if not evaluate_h_formula_equivalent(c_formula.value, row):
            h_formula_bad.append(f"H{row}={c_formula.value!r}")
        # 数值格式“0.00”。
        if c_formula.number_format != "0.00":
            h_format_bad.append(f"H{row} 格式={c_formula.number_format!r}")
        # 结果与 G 列除以 F 列一致。
        # data_only 缓存可能陈旧（工作簿保存前未重算），若已确认 H 公式结构正确
        # （分子G、分母F），则该行结果是否正确本质上等价于“公式是否正确”，
        # 不必再单独依赖可能过期的缓存值来判定结果一致——避免把未重算的正确公式误判为不一致。
        # 仅当公式结构校验未通过（或干脆不是公式）时，才需要用缓存值兜底核验实际结果。
        if is_number(f_value) and is_number(g_value) and float(f_value) != 0:
            expected = float(g_value) / float(f_value)
            formula_confirmed = evaluate_h_formula_equivalent(c_formula.value, row)
            if not formula_confirmed:
                if not is_number(h_value) or abs(float(h_value) - expected) > 1e-6:
                    h_result_bad.append(f"H{row}={h_value!r}，期望≈{expected:.6f}")
            elif is_number(h_value) and abs(float(h_value) - expected) > 1e-6:
                # 公式结构已确认正确，但缓存值仍明显对不上：多半是缓存彻底陈旧/损坏，
                # 而不是公式本身的问题，此处仍需报告以免真正的结果错误被放过。
                h_result_bad.append(f"H{row}={h_value!r}，期望≈{expected:.6f}（公式结构正确，但缓存值不一致，请确认工作簿已重新计算）")
    ok = not h_formula_bad and not h_format_bad and not h_result_bad
    details = []
    if h_formula_bad:
        details.append("公式非=IFERROR(G/F,\"\")或等效：" + "; ".join(h_formula_bad[:8]) + (" ..." if len(h_formula_bad) > 8 else ""))
    if h_format_bad:
        details.append("格式非0.00：" + "; ".join(h_format_bad[:8]) + (" ..." if len(h_format_bad) > 8 else ""))
    if h_result_bad:
        details.append("结果与G/F不一致：" + "; ".join(h_result_bad[:8]) + (" ..." if len(h_result_bad) > 8 else ""))
    results.append(CheckResult(ok, "+5 “汇总”!H3:H62用公式=IFERROR(G/F,\"\")或等效并下拉填充至H62、格式0.00，结果与G列除以F列一致", "；".join(details) or "H3:H62 公式、格式、结果均合规", 5))

    # +5：Sheet3 及同结构子表（Sheet3、Sheet6~Sheet10、Sheet12~Sheet39、Sheet41）。
    #   砼体积：均从各子表 D20 提取，写入“汇总”对应行的 F 列。
    #   总重量合计：均从各子表 I10 提取，写入“汇总”对应行的 G 列。
    #   引用可来自 Excel/WPS，允许 $ 锚定、空格、加不加单引号等等效写法。
    bad_a: list[str] = []
    for sheet in STRUCT_A_SHEETS:
        row = summary_row_for_sheet(sheet)
        if row is None:
            bad_a.append(f"{sheet}: 汇总行缺失")
            continue
        cf = ws_sum_f.cell(row=row, column=6).value
        cg = ws_sum_f.cell(row=row, column=7).value
        # F 列 砼体积 ← 子表 D20。
        if formula_reference(cf) != (sheet, "D20"):
            bad_a.append(f"F{row} 砼体积应取{sheet}!D20，实际={cf!r}")
        # G 列 总重量合计 ← 子表 I10。
        if formula_reference(cg) != (sheet, "I10"):
            bad_a.append(f"G{row} 总重量合计应取{sheet}!I10，实际={cg!r}")
    results.append(CheckResult(not bad_a, "+5 Sheet3及同结构子表：砼体积取各子表D20填入汇总F列、总重量合计取各子表I10填入汇总G列", "; ".join(bad_a[:10]) + (" ..." if len(bad_a) > 10 else ""), 5))

    # +5：Sheet5 及同结构子表（Sheet5、Sheet42~Sheet61）。
    #   砼体积：均从各子表 E21 提取，写入“汇总”对应行的 F 列。
    #   总重量合计：均从各子表 I11 提取，写入“汇总”对应行的 G 列。
    #   引用可来自 Excel/WPS，允许 $ 锚定、空格、加不加单引号等等效写法。
    bad_b: list[str] = []
    for sheet in STRUCT_B_SHEETS:
        row = summary_row_for_sheet(sheet)
        if row is None:
            bad_b.append(f"{sheet}: 汇总行缺失")
            continue
        cf = ws_sum_f.cell(row=row, column=6).value
        cg = ws_sum_f.cell(row=row, column=7).value
        # F 列 砼体积 ← 子表 E21。
        if formula_reference(cf) != (sheet, "E21"):
            bad_b.append(f"F{row} 砼体积应取{sheet}!E21，实际={cf!r}")
        # G 列 总重量合计 ← 子表 I11。
        if formula_reference(cg) != (sheet, "I11"):
            bad_b.append(f"G{row} 总重量合计应取{sheet}!I11，实际={cg!r}")
    results.append(CheckResult(not bad_b, "+5 Sheet5及同结构子表：砼体积取各子表E21填入汇总F列、总重量合计取各子表I11填入汇总G列", "; ".join(bad_b[:10]) + (" ..." if len(bad_b) > 10 else ""), 5))

    # -1：“汇总”!F3:F62 和 !G3:G62 中有任意10个以上单元格为空白（空白数 > 10 触发）。
    #   取 data_only 计算值判断：公式返回空串或单元格无内容均计为空白，与办公软件所见一致。
    # 已按需求删除该扣分规则。

    # -3：“汇总”!F3:G62 使用文本格式，导致数值左上角出现错误提示或不能参与计算。
    # 已按需求删除该扣分规则。

    # -3：出现名为“sheet11”的空白工作表。
    # 已按需求删除该扣分规则。

    # -1：“构件材料数据汇总”没有出现合并单元格的行为。
    # 已按需求删除该扣分规则。

    return results


def print_dimension1(results: list[CheckResult]) -> bool:
    """本地调试用：打印维度1明细并返回是否全部通过。"""
    print("=== 维度1：可用与可修改性 ===")
    all_passed = True
    for result in results:
        icon = "✅" if result.passed else "❌"
        print(f"{icon} {result.label}")
        if result.detail:
            print(f"   {result.detail}")
        if not result.passed:
            all_passed = False
    return all_passed


def scoring_text(label: str) -> str:
    """去掉内部标签开头已有的 +5/-3 等分值，避免最终输出重复分数。"""
    return re.sub(r"^[+-]\d+\s*", "", label).strip()


def print_dimension2(results: list[CheckResult]) -> int:
    """本地调试用：打印维度2明细并返回累计得分。"""
    total = 0
    hit_results = [result for result in results if result.passed]
    for result in hit_results:
        total += result.score

    print(f"维度二：评分结果（{total} / 25）")
    for result in hit_results:
        print(f"{result.score:+d}：{scoring_text(result.label)}")
    return total


SCRIPT_ID = "090"
MAX_SCORE = 25
SUPPORTED_EXTS = {".xlsx", ".xlsm"}


def _find_target_file(dir_path: str) -> str | None:
    """在脚本所在目录中定位待评估的 Excel 文档。

    - 忽略 Office 打开时产生的锁定文件（以 "~$" 开头）；
    - 若同目录存在多个 xls/xlsx/xlsm 文件，优先选择文件名包含 "完成版" 的文件，
      否则取按名称排序后的第一个。
    """
    if not os.path.isdir(dir_path):
        return None
    candidates: list[str] = []
    for name in sorted(os.listdir(dir_path)):
        if name.startswith("~$"):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in SUPPORTED_EXTS:
            candidates.append(os.path.join(dir_path, name))
    if not candidates:
        return None
    for path in candidates:
        if "完成版" in os.path.basename(path):
            return path
    return candidates[0]


def _check_result_to_item(result: CheckResult) -> dict[str, Any]:
    return {
        "rule": scoring_text(result.label),
        "max_delta": result.score,
        "delta": result.score if result.passed else 0,
        "hit": result.passed,
        "detail": "",
    }


def evaluate(dir_path: str) -> dict[str, Any]:
    """统一入口：接收脚本所在目录路径，返回结构化评分字典。

    脚本自行在 dir_path 目录中定位并打开被评估的 Excel 文档；调用方无需再传
    文档文件名。返回结构遵循《脚本接口差异与统一建议.md》§2.2。
    """
    report: dict[str, Any] = {
        "id": SCRIPT_ID,
        "file_name": "",
        "status": "ok",
        "error": None,
        "dim1_pass": False,
        "dim1_reason": "",
        "dim2_items": [],
        "total_score": 0,
        "max_score": MAX_SCORE,
    }

    try:
        target = _find_target_file(dir_path)
        if target is None:
            report["status"] = "error"
            report["error"] = f"目录 {dir_path!r} 中未找到 xlsx或.xlsm 文件"
            return report
        report["file_name"] = os.path.basename(target)

        ext = os.path.splitext(target)[1].lower()
        if ext not in SUPPORTED_EXTS:
            report["dim1_reason"] = f"D1-1 交付文件为xlsx或.xlsm格式，文件可正常打开：当前扩展名 {ext!r}"
            return report

        try:
            wb_formula, wb_values = load_workbooks(target)
        except Exception as exc:
            report["dim1_reason"] = f"D1-1 交付文件为xlsx或.xlsm格式，文件可正常打开：打开失败：{exc}"
            return report

        d1_results = dimension1_checks(target, wb_formula, wb_values)
        if not all(r.passed for r in d1_results):
            fails = [
                r.label + (f"（{r.detail}）" if r.detail else "")
                for r in d1_results
                if not r.passed
            ]
            report["dim1_reason"] = "; ".join(fails)
            return report

        report["dim1_pass"] = True
        d2_results = dimension2_checks(wb_formula, wb_values)
        items = [_check_result_to_item(r) for r in d2_results]
        report["dim2_items"] = items
        report["total_score"] = sum(item["delta"] for item in items)
        return report
    except Exception as exc:  # 兜底：任何未预期异常都归为 status=error
        report["status"] = "error"
        report["error"] = f"{type(exc).__name__}: {exc}"
        return report


if __name__ == "__main__":
    # 仅用于本地调试：默认使用脚本所在目录，也可通过命令行参数指定其他目录。
    debug_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
    print(json.dumps(evaluate(debug_dir), ensure_ascii=False, indent=2))

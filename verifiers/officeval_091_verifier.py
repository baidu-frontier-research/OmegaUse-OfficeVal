# -*- coding: utf-8 -*-
"""
自动评估脚本：对《2026年区域渠道销量统计表_整理完成.xlsx》按“打分细则”自动打分。

对外接口（符合《脚本接口差异与统一建议》§2）：
  def evaluate(dir_path: str) -> dict
    - 入参 dir_path：脚本所在目录路径。脚本自身在该目录里定位并打开被评估的
      .xlsx/.xlsm 文档，调用方无需再传文件名。
    - 返回值：结构化字典，字段见文档 §2.2（id/file_name/status/error/
      dim1_pass/dim1_reason/dim2_items/total_score/max_score）。

评估逻辑（与原实现保持一致）：
  维度1（可用与可修改性）：门槛维度。任一条不满足 -> 直接 0 分，且不再检查维度2。
  维度2（完成度评分细则）：在通过维度1后逐条检查。
      - 加分点：需满足该条内“每一个点”才加分（正分）。
      - 扣分点：只要满足该条内“任意一点”即扣分（负分）。
  最终得分 = 维度2所有命中细则的分数累加（可正可负）。

本地调试用法（不作为主结果输出通道）：
  python officeval_091_verifier.py [脚本所在目录]
  不传参数时默认使用本脚本所在目录。__main__ 仅打印 JSON 便于自测。
"""

import sys
import os
import json

try:
    import openpyxl
except ImportError:
    sys.stderr.write("需要 openpyxl，请先安装： pip install openpyxl\n")
    raise

# 目标字段（用于维度2 +5 细则）
REQUIRED_FIELDS = ["开口料", "保育/育肥料", "蛋禽料", "水禽料", "沉水鱼料", "浮水鱼料"]
CUSTOMER_FIELDS = ["客户", "客户名称"]

# Excel 错误标记
EXCEL_ERRORS = ["#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"]


def norm(v):
    """把单元格值规整成便于比较的字符串（去首尾空白）。"""
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# 维度 1：可用与可修改性（门槛）
# ---------------------------------------------------------------------------
def check_dimension1(path):
    """返回 (通过?, [(条目描述, 是否满足, 说明), ...])"""
    results = []
    ok_all = True

    # 1.1 交付文件为 xlsx 或 .xlsm 格式，文件可正常打开
    ext = os.path.splitext(path)[1].lower()
    fmt_ok = ext in (".xlsx", ".xlsm")
    can_open = False
    wb = None
    open_msg = ""
    if fmt_ok:
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            can_open = True
        except Exception as e:  # 打不开
            open_msg = "打开失败: %s" % e
    else:
        open_msg = "扩展名为 %s，非 xlsx/xlsm" % (ext or "(无)")
    c1 = fmt_ok and can_open
    results.append(("交付文件为xlsx或.xlsm格式，文件可正常打开",
                    c1, open_msg or "格式=%s，打开成功" % ext))
    if not c1:
        return False, results, None

    return ok_all, results, wb


# ---------------------------------------------------------------------------
# 维度 2：完成度评分细则
# ---------------------------------------------------------------------------
def get_sheet(wb, name):
    if name in wb.sheetnames:
        return wb.worksheets[wb.sheetnames.index(name)]
    return None


def rule_two_sheets(wb):
    """+3：文件中仅有两张表格。加分点：需满足‘每一个点’——即“表格”恰为 2 张。

    与办公软件（Excel/WPS）保持一致：在软件里底部标签栏出现的每一张“表”
    都算作一张表格，既包含普通工作表（worksheet），也包含图表工作表
    （chartsheet），并且不论其可见/隐藏/深度隐藏状态，只要存在于文件中即计入。
    细则只要求“仅有两张表格”，因此判定条件为：文件中的表格总数恰为 2。
    """
    ws_names = list(wb.sheetnames)              # 普通工作表（含隐藏/深度隐藏）
    chart_names = [cs.title for cs in getattr(wb, "chartsheets", []) or []]
    n = len(ws_names) + len(chart_names)        # 办公软件标签栏可见的“表格”总数
    ok = (n == 2)
    detail = "表格总数=%d（工作表%d" % (n, len(ws_names))
    if chart_names:
        detail += "，图表表%d" % len(chart_names)
    detail += "） %s" % ("（恰为2）" if ok else "（≠2）")
    return ok, detail


def rule_n41(wb):
    """+1：‘12月’工作表 N41 单元格内容为 2.52。

    与办公软件（Excel/WPS）一致：以 N41 单元格的实际内容为准。
    - 若为数值：其值等于 2.52（用带容差的比较规避二进制浮点误差）即命中；
    - 若为文本：文本内容为“2.52”即命中。
    细则只要求“内容为 2.52”，不对单元格的数字格式/显示样式等额外约束。
    """
    ws = get_sheet(wb, "12月")
    if ws is None:
        return False, "无‘12月’表"
    val = ws["N41"].value
    # 数值与文本 2.52 都算命中
    hit = False
    if isinstance(val, bool):
        hit = False  # 布尔不算数值内容
    elif isinstance(val, (int, float)):
        hit = abs(float(val) - 2.52) < 1e-9
    else:
        hit = norm(val) == "2.52"
    return hit, "N41=%r %s" % (val, "（=2.52）" if hit else "（≠2.52）")


def rule_second_sheet_fields(wb):
    """+5：文档中出现一个 sheet 且名字不是‘12月’：其拥有“客户”或“客户名称”、
    “开口料”、“保育/育肥料”、“蛋禽料”、“水禽料”、“沉水鱼料”、“浮水鱼料”字段，
    且“客户”或“客户名称”列有不同的客户名称内容。

    加分点：需同时满足细则中的每一个点，缺一不可：
      点1：存在一个名字不是‘12月’的 sheet；
      点2：该 sheet 拥有“客户”或“客户名称”字段（表头）；
      点3~8：该 sheet 拥有六类料字段（开口料/保育-育肥料/蛋禽料/水禽料/
             沉水鱼料/浮水鱼料）——即“客户/客户名称”所在表头行同时含这些字段；
      点9：“客户”或“客户名称”列有不同（去重≥2 个）的客户名称内容。

    与办公软件（Excel/WPS）一致：
      - “字段”指办公软件中该表可见的表头单元格文本，因此六类料字段与客户字段
        必须位于同一张表头行（软件里用户看到的就是这一排列标题），而不是散落在
        表格任意角落；
      - 隐藏/深度隐藏的 sheet 在软件标签栏仍会出现且内容真实存在，故一并纳入
        候选（只排除名字为‘12月’的表）。
    """
    # 点1：存在名字不是‘12月’的 sheet（含隐藏/深度隐藏，与软件标签栏一致）
    candidates = [ws for ws in wb.worksheets if ws.title != "12月"]
    if not candidates:
        return False, "不存在名字非‘12月’的其它 sheet"

    detail = []
    for ws in candidates:
        # 定位表头行：同一行内同时能找到“客户/客户名称”字段。
        # 表头文本 -> 列号 映射，用于后续按列取客户名称。
        header_row = None
        header_cells = {}   # 列号 -> 表头文本
        for r in range(1, min(ws.max_row, 30) + 1):
            row_texts = {}
            for c in range(1, ws.max_column + 1):
                t = norm(ws.cell(row=r, column=c).value)
                if t:
                    row_texts[c] = t
            texts = set(row_texts.values())
            if any(cf in texts for cf in CUSTOMER_FIELDS):
                header_row = r
                header_cells = row_texts
                break

        detail = []
        header_texts = set(header_cells.values())

        # 点2：表头含“客户”或“客户名称”字段
        has_customer = header_row is not None and any(
            cf in header_texts for cf in CUSTOMER_FIELDS)
        detail.append("客户字段" + ("✓" if has_customer else "✗"))

        # 点3~8：同一张表头行中含全部六类料字段
        #（软件里用户看到的列标题，须与客户字段同排）
        missing = [f for f in REQUIRED_FIELDS if f not in header_texts]
        has_fields = header_row is not None and len(missing) == 0
        detail.append("6类料字段" + ("✓" if has_fields
                                    else "✗(缺:%s)" % ",".join(missing) if header_row
                                    else "✗(无表头行)"))

        # 点9：“客户/客户名称”列有不同（去重≥2 个）客户名称内容
        distinct_customers = 0
        if has_customer:
            cust_col = None
            for col, txt in header_cells.items():
                if txt in CUSTOMER_FIELDS:
                    cust_col = col
                    break
            if cust_col is not None:
                names = set()
                for r in range(header_row + 1, ws.max_row + 1):
                    t = norm(ws.cell(row=r, column=cust_col).value)
                    if t:
                        names.add(t)
                distinct_customers = len(names)
        has_diff_customers = distinct_customers >= 2
        detail.append("不同客户名称=%d%s" % (
            distinct_customers, "✓" if has_diff_customers else "✗"))

        if has_customer and has_fields and has_diff_customers:
            return True, "表[%s]：%s" % (ws.title, "，".join(detail))
    # 没有任何候选表全部满足
    return False, "无满足全部要求的其它 sheet（最近候选：%s）" % (
        "，".join(detail) if detail else "无")


def rule_deduct_l41(wb):
    """-1：‘12月’L41 公式不是 =SUM(M41:R41)。扣分点：满足即扣。"""
    ws = get_sheet(wb, "12月")
    if ws is None:
        return True, "无‘12月’表"
    f = norm(ws["L41"].value)
    fn = f.replace(" ", "").upper()
    ok_formula = fn in ("=SUM(M41:R41)",)
    hit = not ok_formula  # “不是该公式”则扣分
    return hit, "L41公式=%r %s" % (f, "（不等于=SUM(M41:R41)）" if hit else "（正确）")


def rule_deduct_k57(wb):
    """-1：‘12月’K57 单元格内容不是‘合计’。扣分点：满足（即“不是合计”）则扣分。

    与办公软件（Excel/WPS）一致：以用户在软件里 K57 位置看到的内容为准。
    若 K57 落在合并单元格区域内，软件会把合并区左上角（锚点）单元格的内容显示在
    整个合并区（含 K57），而文件中只有锚点单元格保存该值、K57 本身读到 None。
    因此这里在 K57 为空时回退读取其所在合并区锚点的值，保证判定与软件显示一致。
    细则只要求“内容为‘合计’”，不对字体/对齐/格式等额外约束。
    """
    ws = get_sheet(wb, "12月")
    if ws is None:
        return True, "无‘12月’表"

    cell = ws["K57"]
    v = norm(cell.value)
    # K57 若在合并区内且自身为空，则取合并区锚点（左上角）的显示值
    if v == "":
        for mr in ws.merged_cells.ranges:
            if (mr.min_col <= cell.column <= mr.max_col and
                    mr.min_row <= cell.row <= mr.max_row):
                v = norm(ws.cell(row=mr.min_row, column=mr.min_col).value)
                break

    hit = (v != "合计")
    return hit, "K57=%r %s" % (v, "（不是‘合计’）" if hit else "（是‘合计’）")


def rule_deduct_o41_r41(wb):
    """-1：‘12月’O41 到 R41 单元格至少有一个出现内容。扣分点：满足（有内容）则扣分。

    与办公软件（Excel/WPS）一致：以用户在软件里 O41、P41、Q41、R41 四个位置
    看到的内容为准，只要其中任意一个显示出内容即扣分。
    若某格落在合并单元格区域内、其自身读到 None，软件仍会在该位置显示合并区左上角
    锚点单元格的内容，故在该格为空时回退读取所在合并区锚点的值，避免漏判。
    细则只针对 O41~R41 这四个单元格“是否出现内容”，不涉及其它单元格或格式。
    """
    ws = get_sheet(wb, "12月")
    if ws is None:
        return False, "无‘12月’表"

    def displayed(coord):
        cell = ws[coord]
        v = norm(cell.value)
        if v == "":
            for mr in ws.merged_cells.ranges:
                if (mr.min_col <= cell.column <= mr.max_col and
                        mr.min_row <= cell.row <= mr.max_row):
                    v = norm(ws.cell(row=mr.min_row, column=mr.min_col).value)
                    break
        return v

    filled = []
    for coord in ["O41", "P41", "Q41", "R41"]:
        v = displayed(coord)
        if v != "":
            filled.append("%s=%r" % (coord, v))
    hit = len(filled) > 0
    return hit, ("O41:R41 有内容: " + "；".join(filled)) if hit else "O41:R41 均为空"


# ---------------------------------------------------------------------------
# 主流程（统一接口：evaluate(dir_path) -> dict）
# ---------------------------------------------------------------------------
SCRIPT_ID = "091"

# 维度二加分/扣分细则元数据（保持与原实现一致的规则与顺序）
_PLUS_RULES = [
    (3, "文件中仅有两张表格", rule_two_sheets),
    (1, "“12月”工作表N41单元格内容为2.52", rule_n41),
    (5, ("文档中出现一个sheet且名字不是“12月”：其拥有“客户”或“客户名称”、"
         "“开口料”、“保育/育肥料”、“蛋禽料”、“水禽料”、“沉水鱼料”、"
         "“浮水鱼料”字段，且“客户”或“客户名称”列有不同的客户名称内容"),
     rule_second_sheet_fields),
]
_MINUS_RULES = [
    (-1, "“12月”L41单元格公式不是=SUM(M41:R41)", rule_deduct_l41),
    (-1, "“12月”K57单元格内容不是“合计”", rule_deduct_k57),
    (-1, "“12月”O41到R41单元格至少有一个出现内容", rule_deduct_o41_r41),
]


def _find_target_file(dir_path):
    """在 dir_path 目录中定位待评估的 .xlsx/.xlsm 文档。

    与办公软件的“打开文件”行为一致：只挑用户可见的 Excel 文档，忽略临时文件
    （以 ~$ 开头）与非 xlsx/xlsm 扩展名。若目录里同时存在多个候选，优先选择
    文件名包含“整理完成”的（约定的交付产物命名），其次按文件名排序取第一个。
    """
    if not os.path.isdir(dir_path):
        return None
    candidates = []
    for name in os.listdir(dir_path):
        if name.startswith("~$"):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            candidates.append(name)
    if not candidates:
        return None
    candidates.sort(key=lambda n: (0 if "整理完成" in n else 1, n))
    return os.path.join(dir_path, candidates[0])


def _max_score():
    """维度二满分（仅“加分项”的分值之和）。

    约定：
      - 满分 `max_score`（“总分”）= 所有加分项 max_delta 之和；扣分项不进入满分基数。
      - 实得分 `total_score`（“得分”）= 命中的加分项与扣分项 delta 累加（可正可负）。
    """
    return sum(s for s, _, _ in _PLUS_RULES)


def evaluate(dir_path: str) -> dict:
    """统一入口：接收脚本所在目录路径，自行在该目录里定位并评估 Excel 文档。

    返回结构见《脚本接口差异与统一建议》§2.2。
    """
    result = {
        "id": SCRIPT_ID,
        "file_name": "",
        "status": "ok",
        "error": None,
        "dim1_pass": False,
        "dim1_reason": "",
        "dim2_items": [],
        "total_score": 0,
        "max_score": _max_score(),
    }

    try:
        target = _find_target_file(dir_path)
        if target is None:
            result["status"] = "error"
            result["error"] = "目录中未找到 .xlsx/.xlsm 文档：%s" % dir_path
            result["dim2_items"] = []
            return result
        result["file_name"] = os.path.basename(target)

        # ---- 维度 1 ----
        d1_ok, d1_results, wb = check_dimension1(target)
        result["dim1_pass"] = bool(d1_ok)
        if not d1_ok:
            reasons = [
                "%s（%s）" % (name, msg)
                for name, ok, msg in d1_results if not ok
            ]
            result["dim1_reason"] = "；".join(reasons)
            result["dim2_items"] = []
            result["total_score"] = 0
            return result

        # ---- 维度 2 ----
        dim2_items = []
        total = 0

        for score, desc, fn in _PLUS_RULES:
            hit, detail = fn(wb)
            delta = score if hit else 0
            total += delta
            dim2_items.append({
                "rule": desc,
                "max_delta": score,
                "delta": delta,
                "hit": bool(hit),
                "detail": "",
            })

        for score, desc, fn in _MINUS_RULES:
            hit, detail = fn(wb)
            delta = score if hit else 0
            total += delta
            dim2_items.append({
                "rule": desc,
                "max_delta": score,
                "delta": delta,
                "hit": bool(hit),
                "detail": detail,
            })

        result["dim2_items"] = dim2_items
        result["total_score"] = total
        return result
    except Exception as e:  # 顶层兜底：脚本崩溃 -> status=error
        result["status"] = "error"
        result["error"] = "%s: %s" % (type(e).__name__, e)
        result["dim2_items"] = []
        result["total_score"] = 0
        return result


if __name__ == "__main__":
    # 仅用于本地调试：接收“脚本所在目录路径”，打印 JSON 结果。
    if len(sys.argv) > 1:
        _dir = sys.argv[1]
    else:
        _dir = os.path.dirname(os.path.abspath(__file__))
    print(json.dumps(evaluate(_dir), ensure_ascii=False, indent=2))

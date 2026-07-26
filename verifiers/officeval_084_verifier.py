#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""自动评估“进账单打印_宏按钮版.xlsm”。

评分逻辑：
1. 先检查维度1（可用与可修改性）。任一硬门槛失败则总分为 0，且不再评估维度2。
2. 维度1通过后，逐条检查维度2的得分点和扣分点，命中即累计对应分值。
3. 输出命中的点、未命中的点、证据与最终分。

本脚本主要通过静态方式评估 xlsm：
- zipfile + XML 解析 OOXML 包结构、XLM 宏表、按钮/VML/drawing/打印区域；
- openpyxl 读取工作表、单元格、合并结构、样式、列宽、页面设置。
"""

from __future__ import annotations

import json
import math
import posixpath
import re
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange


SCRIPT_ID = "084"
DEFAULT_FILE = "进账单打印_宏按钮版.xlsm"
SUPPORTED_SUFFIXES = (".xlsm", ".xlsx", ".xls")
DIM2_MAX_SCORE = 18  # P+1(5) + P+2(3) + P+3(5) + P+4(5)
TARGET_SHEET_KEY = "划拨进账单"
DATA_SHEET_KEY = "支票明细"
TICKET_SHEET_NAME = "划拨进账单"
TICKET_BOUNDARY = "B1:W9"
EXPECTED_SOURCE_COLS = "A:I"
# P+4 打印区域内容检测阈值：区域内“有效内容单元格”至少这么多个，
# 且有效内容占整个打印区域单元格数的比例不低于该值，用于排除局部金额区/空白区/整表大面积空白。
MIN_PRINT_AREA_MEANINGFUL = 8
MIN_PRINT_AREA_FILL_RATIO = 0.08

TICKET_FIELD_RANGES = {
    "日期": "F1:I1",
    "缴款单位": "B2:H2",
    "账号": "B3:H3",
    "开户行": "B4:H4",
    "金额大写": "B5:H6",
    "承付情况": "B7:B7",
    "费用类型": "B8:G8",
    "核验种类": "B9:G9",
    "对应序号": "W2:W2",
    "金额小写": "N6:T6",
}
CORE_TICKET_RANGES = list(TICKET_FIELD_RANGES.values())
DYNAMIC_TICKET_FIELDS = {"日期", "缴款单位", "账号", "开户行", "金额大写", "费用类型", "核验种类", "对应序号", "金额小写"}
RIGHT_SIDE_STATIC_RANGES = ["I2:T4", "H8:T9"]
# 右侧应保持不变的字段及其固定文本关键字（细则：中部收款单位、收款账号、银行名称、用途说明）。
# 逐项对固定标签文本做“不变”检测，任一字段标签缺失即视为右侧内容被篡改/删除。
RIGHT_SIDE_FIELD_KEYWORDS = {
    "收款单位": ["收款单位", "收款人", "收 款 单 位", "中部收款单位"],
    "收款账号": ["收款账号", "收款帐号", "账号", "帐号"],
    "银行名称": ["银行名称", "开户银行", "开户行", "银行"],
    "用途说明": ["用途说明", "用途", "摘要", "说明"],
}
COLOR_REGION_RULES = {
    "yellow": ["B2:H7", "B8:G9"],
    "green": ["I2:W4", "U2:W4"],
    "blue": ["I5:W6", "H8:W8"],
}

ALLOWED_SHEET_KEYWORDS = ["支票明细", "填写", "划拨进账单", "打印宏"]
HEADER_KEYWORDS = [
    ["缴款单位", "单位", "付款单位", "客户"],
    ["账号", "结算账号", "付款账号", "银行账号"],
    ["金额", "发生额", "付款金额"],
    ["日期", "缴费日期", "出票日期"],
    ["摘要", "用途", "备注"],
]

INVALID_ROW_PROMPTS = ["请选择数据行", "请选择要打印的记录", "请选择", "数据行", "记录", "无效", "空白"]
INVALID_ROW_GUARDS = ["IF", "ISBLANK", "COUNTA", "COUNT", "ISERROR", "ALERT", "MESSAGE", "STOP", "RETURN"]
PRINT_KEYWORDS = ["PRINT", "PRINT.PREVIEW", "PAGE.SETUP", "打印", "预览"]
HIGH_RISK_MACRO_KEYWORDS = ["EXEC", "SHELL", "CALL", "REGISTER", "DDE", "DELETE"]

UPPER_ALLOWED = set("零壹贰叁肆伍陆柒捌玖拾佰仟万亿元角分整正人民币圆 ")
LOWER_CN_NUMS = set("一二三四五六七八九十百千万")


@dataclass
class RuleResult:
    rule_id: str
    title: str
    points: int
    hit: bool
    message: str
    evidence: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.rule_id,
            "title": self.title,
            "points": self.points,
            "hit": self.hit,
            "message": self.message,
            "evidence": self.evidence,
        }


@dataclass
class ObjInfo:
    kind: str
    sheet: str | None
    ref: str | None
    text: str = ""
    macro: str = ""
    source: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "sheet": self.sheet,
            "ref": self.ref,
            "text": self.text,
            "macro": self.macro,
            "source": self.source,
        }


@dataclass
class WorkbookFeatures:
    path: Path
    exists: bool = False
    suffix_ok: bool = False
    zip_ok: bool = False
    package_error: str = ""
    entries: set[str] = field(default_factory=set)
    workbook_loaded: bool = False
    workbook_error: str = ""
    wb: Any = None
    sheet_names: list[str] = field(default_factory=list)
    sheet_states: dict[str, str] = field(default_factory=dict)
    sheet_paths: dict[str, str] = field(default_factory=dict)
    defined_names: list[dict[str, Any]] = field(default_factory=list)
    print_areas: dict[str, list[str]] = field(default_factory=dict)
    xlm_formulas: list[str] = field(default_factory=list)
    objects: list[ObjInfo] = field(default_factory=list)
    shared_text: str = ""
    wb_values: Any = None
    values_loaded: bool = False
    first_data_sheet: str | None = None
    data_header_row: int | None = None
    data_header_cols: dict[str, int] = field(default_factory=dict)

    @property
    def macro_text(self) -> str:
        parts: list[str] = []
        parts.extend(self.xlm_formulas)
        parts.extend(str(d.get("name", "")) + " " + str(d.get("text", "")) for d in self.defined_names)
        parts.extend(o.text + " " + o.macro for o in self.objects)
        return "\n".join(parts)


@dataclass
class EvaluationReport:
    file: str
    dimension1: list[RuleResult]
    positives: list[RuleResult]
    penalties: list[RuleResult]
    dimension1_pass: bool
    raw_score: int
    final_score: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "file": self.file,
            "dimension1_pass": self.dimension1_pass,
            "raw_score": self.raw_score,
            "final_score": self.final_score,
            "dimension1": [r.to_dict() for r in self.dimension1],
            "positives": [r.to_dict() for r in self.positives],
            "penalties": [r.to_dict() for r in self.penalties],
        }


# ------------------------- 基础工具 -------------------------

def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag.split(":")[-1]


def safe_xml_from_bytes(data: bytes) -> ET.Element | None:
    try:
        return ET.fromstring(data)
    except ET.ParseError:
        try:
            return ET.fromstring(data.decode("utf-8", errors="ignore"))
        except Exception:
            return None


def read_zip_text(zf: zipfile.ZipFile, name: str) -> str:
    return zf.read(name).decode("utf-8", errors="ignore")


def norm_target(base_dir: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target)).replace("\\", "/")


def col_to_int(col: str) -> int:
    return column_index_from_string(col.replace("$", ""))


def int_to_col(idx: int) -> str:
    return get_column_letter(idx)


def normalize_ref(ref: str | None) -> str | None:
    if not ref:
        return None
    ref = str(ref).strip()
    if "!" in ref:
        ref = ref.split("!", 1)[1]
    ref = ref.replace("'", "").replace("$", "")
    if ":" not in ref:
        ref = f"{ref}:{ref}"
    return ref.upper()


def split_sheet_ref(text: str | None) -> tuple[str | None, str | None]:
    if not text:
        return None, None
    text = str(text).strip()
    if "," in text:
        text = text.split(",", 1)[0]
    if "!" in text:
        sheet, ref = text.rsplit("!", 1)
        return sheet.strip("'"), normalize_ref(ref)
    return None, normalize_ref(text)


def range_obj(ref: str | None) -> CellRange | None:
    ref = normalize_ref(ref)
    if not ref:
        return None
    try:
        return CellRange(ref)
    except Exception:
        return None


def range_contains(outer: str, inner: str) -> bool:
    a = range_obj(outer)
    b = range_obj(inner)
    if not a or not b:
        return False
    return a.min_col <= b.min_col and a.max_col >= b.max_col and a.min_row <= b.min_row and a.max_row >= b.max_row


def range_exceeds_boundary(ref: str | None, boundary: str = TICKET_BOUNDARY) -> bool:
    """判断有效区域是否有任一边越过给定边界。"""
    return bool(range_obj(ref) and not range_contains(boundary, str(ref)))


def normalized_valid_ranges(areas: list[str]) -> tuple[list[str], list[str]]:
    """规范化区域列表，同时保留无法解析的区域作为证据。"""
    valid: list[str] = []
    invalid: list[str] = []
    for area in areas:
        ref = normalize_ref(area)
        if not ref or not range_obj(ref):
            invalid.append(str(area))
            continue
        if ref not in valid:
            valid.append(ref)
    return valid, invalid


def ranges_intersect(a_ref: str | None, b_ref: str | None) -> bool:
    a = range_obj(a_ref)
    b = range_obj(b_ref)
    if not a or not b:
        return False
    return not (a.max_col < b.min_col or b.max_col < a.min_col or a.max_row < b.min_row or b.max_row < a.min_row)


def range_area(ref: str | None) -> int:
    r = range_obj(ref)
    if not r:
        return 0
    return (r.max_col - r.min_col + 1) * (r.max_row - r.min_row + 1)


def intersection_area(a_ref: str | None, b_ref: str | None) -> int:
    a = range_obj(a_ref)
    b = range_obj(b_ref)
    if not a or not b or not ranges_intersect(a_ref, b_ref):
        return 0
    min_col = max(a.min_col, b.min_col)
    max_col = min(a.max_col, b.max_col)
    min_row = max(a.min_row, b.min_row)
    max_row = min(a.max_row, b.max_row)
    return max(0, max_col - min_col + 1) * max(0, max_row - min_row + 1)


def coords_to_ref(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    min_col = max(1, min_col)
    min_row = max(1, min_row)
    max_col = max(min_col, max_col)
    max_row = max(min_row, max_row)
    return f"{int_to_col(min_col)}{min_row}:{int_to_col(max_col)}{max_row}"


def text_of(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def range_values_text(ws: Any, ref: str) -> str:
    pieces: list[str] = []
    cr = range_obj(ref)
    if not cr:
        return ""
    for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
        for cell in row:
            value = text_of(cell.value)
            if value:
                pieces.append(value)
    return " ".join(pieces)


def display_text(f: WorkbookFeatures, sheet_name: str | None, ref: str) -> str:
    """优先返回缓存值；没有缓存值时再返回公式文本。"""
    if not sheet_name:
        return ""
    if f.values_loaded and f.wb_values and sheet_name in f.wb_values.sheetnames:
        txt = range_values_text(f.wb_values[sheet_name], ref)
        if txt:
            return txt
    if f.workbook_loaded and sheet_name in f.wb.sheetnames:
        return range_values_text(f.wb[sheet_name], ref)
    return ""


def cells_have_formula_or_value(ws: Any, ref: str) -> tuple[int, int, int]:
    value_count = formula_count = total = 0
    cr = range_obj(ref)
    if not cr:
        return 0, 0, 0
    for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
        for cell in row:
            total += 1
            if cell.value not in (None, ""):
                value_count += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    return value_count, formula_count, total


def print_area_content_stats(ws: Any, ref: str) -> dict[str, Any]:
    """统计打印区域内的“有效内容/装饰”密度，用于排除局部/空白/整表大面积空白。

    - meaningful：既包括非空单元格（值/公式）也包括带边框或非默认填充的“票据框结构”单元格；
    - total：打印区域包含的单元格总数；
    - meaningful_ratio：meaningful / total，占比过低即视为大面积空白。
    """
    cr = range_obj(ref)
    if not cr:
        return {"total": 0, "values": 0, "formulas": 0, "meaningful": 0, "meaningful_ratio": 0.0}
    total = values = formulas = meaningful = 0
    for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
        for cell in row:
            total += 1
            has_content = False
            if cell.value not in (None, ""):
                values += 1
                has_content = True
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
            sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
            if any(side and side.style for side in sides):
                has_content = True
            elif cell.fill and cell.fill.fill_type not in (None, "none"):
                rgb = color_rgb(cell)
                # 近白背景不算“票据结构”，避免整表底色被误判为有内容。
                if rgb and (max(rgb) - min(rgb) > 20 or min(rgb) < 245):
                    has_content = True
            if has_content:
                meaningful += 1
    ratio = (meaningful / total) if total else 0.0
    return {"total": total, "values": values, "formulas": formulas, "meaningful": meaningful, "meaningful_ratio": round(ratio, 4)}


def collect_range_details(ws: Any, refs: list[str] | dict[str, str]) -> dict[str, dict[str, Any]]:
    items = refs.items() if isinstance(refs, dict) else ((ref, ref) for ref in refs)
    details: dict[str, dict[str, Any]] = {}
    for label, ref in items:
        cr = range_obj(ref)
        detail: dict[str, Any] = {"ref": ref, "values": 0, "formulas": 0, "merged": merged_contains_or_intersects(ws, ref), "bordered": 0, "filled": 0, "formula_samples": [], "value_samples": []}
        if cr:
            for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
                for cell in row:
                    value = text_of(cell.value)
                    if value:
                        detail["values"] += 1
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            detail["formulas"] += 1
                            if len(detail["formula_samples"]) < 3:
                                detail["formula_samples"].append(cell.value)
                        elif len(detail["value_samples"]) < 3:
                            detail["value_samples"].append(value)
                    sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
                    if any(side and side.style for side in sides):
                        detail["bordered"] += 1
                    if cell.fill and cell.fill.fill_type not in (None, "none"):
                        detail["filled"] += 1
        details[str(label)] = detail
    return details


def formula_references_data_source(formula: str, data_sheet_name: str | None = DATA_SHEET_KEY, expected_cols: str = EXPECTED_SOURCE_COLS) -> bool:
    text = formula.upper().replace("'", "").replace("$", "")
    sheet = (data_sheet_name or DATA_SHEET_KEY).upper()
    if sheet not in text:
        return False
    if expected_cols in text or "A:O" in text:
        return True
    return bool(re.search(rf"{re.escape(sheet)}![A-I](?:\d+|:[A-I])", text))


def formula_has_current_row_logic(formula: str) -> bool:
    text = formula.upper()
    return any(k in text for k in ["ROW(", "CELL(", "INDEX(", "OFFSET(", "MATCH(", "INDIRECT(", "LOOKUP(", "VLOOKUP(", "XLOOKUP("])


def extract_range_refs_from_text(text: str) -> list[str]:
    refs: list[str] = []
    for match in re.findall(r"\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+", text.upper()):
        ref = normalize_ref(match.replace(" ", ""))
        if ref and range_obj(ref) and ref not in refs:
            refs.append(ref)
    return refs


def range_has_content_or_style_outside(ws: Any, inner_ref: str, outer_ref: str) -> bool:
    inner = range_obj(inner_ref)
    outer = range_obj(outer_ref)
    if not inner or not outer:
        return False
    for row in ws.iter_rows(min_row=inner.min_row, max_row=inner.max_row, min_col=inner.min_col, max_col=inner.max_col):
        for cell in row:
            if outer.min_col <= cell.column <= outer.max_col and outer.min_row <= cell.row <= outer.max_row:
                continue
            if cell.value not in (None, ""):
                return True
            if cell.fill and cell.fill.fill_type not in (None, "none"):
                rgb = color_rgb(cell)
                # 近白背景色（如模板外边距 F8FAFC）不视为票据内容超界；有明显饱和度或非浅色才算可见超界。
                if rgb and (max(rgb) - min(rgb) > 20 or min(rgb) < 245):
                    return True
            sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
            if any(side and side.style for side in sides):
                return True
            if merged_contains_or_intersects(ws, cell.coordinate):
                return True
    return False


def merged_intersects(ws: Any, ref: str) -> bool:
    for mr in ws.merged_cells.ranges:
        if ranges_intersect(str(mr), ref):
            return True
    return False


def merged_contains_or_intersects(ws: Any, ref: str) -> bool:
    target = range_obj(ref)
    if not target:
        return False
    for mr in ws.merged_cells.ranges:
        mref = str(mr)
        if range_contains(mref, ref) or ranges_intersect(mref, ref):
            return True
    return False


# ------------------------- 特征抽取 -------------------------

def extract_features(path: Path) -> WorkbookFeatures:
    f = WorkbookFeatures(path=path)
    f.exists = path.exists()
    f.suffix_ok = path.suffix.lower() in (".xlsm", ".xlsx")
    if not f.exists:
        f.package_error = "文件不存在"
        return f

    try:
        with zipfile.ZipFile(path) as zf:
            f.zip_ok = True
            f.entries = set(zf.namelist())
            parse_workbook_package(zf, f)
            parse_shared_strings(zf, f)
            parse_xlm_formulas(zf, f)
            parse_objects(zf, f)
    except Exception as exc:
        f.package_error = f"无法作为 xlsm/zip 打开：{exc}"

    try:
        f.wb = load_workbook(path, keep_vba=True, data_only=False)
        f.workbook_loaded = True
        try:
            f.wb_values = load_workbook(path, keep_vba=False, data_only=True)
            f.values_loaded = True
        except Exception:
            f.values_loaded = False
        if not f.sheet_names:
            f.sheet_names = list(f.wb.sheetnames)
            f.sheet_states = {name: f.wb[name].sheet_state for name in f.sheet_names}
        infer_data_sheet_and_headers(f)
    except Exception as exc:
        f.workbook_error = f"openpyxl 加载失败：{exc}"

    return f


def parse_workbook_package(zf: zipfile.ZipFile, f: WorkbookFeatures) -> None:
    if "xl/workbook.xml" not in f.entries:
        return

    rels = parse_rels(zf, "xl/_rels/workbook.xml.rels", "xl")
    root = safe_xml_from_bytes(zf.read("xl/workbook.xml"))
    if root is None:
        return

    sheet_rids: dict[str, str] = {}
    for elem in root.iter():
        if local_name(elem.tag) == "sheet":
            name = elem.attrib.get("name", "")
            if not name:
                continue
            f.sheet_names.append(name)
            f.sheet_states[name] = elem.attrib.get("state", "visible")
            rid = ""
            for key, value in elem.attrib.items():
                if local_name(key) == "id":
                    rid = value
                    break
            if rid:
                sheet_rids[name] = rid
                target = rels.get(rid, {}).get("target")
                if target:
                    f.sheet_paths[name] = target

    for elem in root.iter():
        if local_name(elem.tag) != "definedName":
            continue
        item = dict(elem.attrib)
        item["text"] = (elem.text or "").strip()
        f.defined_names.append(item)
        if item.get("name") == "_xlnm.Print_Area":
            sheet_name = None
            if "localSheetId" in item:
                try:
                    sheet_name = f.sheet_names[int(item["localSheetId"])]
                except Exception:
                    sheet_name = None
            ref_sheet, ref = split_sheet_ref(item.get("text"))
            sheet_name = sheet_name or ref_sheet
            if sheet_name and ref:
                f.print_areas.setdefault(sheet_name, []).append(ref)


def parse_rels(zf: zipfile.ZipFile, rels_path: str, base_dir: str) -> dict[str, dict[str, str]]:
    if rels_path not in zf.namelist():
        return {}
    root = safe_xml_from_bytes(zf.read(rels_path))
    if root is None:
        return {}
    result: dict[str, dict[str, str]] = {}
    for elem in root.iter():
        if local_name(elem.tag) != "Relationship":
            continue
        rid = elem.attrib.get("Id")
        target = elem.attrib.get("Target", "")
        typ = elem.attrib.get("Type", "")
        if not rid:
            continue
        result[rid] = {"target": norm_target(base_dir, target), "type": typ}
    return result


def parse_shared_strings(zf: zipfile.ZipFile, f: WorkbookFeatures) -> None:
    if "xl/sharedStrings.xml" not in f.entries:
        return
    root = safe_xml_from_bytes(zf.read("xl/sharedStrings.xml"))
    if root is None:
        return
    texts: list[str] = []
    for elem in root.iter():
        if local_name(elem.tag) == "t" and elem.text:
            texts.append(elem.text)
    f.shared_text = "\n".join(texts)


def parse_xlm_formulas(zf: zipfile.ZipFile, f: WorkbookFeatures) -> None:
    for name in sorted(n for n in f.entries if n.startswith("xl/macrosheets/") and n.endswith(".xml")):
        root = safe_xml_from_bytes(zf.read(name))
        if root is None:
            continue
        for elem in root.iter():
            if local_name(elem.tag) == "f" and elem.text:
                f.xlm_formulas.append(elem.text.strip())


def sheet_rels_path(sheet_path: str) -> str:
    folder = posixpath.dirname(sheet_path)
    base = posixpath.basename(sheet_path)
    return f"{folder}/_rels/{base}.rels"


def reverse_sheet_object_map(zf: zipfile.ZipFile, f: WorkbookFeatures) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for sheet_name, sheet_path in f.sheet_paths.items():
        rels_path = sheet_rels_path(sheet_path)
        rels = parse_rels(zf, rels_path, posixpath.dirname(sheet_path))
        for rel in rels.values():
            target = rel.get("target", "")
            typ = rel.get("type", "")
            if "drawing" in typ or "vmlDrawing" in typ or target.startswith("xl/drawings/"):
                mapping[target] = sheet_name
    return mapping


def parse_objects(zf: zipfile.ZipFile, f: WorkbookFeatures) -> None:
    obj_to_sheet = reverse_sheet_object_map(zf, f)

    # VML：常见的表单按钮存放在 vmlDrawing*.vml 中。
    for name in sorted(n for n in f.entries if n.startswith("xl/drawings/vmlDrawing") and n.endswith(".vml")):
        text = read_zip_text(zf, name)
        sheet = obj_to_sheet.get(name)
        if sheet is None and f.sheet_names:
            # 无法从 rels 精确映射时，按钮通常位于第一个数据工作表。
            sheet = f.sheet_names[0]
        button_like = "ObjectType=\"Button\"" in text or "ObjectType='Button'" in text
        anchors = re.findall(r"<[^>]*Anchor[^>]*>(.*?)</[^>]*Anchor>", text, flags=re.I | re.S)
        macros = re.findall(r"<[^>]*FmlaMacro[^>]*>(.*?)</[^>]*FmlaMacro>", text, flags=re.I | re.S)
        shape_texts = re.findall(r"<[^>]*TextBox[^>]*>(.*?)</[^>]*TextBox>", text, flags=re.I | re.S)
        plain_shape_text = " ".join(re.sub(r"<[^>]+>", " ", t) for t in shape_texts)
        if button_like or anchors or macros:
            if anchors:
                for idx, anchor in enumerate(anchors):
                    ref = vml_anchor_to_ref(anchor)
                    f.objects.append(ObjInfo(
                        kind="button" if button_like else "vml_object",
                        sheet=sheet,
                        ref=ref,
                        text=plain_shape_text,
                        macro=macros[idx] if idx < len(macros) else "",
                        source=name,
                    ))
            else:
                f.objects.append(ObjInfo(
                    kind="button" if button_like else "vml_object",
                    sheet=sheet,
                    ref=None,
                    text=plain_shape_text,
                    macro=" ".join(macros),
                    source=name,
                ))

    # DrawingML：图片、形状等对象及其锚点。
    for name in sorted(n for n in f.entries if n.startswith("xl/drawings/drawing") and n.endswith(".xml")):
        root = safe_xml_from_bytes(zf.read(name))
        if root is None:
            continue
        sheet = obj_to_sheet.get(name)
        for anchor in root.iter():
            if local_name(anchor.tag) not in {"twoCellAnchor", "oneCellAnchor", "absoluteAnchor"}:
                continue
            kind = "shape"
            text_pieces: list[str] = []
            for child in anchor.iter():
                lname = local_name(child.tag)
                if lname == "pic":
                    kind = "picture"
                elif lname == "sp" and kind != "picture":
                    kind = "shape"
                elif lname == "t" and child.text:
                    text_pieces.append(child.text)
            ref = drawing_anchor_to_ref(anchor)
            f.objects.append(ObjInfo(kind=kind, sheet=sheet, ref=ref, text=" ".join(text_pieces), source=name))

    # ctrlProps 作为控件存在性的补充证据。
    for name in sorted(n for n in f.entries if n.startswith("xl/ctrlProps/") and n.endswith(".xml")):
        f.objects.append(ObjInfo(kind="control_prop", sheet=None, ref=None, source=name))


def vml_anchor_to_ref(anchor: str) -> str | None:
    nums = [int(x) for x in re.findall(r"-?\d+", anchor)]
    if len(nums) < 8:
        return None
    left_col, _left_dx, top_row, _top_dy, right_col, _right_dx, bottom_row, _bottom_dy = nums[:8]
    return coords_to_ref(left_col + 1, top_row + 1, right_col + 1, bottom_row + 1)


def drawing_anchor_to_ref(anchor_elem: ET.Element) -> str | None:
    from_col = from_row = to_col = to_row = None
    current_part = None
    for elem in anchor_elem.iter():
        lname = local_name(elem.tag)
        if lname == "from":
            current_part = "from"
        elif lname == "to":
            current_part = "to"
        elif lname == "col" and elem.text is not None:
            if current_part == "from":
                from_col = int(elem.text) + 1
            elif current_part == "to":
                to_col = int(elem.text) + 1
        elif lname == "row" and elem.text is not None:
            if current_part == "from":
                from_row = int(elem.text) + 1
            elif current_part == "to":
                to_row = int(elem.text) + 1
    if from_col and from_row:
        return coords_to_ref(from_col, from_row, to_col or from_col, to_row or from_row)
    return None


def infer_data_sheet_and_headers(f: WorkbookFeatures) -> None:
    if not f.workbook_loaded:
        return
    if DATA_SHEET_KEY in f.wb.sheetnames:
        f.first_data_sheet = DATA_SHEET_KEY
    else:
        for name in f.wb.sheetnames:
            if "打印宏" not in name and "划拨进账单" not in name:
                f.first_data_sheet = name
                break
        if f.first_data_sheet is None and f.wb.sheetnames:
            f.first_data_sheet = f.wb.sheetnames[0]

    if not f.first_data_sheet:
        return
    ws = f.wb[f.first_data_sheet]
    best_row = None
    best_score = -1
    best_cols: dict[str, int] = {}
    max_col = min(ws.max_column or 1, 30)
    max_row = min(ws.max_row or 1, 30)
    for row in range(1, max_row + 1):
        row_texts = [text_of(ws.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        joined = " ".join(row_texts)
        score = 0
        cols: dict[str, int] = {}
        for group in HEADER_KEYWORDS:
            for kw in group:
                if kw in joined:
                    score += 1
                    for col, txt in enumerate(row_texts, start=1):
                        if kw in txt:
                            cols[group[0]] = col
                            break
                    break
        if score > best_score:
            best_score = score
            best_row = row
            best_cols = cols
    if best_score >= 2:
        f.data_header_row = best_row
        f.data_header_cols = best_cols


# ------------------------- 工作簿检测辅助 -------------------------

def get_ticket_sheet_name(f: WorkbookFeatures) -> str | None:
    for name in f.sheet_names:
        if TARGET_SHEET_KEY in name:
            return name
    if f.workbook_loaded:
        for name in f.wb.sheetnames:
            if TARGET_SHEET_KEY in name:
                return name
    return None


def print_area_for(f: WorkbookFeatures, sheet_name: str) -> list[str]:
    areas = list(f.print_areas.get(sheet_name, []))
    if f.workbook_loaded and sheet_name in f.wb.sheetnames:
        ws = f.wb[sheet_name]
        pa = getattr(ws, "print_area", None)
        if pa:
            if isinstance(pa, str):
                _, ref = split_sheet_ref(pa)
                if ref and ref not in areas:
                    areas.append(ref)
            else:
                for item in pa:
                    _, ref = split_sheet_ref(str(item))
                    if ref and ref not in areas:
                        areas.append(ref)
    return areas


def has_xlm_entry(f: WorkbookFeatures) -> bool:
    has_macro_sheet = any(n.startswith("xl/macrosheets/") and n.endswith(".xml") for n in f.entries)
    for dn in f.defined_names:
        name = str(dn.get("name", ""))
        target = str(dn.get("text", ""))
        if (dn.get("xlm") == "1" or dn.get("function") == "1") and ("打印宏" in target or "按钮" in name or "单击" in name):
            return has_macro_sheet
    return False


def has_print_button(f: WorkbookFeatures) -> bool:
    """检测第一个数据工作表是否存在可点击的“打印进账单”按钮或控件。

    细则要求按钮位于第一个数据工作表且可点击，因此这里做三重校验：
    1. 至少存在一个 button/vml_object/control_prop 对象锚定在 ``first_data_sheet``。
       ``control_prop`` 自身通常没有 sheet/ref，此时必须借助同一 sheet 上有 macro 绑定或按钮语义的
       VML/Drawing 对象来印证，避免“缺少按钮的文件因存在孤立 ctrlProps 而逃过扣分”。
    2. 至少存在一处“可点击性”证据：VML 按钮通常绑定 FmlaMacro（obj.macro 非空）或 kind 明确为 ``button``；
       也允许由定义名/XLM 宏名体现（如 ``打印宏``、``单击`` 等）作为兼容性兜底。
    3. 存在“打印进账单”语义：对象文本/宏名/定义名中包含“打印”“进账单”“打印进账单”等关键字。
    """
    data_sheet = f.first_data_sheet
    if not data_sheet:
        return False

    print_semantic_kws = ["打印进账单", "打印票据", "打印", "进账单", "PrintTicket", "PrintReceipt"]
    button_kws = ["按钮", "Button", "button", "单击", "OnClick", "点击"]

    macro_text = f.macro_text
    defined_name_text = " ".join(
        str(d.get("name", "")) + " " + str(d.get("text", "")) for d in f.defined_names
    )

    # 收集锚定在第一个数据工作表上的按钮/形状/控件对象
    on_sheet_button_like = [
        o for o in f.objects
        if o.kind in {"button", "vml_object"} and o.sheet == data_sheet
    ]
    # ctrlProps 常缺 sheet；仅当它与同 sheet 的按钮式对象共存时才承认它是佐证
    ctrl_props = [o for o in f.objects if o.kind == "control_prop"]

    if not on_sheet_button_like:
        # 没有任何锚定在数据工作表上的按钮式对象，则视为“数据工作表无可点击按钮”
        return False

    # 语义命中：对象自身文本/宏 或 定义名/共享字符串中出现“打印进账单”等关键字
    object_text_first_sheet = " ".join(o.text + " " + o.macro + " " + o.source for o in on_sheet_button_like)
    semantic_pool = object_text_first_sheet + " " + defined_name_text + " " + f.shared_text
    print_semantic_ok = any(k in semantic_pool for k in print_semantic_kws)
    if not print_semantic_ok:
        return False

    # 可点击性：VML 按钮 kind 为 button，或对象带 FmlaMacro 绑定，或与 XLM 宏/ctrlProps 有关联
    clickable = any(o.kind == "button" for o in on_sheet_button_like)
    if not clickable:
        clickable = any(bool(o.macro.strip()) for o in on_sheet_button_like)
    if not clickable and has_xlm_entry(f) and ctrl_props:
        # 存在 XLM 打印宏 + 数据工作表上的形状 + ctrlProps 三者共存，也视为可点击控件
        clickable = any(k in defined_name_text for k in button_kws + ["打印宏"])

    return clickable


def ticket_sheet_content_stats(f: WorkbookFeatures, sheet_name: str) -> dict[str, Any]:
    stats = {"values": 0, "formulas": 0, "total": 0, "merged": 0, "bordered": 0, "filled": 0}
    if not f.workbook_loaded or sheet_name not in f.wb.sheetnames:
        return stats
    ws = f.wb[sheet_name]
    cr = range_obj(TICKET_BOUNDARY)
    if not cr:
        return stats
    for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
        for cell in row:
            stats["total"] += 1
            if cell.value not in (None, ""):
                stats["values"] += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    stats["formulas"] += 1
            if cell.fill and cell.fill.fill_type not in (None, "none"):
                stats["filled"] += 1
            sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
            if any(side and side.style for side in sides):
                stats["bordered"] += 1
    stats["merged"] = sum(1 for mr in ws.merged_cells.ranges if ranges_intersect(str(mr), TICKET_BOUNDARY))
    return stats


def object_cover_ratio(objects: list[ObjInfo], sheet: str | None, target_ref: str) -> float:
    target_area = range_area(target_ref)
    if target_area == 0:
        return 0.0
    covered = 0
    for obj in objects:
        if sheet and obj.sheet and obj.sheet != sheet:
            continue
        if obj.ref and ranges_intersect(obj.ref, target_ref):
            covered += intersection_area(obj.ref, target_ref)
    return min(1.0, covered / target_area)


def worksheet_has_protection(f: WorkbookFeatures, sheet_name: str | None) -> bool:
    if not f.workbook_loaded or not sheet_name or sheet_name not in f.wb.sheetnames:
        return False
    return bool(getattr(f.wb[sheet_name].protection, "sheet", False))


# ------------------------- 维度1 -------------------------

def evaluate_dimension1(f: WorkbookFeatures) -> list[RuleResult]:
    results: list[RuleResult] = []

    file_ok = f.exists and f.suffix_ok and f.zip_ok and "[Content_Types].xml" in f.entries and "xl/workbook.xml" in f.entries and f.workbook_loaded
    results.append(RuleResult(
        "D1.1",
        "交付文件为 .xlsm 或 .xlsx 格式，文件可正常打开",
        0,
        file_ok,
        "文件格式和可打开性通过" if file_ok else "文件不是有效 .xlsm/.xlsx，或无法被 zip/openpyxl 正常打开",
        {"exists": f.exists, "suffix_ok": f.suffix_ok, "zip_ok": f.zip_ok, "workbook_loaded": f.workbook_loaded, "error": f.package_error or f.workbook_error},
    ))

    return results


# ------------------------- 维度2得分点 -------------------------

def evaluate_positive_points(f: WorkbookFeatures) -> list[RuleResult]:
    results: list[RuleResult] = []
    macro_text = f.macro_text.upper()
    all_text = (f.macro_text + "\n" + f.shared_text).upper()
    ticket_sheet = get_ticket_sheet_name(f)

    current_row_hit = False
    formula_refs: list[str] = []
    formula_dynamic_keywords: set[str] = set()
    source_mapped_regions: list[str] = []
    dynamic_formula_regions: list[str] = []
    macro_dynamic_keywords = [k for k in ["ACTIVE.CELL", "GET.CELL", "SELECTION", "ACTIVECELL", "ROW", "CELL", "OFFSET", "INDEX", "INDIRECT", "MATCH"] if k in macro_text]
    macro_source_keywords = [k for k in [DATA_SHEET_KEY, EXPECTED_SOURCE_COLS, "$A:$I"] if k.upper() in macro_text]
    macro_current_row_evidence = bool(macro_dynamic_keywords and macro_source_keywords)
    macro_reads_active_selection = False
    macro_uses_active_row = False
    macro_writes_row_index = False
    macro_active_row_controls_ticket = False
    if f.workbook_loaded and ticket_sheet:
        ws = f.wb[ticket_sheet]
        for label, ref in TICKET_FIELD_RANGES.items():
            cr = range_obj(ref)
            if not cr:
                continue
            region_has_source = False
            region_has_dynamic = False
            for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_refs.append(cell.value)
                        if formula_references_data_source(cell.value, f.first_data_sheet):
                            region_has_source = True
                        if formula_has_current_row_logic(cell.value):
                            region_has_dynamic = True
                            for kw in ["ROW", "CELL", "INDEX", "OFFSET", "MATCH", "INDIRECT", "LOOKUP", "VLOOKUP", "XLOOKUP"]:
                                if f"{kw}(" in cell.value.upper():
                                    formula_dynamic_keywords.add(kw)
            if region_has_source:
                source_mapped_regions.append(label)
            if region_has_dynamic:
                dynamic_formula_regions.append(label)
        dynamic_key_source_regions = set(source_mapped_regions) & set(dynamic_formula_regions) & DYNAMIC_TICKET_FIELDS
        # 只在多个关键字段都使用 INDEX/OFFSET/ROW 等当前行逻辑时，才认为票据字段本身具备动态取数能力；
        # 但这还不能证明“用户选中哪一行就打印哪一行”。如果行号只来自旋钮/固定辅助格，WPS 中
        # 选中业务行后点击按钮仍可能打印旧行。必须同时检测宏是否读取 ACTIVE.CELL/SELECTION 等
        # 办公软件当前选区，并把该行写入行索引，或直接用活动行取数。
        formula_current_row_evidence = len(dynamic_key_source_regions) >= 4 and bool({"缴款单位", "账号", "金额大写", "金额小写"} & dynamic_key_source_regions)
        macro_reads_active_selection = any(k in macro_text for k in ["ACTIVE.CELL", "ACTIVECELL", "SELECTION", "GET.CELL"])
        macro_uses_active_row = "ROW(" in macro_text or "CELL(" in macro_text
        macro_writes_row_index = any(k in macro_text for k in ["SET.VALUE", "FORMULA", "SET.NAME"]) and any(k in macro_text for k in ["$S$1", "W2", "对应序号"])
        macro_active_row_controls_ticket = macro_current_row_evidence and (macro_writes_row_index or macro_uses_active_row)
        # P+1 的细则是“点击按钮前用户选中业务记录所在行，宏能识别该行”。因此公式动态取数必须由
        # 宏读取的当前选中行驱动；仅有 INDEX(..., W2, ...) 或 W2=支票明细!S1 不足以加分。
        current_row_hit = formula_current_row_evidence and macro_active_row_controls_ticket
    results.append(RuleResult(
        "P+1",
        "“支票明细”当前行识别逻辑，选中业务记录后可取缴款单位、账号、金额、日期等数据",
        5,
        current_row_hit,
        "检测到宏读取当前选中行并驱动票据字段取数" if current_row_hit else "未检测到宏将办公软件当前选中行用于票据字段取数",
        {"macro_dynamic_keywords": macro_dynamic_keywords, "macro_source_keywords": macro_source_keywords, "macro_reads_active_selection": macro_reads_active_selection, "macro_uses_active_row": macro_uses_active_row, "macro_writes_row_index": macro_writes_row_index, "macro_active_row_controls_ticket": macro_active_row_controls_ticket, "source_mapped_regions": source_mapped_regions, "dynamic_formula_regions": dynamic_formula_regions, "dynamic_formula_keywords": sorted(formula_dynamic_keywords), "formula_refs_sample": formula_refs[:5], "xlm_formula_count": len(f.xlm_formulas)},
    ))

    prompt_terms = [p for p in INVALID_ROW_PROMPTS if p.upper() in all_text]
    guard_terms = [k for k in INVALID_ROW_GUARDS if k in macro_text]
    # RETURN 是正常宏结束指令，不能单独证明“无效行提示/停止”。需要提示语，或 IF/ISBLANK/COUNTA 等校验条件配合 RETURN/ALERT/MESSAGE/STOP。
    prompt_or_stop_logic = bool(prompt_terms) or (any(k in guard_terms for k in ["IF", "ISBLANK", "COUNTA", "COUNT", "ISERROR"]) and any(k in guard_terms for k in ["RETURN", "ALERT", "MESSAGE", "STOP"]))
    # P+2 细则说的是“当前单元格不在有效数据行内”。如果宏只校验 S1/W2 这样的辅助行号，
    # 而不读取办公软件当前活动单元格/选区，那么用户选中无效行时仍可能绕过校验。
    prompt_hit = prompt_or_stop_logic and macro_reads_active_selection and macro_uses_active_row
    results.append(RuleResult(
        "P+2",
        "当前单元格不在有效数据行内时弹出提示或停止执行，不直接打印空白票据",
        3,
        prompt_hit,
        "检测到针对当前活动单元格的提示/停止执行逻辑" if prompt_hit else "未检测到宏基于办公软件当前活动单元格的无效行提示或停止逻辑",
        {"prompt_terms": prompt_terms, "guard_terms": guard_terms, "prompt_or_stop_logic": prompt_or_stop_logic, "macro_reads_active_selection": macro_reads_active_selection, "macro_uses_active_row": macro_uses_active_row},
    ))

    ticket_area_hit = False
    core_evidence: dict[str, Any] = {}
    if f.workbook_loaded and ticket_sheet:
        ws = f.wb[ticket_sheet]
        region_details = collect_range_details(ws, TICKET_FIELD_RANGES)
        covered_core = [lbl for lbl, d in region_details.items() if d["values"] > 0 or d["formulas"] > 0 or d["merged"]]
        non_empty_core = [lbl for lbl, d in region_details.items() if d["values"] > 0 or d["formulas"] > 0]
        merged_core = [lbl for lbl, d in region_details.items() if d["merged"]]
        # 逐项强制：TICKET_FIELD_RANGES 中每一项都必须存在（值/公式/合并任一种映射），
        # 而不再只要求“>=8 个字段被覆盖”。任一字段缺失即视为票据区域不完整。
        missing_core = [lbl for lbl in TICKET_FIELD_RANGES if lbl not in covered_core]
        required_present = not missing_core
        # 逐项强制：DYNAMIC_TICKET_FIELDS 中每一项都必须有值或公式，缺一不可。
        missing_dynamic = [lbl for lbl in DYNAMIC_TICKET_FIELDS if lbl not in non_empty_core]
        dynamic_ok = not missing_dynamic
        # 逐项强制:关键动态字段（面向支票明细取数）必须真正建立到 A:I 的映射。
        dynamic_key_labels = {"日期", "缴款单位", "账号", "金额大写", "金额小写", "费用类型", "核验种类"}
        missing_source_mapped = [lbl for lbl in dynamic_key_labels if lbl not in source_mapped_regions]
        source_ok = not missing_source_mapped
        # 右侧“中部收款单位、收款账号、银行名称、用途说明”做固定文本或公式不变检测：
        # 逐字段检查其关键字/公式是否出现在右侧静态区域（I2:T4 与 H8:T9）内，任一字段缺失即视为被篡改。
        right_side_details = collect_range_details(ws, {r: r for r in RIGHT_SIDE_STATIC_RANGES})
        right_total_non_empty = sum(d["values"] for d in right_side_details.values())
        right_side_text = " ".join(display_text(f, ticket_sheet, r) for r in RIGHT_SIDE_STATIC_RANGES)
        right_side_formula_text = " ".join(
            (sample for d in right_side_details.values() for sample in d["formula_samples"])
        )
        right_side_combined = right_side_text + " " + right_side_formula_text
        right_field_presence = {
            field: any(kw in right_side_combined for kw in kws)
            for field, kws in RIGHT_SIDE_FIELD_KEYWORDS.items()
        }
        missing_right_fields = [field for field, ok in right_field_presence.items() if not ok]
        # 右侧要求：四个规定字段均能被检测到（固定文本或公式不变），且区域整体有足够非空单元格。
        right_side_ok = not missing_right_fields and right_total_non_empty >= 4
        macro_selects_ticket = ticket_sheet in f.macro_text and any(k in macro_text for k in ["SELECT", "ACTIVATE", "FORMULA", "SET.VALUE"])
        macro_selected_refs = extract_range_refs_from_text(f.macro_text)
        selected_cover_core = [label for label, ref in TICKET_FIELD_RANGES.items() if any(range_contains(selected_ref, ref) for selected_ref in macro_selected_refs)]
        missing_selected = [lbl for lbl in TICKET_FIELD_RANGES if lbl not in selected_cover_core]
        selected_area_ok = not missing_selected
        # P+3 不只要求票据区域被选中/有公式，还要求宏执行后填充的是用户当前选中业务行的数据。
        # 如果 P+1 的当前行联动证据不成立，票据区即使有 INDEX(..., W2, ...) 也可能显示旧行，不能加分。
        ticket_area_hit = required_present and dynamic_ok and source_ok and right_side_ok and (macro_selects_ticket or selected_area_ok) and selected_area_ok and current_row_hit
        core_evidence = {
            "covered_core_ranges": covered_core,
            "missing_core_ranges": missing_core,
            "non_empty_core_ranges": non_empty_core,
            "missing_dynamic_ranges": missing_dynamic,
            "merged_core_ranges": merged_core,
            "source_mapped_regions": source_mapped_regions,
            "missing_source_mapped": missing_source_mapped,
            "right_side_non_empty": right_total_non_empty,
            "right_side_field_presence": right_field_presence,
            "missing_right_fields": missing_right_fields,
            "right_side_ok": right_side_ok,
            "macro_selects_ticket": macro_selects_ticket,
            "macro_selected_refs": macro_selected_refs,
            "selected_cover_core_ranges": selected_cover_core,
            "missing_selected_ranges": missing_selected,
            "selected_area_ok": selected_area_ok,
            "current_row_hit": current_row_hit,
            "macro_active_row_controls_ticket": macro_active_row_controls_ticket,
        }
    results.append(RuleResult(
        "P+3",
        "宏执行后自动选定/填充“划拨进账单”票据区域，覆盖日期、缴款单位、账号、金额等核心区域，来源为支票明细 A:I",
        5,
        ticket_area_hit,
        "核心票据区域结构完整，来源映射及右侧静态区域证据充分" if ticket_area_hit else "未检测到足够证据证明宏自动选定/填充完整票据区域",
        core_evidence,
    ))

    print_area_hit = False
    areas: list[str] = []
    print_area_evidence: dict[str, Any] = {}
    if ticket_sheet:
        raw_areas = print_area_for(f, ticket_sheet)
        areas, _bad = normalized_valid_ranges(raw_areas)
        # 细则允许两种情况：打印区域为 B1:W9，或“覆盖完整票据框”的更大区域（允许合理外扩）。
        # 因此不再强制严格等于 B1:W9，只要求区域包含完整票据框 B1:W9 即可；
        # 但要排除“只打印局部金额区/空白区/整表大面积空白”，故追加内容与空白检测。
        candidate_areas = [area for area in areas if range_contains(area, TICKET_BOUNDARY)]
        area_reports: list[dict[str, Any]] = []
        for area in candidate_areas:
            report_item: dict[str, Any] = {"area": area}
            if f.workbook_loaded and ticket_sheet in f.wb.sheetnames:
                stats = print_area_content_stats(f.wb[ticket_sheet], area)
                report_item.update(stats)
                # 覆盖完整票据框；区域内存在票据内容；且不存在大面积空白（有效内容占比达标或外扩幅度有限）。
                report_item["ok"] = (
                    stats["meaningful"] >= MIN_PRINT_AREA_MEANINGFUL
                    and stats["meaningful_ratio"] >= MIN_PRINT_AREA_FILL_RATIO
                )
            else:
                report_item["ok"] = True  # 无法读取工作表内容时，仅凭区域包含完整票据框判定
            area_reports.append(report_item)
            if report_item["ok"]:
                print_area_hit = True
        print_area_evidence = {
            "ticket_sheet": ticket_sheet,
            "print_areas": areas,
            "expected": TICKET_BOUNDARY,
            "candidate_areas_cover_ticket": candidate_areas,
            "area_content_reports": area_reports,
            "min_meaningful": MIN_PRINT_AREA_MEANINGFUL,
            "min_fill_ratio": MIN_PRINT_AREA_FILL_RATIO,
        }
    results.append(RuleResult(
        "P+4",
        "宏执行后将打印区域设置为 B1:W9 或覆盖完整票据框",
        5,
        print_area_hit,
        "打印区域覆盖完整票据框且内含票据内容、无大面积空白" if print_area_hit else "打印区域未覆盖完整票据框，或存在局部/空白/整表大面积空白",
        print_area_evidence,
    ))

    return results


# ------------------------- 维度2扣分项 -------------------------

def evaluate_penalties(f: WorkbookFeatures) -> list[RuleResult]:
    results: list[RuleResult] = []
    data_sheet = f.first_data_sheet

    add_penalty(results, "P-1", -5, "第一个数据工作表中没有可点击的“打印进账单”按钮或控件", not has_print_button(f),
                "未在第一个数据工作表检测到锚定的可点击“打印进账单”按钮/控件",
                {"first_data_sheet": data_sheet,
                 "objects_on_first_sheet": [o.to_dict() for o in f.objects if o.sheet == data_sheet][:10],
                 "all_objects_sample": [o.to_dict() for o in f.objects[:10]],
                 "defined_names": f.defined_names})

    width_hit, width_evidence = check_account_width_penalty(f, data_sheet)
    add_penalty(results, "P-14", -1, "“支票明细”中“结算账号”一列列宽超过30字符或小于10字符", width_hit,
                "结算账号列宽不在 [10, 30] 范围内", width_evidence)

    return results


def add_penalty(results: list[RuleResult], rule_id: str, points: int, title: str, hit: bool, message: str, evidence: dict[str, Any]) -> None:
    results.append(RuleResult(rule_id, title, points, hit, message if hit else "未命中扣分项", evidence))


def check_multipage_penalty(f: WorkbookFeatures, ticket_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not ticket_sheet or ticket_sheet not in f.wb.sheetnames:
        return False, {}
    ws = f.wb[ticket_sheet]
    areas = print_area_for(f, ticket_sheet)
    too_large = False
    for area in areas:
        cr = range_obj(area)
        if cr and ((cr.max_col - cr.min_col + 1) > 24 or (cr.max_row - cr.min_row + 1) > 12):
            too_large = True
    fit_w = ws.page_setup.fitToWidth
    fit_h = ws.page_setup.fitToHeight
    fit_page = bool(getattr(ws.sheet_properties.pageSetUpPr, "fitToPage", False)) if ws.sheet_properties.pageSetUpPr else False
    breaks = len(getattr(ws.row_breaks, "brk", [])) + len(getattr(ws.col_breaks, "brk", []))
    hit = (too_large and not fit_page and not fit_w) or breaks > 0
    return hit, {"print_areas": areas, "fitToWidth": fit_w, "fitToHeight": fit_h, "fitToPage": fit_page, "manual_breaks": breaks, "too_large": too_large}


def _rgb_from_color(color: Any) -> tuple[int, int, int] | None:
    color_type = getattr(color, "type", None)
    raw = None
    if color_type == "rgb" or getattr(color, "rgb", None):
        raw = getattr(color, "rgb", None)
    elif color_type == "indexed":
        idx = getattr(color, "indexed", None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            raw = COLOR_INDEX[idx]
    if not raw or raw == "00000000":
        return None
    raw = str(raw)
    if len(raw) == 8:
        raw = raw[2:]
    if len(raw) != 6:
        return None
    try:
        rgb = (int(raw[0:2], 16), int(raw[2:4], 16), int(raw[4:6], 16))
    except ValueError:
        return None
    tint = getattr(color, "tint", 0) or 0
    if tint:
        rgb = apply_tint(rgb, float(tint))
    return rgb


def apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    def channel(value: int) -> int:
        if tint < 0:
            return round(value * (1 + tint))
        return round(value * (1 - tint) + 255 * tint)
    return (max(0, min(255, channel(rgb[0]))), max(0, min(255, channel(rgb[1]))), max(0, min(255, channel(rgb[2]))))


def color_rgb(cell: Any, wb: Any = None) -> tuple[int, int, int] | None:
    if not cell.fill or cell.fill.fill_type in (None, "none"):
        return None
    for color in (cell.fill.fgColor, cell.fill.bgColor):
        rgb = _rgb_from_color(color)
        if rgb:
            return rgb
    return None


def classify_color(rgb: tuple[int, int, int]) -> str | None:
    r, g, b = rgb
    # 兼容 Excel 常用浅黄色 FEF3C7、深绿/深青 0F766E、浅蓝 E0F2FE/DBEAFE 等模板色。
    if r >= 220 and g >= 200 and 120 <= b <= 220 and r >= g >= b:
        return "yellow"
    if g >= 80 and g >= r * 1.8 and g >= b * 0.9 and b >= r * 1.5:
        return "green"
    if b >= 180 and b >= r + 15 and b >= g + 5:
        return "blue"
    return None


def check_color_penalty(f: WorkbookFeatures, ticket_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not ticket_sheet or ticket_sheet not in f.wb.sheetnames:
        return False, {}
    ws = f.wb[ticket_sheet]
    found_by_region: dict[str, bool] = {"yellow": False, "green": False, "blue": False}
    samples: dict[str, list[str]] = {"yellow": [], "green": [], "blue": []}
    for expected, refs in COLOR_REGION_RULES.items():
        for ref in refs:
            cr = range_obj(ref)
            if not cr:
                continue
            for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
                for cell in row:
                    rgb = color_rgb(cell, f.wb)
                    if not rgb:
                        continue
                    kind = classify_color(rgb)
                    if kind == expected:
                        found_by_region[expected] = True
                        if len(samples[expected]) < 5:
                            samples[expected].append(f"{cell.coordinate}#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")
    missing = [c for c in ["yellow", "green", "blue"] if not found_by_region[c]]
    return len(missing) >= 2, {"found_by_region": found_by_region, "missing": missing, "samples": samples, "region_rules": COLOR_REGION_RULES}


def check_button_overlap_penalty(f: WorkbookFeatures, data_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not data_sheet:
        return False, {}
    header_row = f.data_header_row or 1
    data_ref = f"A{header_row}:I{min(header_row + 20, 100)}"
    overlaps = []
    for obj in f.objects:
        if obj.kind not in {"button", "control_prop"}:
            continue
        if obj.sheet and obj.sheet != data_sheet:
            continue
        if obj.ref and ranges_intersect(obj.ref, data_ref):
            overlaps.append(obj.to_dict())
    return bool(overlaps), {"data_ref": data_ref, "overlapping_buttons": overlaps}


def check_data_area_penalty(f: WorkbookFeatures, data_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not data_sheet or data_sheet not in f.wb.sheetnames:
        return True, {"reason": "无法读取第一个数据工作表"}
    ws = f.wb[data_sheet]
    max_row = min(max(ws.max_row or 1, 1), 100)
    target_ref = f"A1:I{max_row}"
    total = non_empty = garbled = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            total += 1
            txt = text_of(cell.value)
            if txt:
                non_empty += 1
                if txt.count("�") > 0 or len(re.findall(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", txt)) > 0:
                    garbled += 1
    cover = object_cover_ratio(f.objects, data_sheet, target_ref)
    header_ok = bool(f.data_header_row and len(f.data_header_cols) >= 2)
    hit = non_empty < 5 or not header_ok or (non_empty and garbled / non_empty > 0.1) or cover > 0.3
    return hit, {"target_ref": target_ref, "total_cells": total, "non_empty": non_empty, "header_row": f.data_header_row, "header_cols": f.data_header_cols, "garbled_cells": garbled, "object_cover_ratio": round(cover, 3)}


def check_unrelated_sheet_penalty(f: WorkbookFeatures) -> tuple[bool, dict[str, Any]]:
    extras = []
    if not f.workbook_loaded:
        return False, {}
    for name in f.wb.sheetnames:
        if any(k in name for k in ALLOWED_SHEET_KEYWORDS):
            continue
        ws = f.wb[name]
        used = (ws.max_row or 0) * (ws.max_column or 0)
        non_empty = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), min_col=1, max_col=min(ws.max_column or 1, 30)):
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty += 1
        suspicious = used > 200 or non_empty == 0 or TARGET_SHEET_KEY in name or "空白" in name or "副本" in name
        if suspicious:
            extras.append({"sheet": name, "used_cells_estimate": used, "sample_non_empty": non_empty})
    return bool(extras), {"suspicious_extra_sheets": extras}


def check_uppercase_amount_penalty(f: WorkbookFeatures, ticket_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not ticket_sheet or ticket_sheet not in f.wb.sheetnames:
        return False, {}
    ws = f.wb[ticket_sheet]
    formula_text = range_values_text(ws, "B5:H6")
    upper_text = display_text(f, ticket_sheet, "B5:H6")
    small_text = display_text(f, ticket_sheet, "N6:T6")
    # 公式文本中会包含工作表名/函数名等汉字，非法汉字检查只对缓存显示值生效。
    formula_only = "=" in upper_text or upper_text == formula_text
    lower_num_chars = sorted(set(upper_text) & LOWER_CN_NUMS) if not formula_only else []
    chinese_chars = re.findall(r"[一-鿿]", upper_text) if not formula_only else []
    invalid_chars = sorted({ch for ch in chinese_chars if ch not in UPPER_ALLOWED and ch not in {"大", "写", "金", "额"}})
    small_amount = extract_ticket_small_amount(small_text)
    upper_amount = chinese_money_to_number(upper_text)
    mismatch = (not formula_only) and small_amount is not None and upper_amount is not None and abs(small_amount - upper_amount) > 0.02
    hit = bool(lower_num_chars or invalid_chars or mismatch)
    return hit, {"upper_text": upper_text, "formula_text_sample": formula_text[:300], "small_text": small_text, "ordinary_chinese_digits": lower_num_chars, "invalid_chinese_chars": invalid_chars, "small_amount": small_amount, "upper_amount": upper_amount, "mismatch": mismatch, "formula_only": formula_only}


def check_verify_text_penalty(f: WorkbookFeatures, ticket_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not ticket_sheet or ticket_sheet not in f.wb.sheetnames:
        return False, {}
    lower = display_text(f, ticket_sheet, "B8:G9")
    top = display_text(f, ticket_sheet, "B1:W3")
    all_ticket = display_text(f, ticket_sheet, TICKET_BOUNDARY)
    required = ["公共照明维护费", "票据核验"]
    missing = [t for t in required if t not in all_ticket]
    moved_top = [t for t in required if t in top and t not in lower]
    hit = bool(missing or moved_top)
    return hit, {"lower_left_text": lower, "top_text": top, "missing": missing, "moved_to_top": moved_top}


def check_merge_penalty(f: WorkbookFeatures, ticket_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not ticket_sheet or ticket_sheet not in f.wb.sheetnames:
        return False, {}
    ws = f.wb[ticket_sheet]
    important = ["F1:I1", "B2:H2", "B3:H3", "B5:H6", "N6:T6", "B8:G8", "B9:G9"]
    ok = [ref for ref in important if merged_contains_or_intersects(ws, ref)]
    missing = [ref for ref in important if ref not in ok]
    hit = len(ok) < max(3, math.ceil(len(important) * 0.5))
    return hit, {"merged_core_ranges": ok, "missing_or_not_merged": missing, "all_merges_in_ticket": [str(mr) for mr in ws.merged_cells.ranges if ranges_intersect(str(mr), TICKET_BOUNDARY)]}


def check_account_width_penalty(f: WorkbookFeatures, data_sheet: str | None) -> tuple[bool, dict[str, Any]]:
    if not f.workbook_loaded or not data_sheet or data_sheet not in f.wb.sheetnames:
        return False, {}
    ws = f.wb[data_sheet]
    account_col = None
    if f.data_header_row:
        for col in range(1, min(ws.max_column or 1, 30) + 1):
            txt = text_of(ws.cell(row=f.data_header_row, column=col).value)
            if "结算账号" in txt:
                account_col = col
                break
    if not account_col:
        return False, {"reason": "未找到结算账号列"}
    letter = int_to_col(account_col)
    width = ws.column_dimensions[letter].width
    # openpyxl 中未显式设置的列宽可能是 None，Excel 默认约 8.43。按细则小于 10 处理。
    effective_width = width if width is not None else 8.43
    hit = effective_width < 10 or effective_width > 30
    return hit, {"column": letter, "width": effective_width, "header_row": f.data_header_row}


# ------------------------- 金额解析 -------------------------

def extract_first_number(text: str) -> float | None:
    if not text:
        return None
    m = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text.replace("，", ","))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def extract_ticket_small_amount(text: str) -> float | None:
    """解析金额小写区域。

    票据金额小写常被拆到多个单元格显示，如“￥ 2 3 2 2 8 5”表示 2322.85。
    若文本是常规数字/带小数，则直接返回该数字；若是拆位显示，则把最后两位视为角分。
    """
    if not text:
        return None
    normalized = text.replace("，", ",").replace("￥", " ").replace("¥", " ")
    normal = re.search(r"[-+]?\d[\d,]*\.\d+", normalized)
    if normal:
        try:
            return float(normal.group(0).replace(",", ""))
        except ValueError:
            return None
    tokens = re.findall(r"\d", normalized)
    # 公式文本中也会有很多数字，只有纯展示拆位形态才用该逻辑。
    stripped = re.sub(r"[\d\s,￥¥.]", "", text)
    if tokens and not stripped:
        number = int("".join(tokens))
        return round(number / 100, 2)
    return extract_first_number(text)


def chinese_money_to_number(text: str) -> float | None:
    if not text:
        return None
    text = text.replace("人民币", "").replace(" ", "").replace("正", "整").replace("圆", "元")
    if not any(ch in text for ch in "壹贰叁肆伍陆柒捌玖零"):
        return None
    digit = {"零": 0, "壹": 1, "贰": 2, "叁": 3, "肆": 4, "伍": 5, "陆": 6, "柒": 7, "捌": 8, "玖": 9}
    small_unit = {"拾": 10, "佰": 100, "仟": 1000}

    def parse_section(section: str) -> int:
        total = 0
        num = 0
        for ch in section:
            if ch in digit:
                num = digit[ch]
            elif ch in small_unit:
                total += (num or 1) * small_unit[ch]
                num = 0
        return total + num

    integer_part = text
    fraction = 0.0
    if "元" in text:
        integer_part = text.split("元", 1)[0]
        frac = text.split("元", 1)[1]
        j = re.search(r"([零壹贰叁肆伍陆柒捌玖])角", frac)
        f = re.search(r"([零壹贰叁肆伍陆柒捌玖])分", frac)
        if j:
            fraction += digit[j.group(1)] * 0.1
        if f:
            fraction += digit[f.group(1)] * 0.01

    total_int = 0
    if "亿" in integer_part:
        before, integer_part = integer_part.split("亿", 1)
        total_int += parse_section(before) * 100000000
    if "万" in integer_part:
        before, integer_part = integer_part.split("万", 1)
        total_int += parse_section(before) * 10000
    total_int += parse_section(integer_part)
    return round(total_int + fraction, 2)


# ------------------------- 总评估与输出 -------------------------

def _build_report(f: WorkbookFeatures) -> EvaluationReport:
    d1 = evaluate_dimension1(f)
    d1_pass = all(r.hit for r in d1)
    if not d1_pass:
        return EvaluationReport(str(f.path), d1, [], [], False, 0, 0)

    positives = evaluate_positive_points(f)
    penalties = evaluate_penalties(f)
    raw_score = sum(r.points for r in positives if r.hit) + sum(r.points for r in penalties if r.hit)
    # 分数细则允许负分，总分不做非负截断。
    final_score = raw_score
    return EvaluationReport(str(f.path), d1, positives, penalties, True, raw_score, final_score)


def _locate_workbook(directory: Path) -> Path | None:
    """在指定目录中定位待评估的 Excel 文档。

    优先返回默认文件名（``DEFAULT_FILE``）；否则按 xlsm/xlsx/xls 顺序取第一个可用文件，
    并跳过 Excel 打开时留下的 ``~$`` 临时锁文件。
    """
    preferred = directory / DEFAULT_FILE
    if preferred.exists() and preferred.is_file():
        return preferred
    for suffix in SUPPORTED_SUFFIXES:
        for candidate in sorted(directory.glob(f"*{suffix}")):
            if candidate.is_file() and not candidate.name.startswith("~$"):
                return candidate
    return None


def _dim2_item(rule: RuleResult) -> dict[str, Any]:
    return {
        "rule": rule.title,
        "max_delta": rule.points,
        "delta": rule.points if rule.hit else 0,
        "hit": rule.hit,
        "detail": "",
    }


def evaluate(dir_path: str) -> dict[str, Any]:
    """脚本统一入口。

    参数 ``dir_path`` 为脚本所在目录的路径；脚本自行在该目录内定位并打开被评估的 Excel 文档。
    返回结构化字典（字段定义见项目《脚本接口差异与统一建议》§2.2）。
    """
    result: dict[str, Any] = {
        "id": SCRIPT_ID,
        "file_name": None,
        "status": "ok",
        "error": None,
        "dim1_pass": False,
        "dim1_reason": "",
        "dim2_items": [],
        "total_score": 0,
        "max_score": DIM2_MAX_SCORE,
    }
    try:
        directory = Path(dir_path)
        if not directory.exists() or not directory.is_dir():
            result["status"] = "error"
            result["error"] = f"目录不存在或不是目录：{dir_path}"
            return result

        target = _locate_workbook(directory)
        if target is None:
            result["status"] = "error"
            result["error"] = f"未在目录中找到可评估的 Excel 文件（.xlsm/.xlsx/.xls）：{dir_path}"
            return result
        result["file_name"] = target.name

        features = extract_features(target)
        report = _build_report(features)

        result["dim1_pass"] = report.dimension1_pass
        if not report.dimension1_pass:
            failed_titles = [r.title for r in report.dimension1 if not r.hit]
            result["dim1_reason"] = "；".join(failed_titles)
            result["total_score"] = 0
            return result

        result["dim2_items"] = [_dim2_item(r) for r in report.positives] + [_dim2_item(r) for r in report.penalties]
        result["total_score"] = report.final_score
        return result
    except Exception as exc:  # 顶层兜底：脚本自身异常不应向上抛出
        result["status"] = "error"
        result["error"] = f"{type(exc).__name__}: {exc}"
        return result


if __name__ == "__main__":
    # 仅用于本地调试：将命令行第一个参数视作脚本所在目录，缺省时使用脚本所在目录。
    debug_dir = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).resolve().parent)
    print(json.dumps(evaluate(debug_dir), ensure_ascii=False, indent=2, default=str))

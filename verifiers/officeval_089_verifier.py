"""
自动评估脚本：企业财务统计总表_已添加财务图表.xlsx

评估逻辑：
1. 先检查维度1（可用与可修改性）。维度1不通过，直接返回 total_score=0。
2. 维度1通过后，按维度2逐项自动检测。满足得分点则累加正分；命中扣分点则累加负分。
3. 最后以结构化字典的形式返回评估结果。

依赖：openpyxl
安装：pip install openpyxl
"""
import json
import os
import re
import sys
import zipfile
from collections import Counter

import openpyxl
from openpyxl.chart import LineChart, BarChart

SCRIPT_ID = "089"
SOURCE_SHEET = "企业财务统计总表"
TARGET_SHEET = "Sheet2"
EXPECTED_TIMES = ["2021年度", "2022年度", "2023年度", "2024年度", "2025年度",
                  "2026Q1", "2026Q2", "2026Q3", "2026Q4"]
TARGET_FILE_NAME = "企业财务统计总表_已添加财务图表.xlsx"


# ───────────────────────────── 基础工具函数 ─────────────────────────────
def norm_text(v):
    return "" if v is None else str(v).strip()


def normalize_ref(ref):
    """把 Excel 引用统一成便于比较的形式：去掉 $、单引号、空格。"""
    if ref is None:
        return ""
    return str(ref).replace("$", "").replace("'", "").replace(" ", "").strip()


def ref_match(actual, expected):
    return normalize_ref(actual) == normalize_ref(expected)


def ref_in(actual, expected_list):
    return any(ref_match(actual, x) for x in expected_list)


def formula_or_ref(obj):
    """openpyxl 不同对象版本可能使用 .f 或 .ref 存储引用。"""
    if obj is None:
        return None
    for attr in ("f", "ref"):
        try:
            value = getattr(obj, attr)
            if value:
                return value
        except Exception:
            pass
    return None


def ser_val_ref(ser):
    try:
        return formula_or_ref(ser.val.numRef)
    except Exception:
        return None


def ser_cat_ref(ser):
    try:
        if ser.cat.strRef:
            return formula_or_ref(ser.cat.strRef)
    except Exception:
        pass
    try:
        if ser.cat.numRef:
            return formula_or_ref(ser.cat.numRef)
    except Exception:
        pass
    return None


def ser_name(ser):
    """读取图表系列名称，支持字面值和单元格引用缓存值。"""
    try:
        tx = ser.title
        if tx is None:
            return ""
        v = getattr(tx, "v", None)
        if v:
            return str(v).strip()
        str_ref = getattr(tx, "strRef", None)
        if str_ref is not None:
            # 优先读缓存值
            try:
                pts = str_ref.strCache.pt
                if pts:
                    return str(pts[0].v).strip()
            except Exception:
                pass
            # 退而求其次返回引用本身
            f = formula_or_ref(str_ref)
            return f or ""
        return str(tx).strip()
    except Exception:
        return ""


def chart_title(chart):
    """读取图表标题文本。"""
    try:
        title = chart.title
        if title is None:
            return ""
        # openpyxl 标准结构：chart.title.tx.rich.p[].r[].t
        try:
            parts = []
            for para in title.tx.rich.p:
                for run in para.r:
                    parts.append(run.t or "")
            if parts:
                return "".join(parts).strip()
        except Exception:
            pass
        # 兼容其他结构
        try:
            parts = []
            for para in title.tx.rich.paragraphs:
                for run in para.runs:
                    parts.append(run.text or "")
            if parts:
                return "".join(parts).strip()
        except Exception:
            pass
        return str(title).strip()
    except Exception:
        return ""


def legend_pos(chart):
    try:
        return chart.legend.legendPos if chart.legend else None
    except Exception:
        return None


def legend_entries_match_series_style(chart):
    """图例项与对应系列样式的映射检查。
    Excel/WPS 图例色块本身没有独立存储的颜色属性，规范上始终跟随其代表系列的线条/填充颜色渲染，
    openpyxl 也读不到“渲染后的图例色块颜色”。因此这里退一步核实：
    1) 图例条目数量与顺序和系列数量、顺序一致（即图例确实逐一对应到每个系列）；
    2) 每个系列自身的线条颜色是可读的、且各不相同（这是图例色块能够“不同色”的前提）。
    任一系列颜色读取不到（为空或"None"）时，视为无法确认图例与折线颜色一致，不给分。"""
    if not chart or not chart.legend or not chart.series:
        return False
    series = list(chart.series)
    if len(series) < 2:
        return False
    fills = [line_fill(s) for s in series]
    if any((not f) or f == "None" for f in fills):
        return False
    return len(set(fills)) == len(fills)


def is_bottom_legend(chart):
    return legend_pos(chart) in ("b", "bot", "bottom")


def has_marker(ser):
    """判断折线是否带数据标记。symbol=None 在 Excel 中通常代表自动标记，也视作有标记。"""
    try:
        marker = ser.marker
        if marker is None:
            return False
        symbol = marker.symbol
        return symbol is None or symbol not in ("none", "None")
    except Exception:
        return False


def marker_symbol(ser):
    try:
        return ser.marker.symbol
    except Exception:
        return None


def marker_size(ser):
    try:
        return ser.marker.size
    except Exception:
        return None


def line_width(ser):
    try:
        return ser.graphicalProperties.line.width
    except Exception:
        return None


def line_fill(ser):
    try:
        return str(ser.graphicalProperties.line.solidFill)
    except Exception:
        return ""


def fill_color(ser):
    try:
        return str(ser.graphicalProperties.solidFill)
    except Exception:
        return ""


def anchor_from(chart):
    """返回图表左上角锚点：(col, row)，openpyxl为0基坐标。"""
    try:
        return chart.anchor._from.col, chart.anchor._from.row
    except Exception:
        return 999, 999


def anchor_to(chart):
    try:
        return chart.anchor.to.col, chart.anchor.to.row
    except Exception:
        return None, None


def is_excel_editable_chart(chart):
    return isinstance(chart, (LineChart, BarChart))


def sorted_charts(charts):
    return sorted(charts, key=lambda c: (anchor_from(c)[1], anchor_from(c)[0]))


def all_series_cat_ref(chart, expected="企业财务统计总表!A4:A12"):
    return bool(chart.series) and all(ref_match(ser_cat_ref(s), expected) for s in chart.series)


def series_names(chart):
    return [ser_name(s) for s in chart.series]


def series_val_refs(chart):
    return [ser_val_ref(s) for s in chart.series]


def find_by_title(charts, *keywords):
    for c in charts:
        t = chart_title(c)
        if all(k in t for k in keywords):
            return c
    return None


def find_line_chart(charts, *keywords):
    for c in charts:
        t = chart_title(c)
        if isinstance(c, LineChart) and all(k in t for k in keywords):
            return c
    return None


def find_bar_chart(charts, *keywords):
    for c in charts:
        t = chart_title(c)
        if isinstance(c, BarChart) and all(k in t for k in keywords):
            return c
    return None


def source_values(ws, ref_col):
    return [ws.cell(r, ref_col).value for r in range(4, 13)]


def pct_two_decimal_ok(ws_formula, col):
    """判断某列(4~12行)是否为百分比格式且保留两位小数。
    办公软件中百分比+两位小数的数字格式即形如 0.00%（可带千分位/前缀），核心特征：含 % 且小数位为两位。"""
    for row in range(4, 13):
        fmt = norm_text(ws_formula.cell(row, col).number_format)
        if "%" not in fmt:
            return False
        # 取 % 之前的小数位数，要求恰为两位（如 0.00% / #,##0.00%）
        m = re.search(r"\.(0+)%", fmt)
        if not m or len(m.group(1)) != 2:
            return False
    return True


def _is_pct_two_decimal_fmt(fmt):
    """判断单个数字格式字符串是否为百分比且保留两位小数（0.00% / #,##0.00% 等）。"""
    fmt = norm_text(fmt)
    if "%" not in fmt:
        return False
    m = re.search(r"\.(0+)%", fmt)
    return bool(m) and len(m.group(1)) == 2


def numfmt_source_id(numfmt_obj):
    """读取 c:numFmt 上的 sourceLinked 属性；True 表示未自定义，沿用单元格格式。"""
    try:
        return getattr(numfmt_obj, "sourceLinked", None)
    except Exception:
        return None


def series_dlbls_numfmt(ser):
    """读取某系列 dLbls 的 numFmt；取不到返回 None。"""
    try:
        dlbls = ser.dLbls
        if dlbls is None:
            return None
        return getattr(dlbls, "numFmt", None)
    except Exception:
        return None


def chart_dlbls_numfmt(chart):
    """读取图表级 dLbls 的 numFmt；取不到返回 None。"""
    try:
        dlbls = chart.dLbls
        if dlbls is None:
            return None
        return getattr(dlbls, "numFmt", None)
    except Exception:
        return None


def axis_numfmt(chart):
    """读取数值轴（y 轴）的 numFmt，用作提示值/坐标轴显示格式的兜底来源。"""
    try:
        return getattr(chart.y_axis, "number_format", None)
    except Exception:
        return None


def series_display_numfmt(chart, ser):
    """确定某系列在图表上实际显示的数值格式字符串：
    优先级：系列 dLbls.numFmt > 图表 dLbls.numFmt > 数值轴 numFmt。
    三者均未设置（或未开启数据标签）时返回 None，代表图表未按两位小数百分比显示数值。"""
    for numfmt in (series_dlbls_numfmt(ser), chart_dlbls_numfmt(chart)):
        if numfmt:
            return str(numfmt)
    fmt = axis_numfmt(chart)
    return str(fmt) if fmt else None


def chart_labels_pct_two_decimal_ok(chart):
    """判断图表所有系列的数据标签/提示值格式是否均为百分比两位小数。
    要求图表确实开启了数据标签（系列或图表级 dLbls 存在），且其 numFmt 满足 0.00% 形式；
    若仅有坐标轴格式而没有数据标签，则视为未满足“图表数值/提示值/数据标签两位小数”的要求。"""
    if not chart or not chart.series:
        return False
    for ser in chart.series:
        has_dlbls = series_dlbls_numfmt(ser) is not None or chart_dlbls_numfmt(chart) is not None
        if not has_dlbls:
            return False
        fmt = series_display_numfmt(chart, ser)
        if not _is_pct_two_decimal_fmt(fmt):
            return False
    return True


def evaluate(dir_path: str) -> dict:
    """评估入口。

    参数：
        dir_path: 脚本所在目录的路径。脚本自己在该目录里定位并打开被评估的文档。

    返回：
        结构化评估结果 dict（见文件顶部说明与 §2.2 约定）。
    """
    result = {
        "id": SCRIPT_ID,
        "file_name": TARGET_FILE_NAME,
        "status": "ok",
        "error": None,
        "dim1_pass": False,
        "dim1_reason": "",
        "dim2_items": [],
        "total_score": 0,
        "max_score": 0,
    }

    # 每项：{"rule", "max_delta", "delta", "hit", "detail"}
    dim2_items = []
    score = 0

    def add_item(max_delta, rule, ok, detail=""):
        nonlocal score
        delta = max_delta if ok else 0
        # 扣分项：max_delta 为负数，命中即代表"触发扣分"
        if ok:
            score += delta
        dim2_items.append({
            "rule": rule,
            "max_delta": max_delta,
            "delta": delta,
            "hit": ok,
            "detail": "",
        })

    def hit(max_delta, rule, detail=""):
        add_item(max_delta, rule, True, detail)

    def miss(max_delta, rule, detail=""):
        add_item(max_delta, rule, False, detail)

    try:
        # ─────────── 在目录内定位被评估文档 ───────────
        file_path = os.path.join(dir_path, TARGET_FILE_NAME)
        if not os.path.exists(file_path):
            # 未找到默认文件名时，回退：扫描目录下第一个 .xlsx/.xlsm 文件
            candidates = []
            try:
                for name in os.listdir(dir_path):
                    ext = os.path.splitext(name)[1].lower()
                    if ext in (".xlsx", ".xlsm") and not name.startswith("~$"):
                        candidates.append(name)
            except Exception:
                candidates = []
            if candidates:
                file_path = os.path.join(dir_path, candidates[0])
                result["file_name"] = candidates[0]
            else:
                result["status"] = "error"
                result["error"] = f"目录中未找到被评估文档：{dir_path}"
                return result
        else:
            result["file_name"] = TARGET_FILE_NAME

        # ─────────── 维度1：可用与可修改性 ───────────
        dim1_pass = True
        dim1_reason = ""

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".xlsx", ".xlsm"):
            dim1_pass = False
            dim1_reason = "文件格式不是 .xlsx/.xlsm"

        if dim1_pass and not os.path.exists(file_path):
            dim1_pass = False
            dim1_reason = f"文件不存在：{file_path}"

        wb = None
        if dim1_pass:
            try:
                # 检查zip结构，能尽早发现损坏的xlsx
                if not zipfile.is_zipfile(file_path):
                    raise ValueError("不是有效的 Office Open XML 文件")
                wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as e:
                dim1_pass = False
                dim1_reason = f"文件无法正常打开：{e}"

        ws1 = ws2 = None
        if dim1_pass:
            missing_sheets = [
                name for name in (SOURCE_SHEET, TARGET_SHEET)
                if name not in wb.sheetnames
            ]
            if missing_sheets:
                dim1_pass = False
                dim1_reason = f"缺少必需工作表：{', '.join(missing_sheets)}"
            else:
                ws1 = wb[SOURCE_SHEET]
                ws2 = wb[TARGET_SHEET]

        result["dim1_pass"] = dim1_pass
        result["dim1_reason"] = dim1_reason

        if not dim1_pass:
            # 维度1未通过，不再检查维度2，直接返回
            # max_score 按维度2 正分项满分之和填写（27 项正分 = 57），便于对齐
            result["max_score"] = 57
            result["total_score"] = 0
            return result

        # ─────────── 定位目标图表 ───────────
        charts = sorted_charts(ws2._charts)
        line_profit = find_line_chart(charts, "毛利率", "净利率", "趋势")
        bar_profit = find_bar_chart(charts, "毛利率", "净利率", "柱形")
        roe_chart = find_line_chart(charts, "净资产收益率")
        debt_chart = find_line_chart(charts, "资产负债率")
        growth_chart = find_line_chart(charts, "营业收入增长率", "净利润增长率")

        # 若底部增长率图标题不完整，也尝试用双系列 K/L 引用定位
        if growth_chart is None:
            for c in charts:
                vals = [normalize_ref(x) for x in series_val_refs(c)]
                if isinstance(c, LineChart) and "企业财务统计总表!K4:K12" in vals and "企业财务统计总表!L4:L12" in vals:
                    growth_chart = c
                    break

        # ─────────── 维度2：完成度评分 ───────────

        # +5：Sheet2中出现五个图表
        if len(charts) == 5:
            hit(5, "Sheet2中出现五个图表")
        else:
            miss(5, "Sheet2中出现五个图表", f"实际 {len(charts)} 个")

        # +1：Sheet2工作表：只放置财务分析标题、5个图表及必要图例
        non_empty_cells = [(cell.coordinate, cell.value)
                           for row in ws2.iter_rows() for cell in row
                           if cell.value not in (None, "")]
        images_count = len(getattr(ws2, "_images", []))
        sheet2_is_second = TARGET_SHEET in wb.sheetnames and wb.sheetnames.index(TARGET_SHEET) == 1
        has_title_cell = any(isinstance(v, str) and norm_text(v) for _, v in non_empty_cells)
        has_five_charts = len(charts) == 5
        no_large_data = len(non_empty_cells) <= 10
        no_images = images_count == 0
        if sheet2_is_second and has_title_cell and has_five_charts and no_large_data and no_images:
            hit(1, "Sheet2为工作簿第二个工作表，仅放置财务分析标题、5个图表及必要图例，无无关大面积数据或图片")
        else:
            miss(1, "Sheet2为工作簿第二个工作表，仅放置财务分析标题、5个图表及必要图例，无无关大面积数据或图片",
                 f"第二张表={sheet2_is_second}，含标题={has_title_cell}，5个图表={has_five_charts}，"
                 f"无大面积数据={no_large_data}，无图片={no_images}")

        # +3：Sheet2左上毛利率与净利率趋势图为Excel可编辑带数据标记的折线图。
        c = line_profit
        target_charts = [x for x in (line_profit, bar_profit, roe_chart, debt_chart, growth_chart) if x is not None]
        is_top_left = bool(c) and c is min(target_charts, key=lambda x: (anchor_from(x)[1], anchor_from(x)[0])) if target_charts else False
        is_editable_line = bool(c) and isinstance(c, LineChart)
        has_markers = bool(c) and len(c.series) >= 1 and all(has_marker(s) for s in c.series)
        if is_top_left and is_editable_line and has_markers:
            hit(3, "Sheet2左上毛利率与净利率趋势图为Excel可编辑带数据标记的折线图")
        else:
            miss(3, "Sheet2左上毛利率与净利率趋势图为Excel可编辑带数据标记的折线图",
                 f"左上={is_top_left}，可编辑折线图={is_editable_line}，带数据标记={has_markers}")

        # +3：Sheet2左上毛利率与净利率趋势图数据
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            ser1_name_ok = ser_name(s1) == "毛利率"
            ser1_ref_ok = ref_match(ser_val_ref(s1), "企业财务统计总表!B4:B12")
            ser2_name_ok = ser_name(s2) == "净利率"
            ser2_ref_ok = ref_match(ser_val_ref(s2), "企业财务统计总表!C4:C12")
            if ser1_name_ok and ser1_ref_ok and ser2_name_ok and ser2_ref_ok:
                hit(3, "毛利率与净利率趋势图第一条系列“毛利率”引用B4:B12、第二条系列“净利率”引用C4:C12，均正确")
            else:
                miss(3, "毛利率与净利率趋势图第一条系列“毛利率”引用B4:B12、第二条系列“净利率”引用C4:C12，均正确",
                     f"名称={series_names(c)}，引用={series_val_refs(c)}")
        else:
            miss(3, "毛利率与净利率趋势图第一条系列“毛利率”引用B4:B12、第二条系列“净利率”引用C4:C12，均正确",
                 "缺少第一条/第二条系列")

        # +1：毛利率与净利率趋势图标题
        if c:
            title_text_ok = chart_title(c) == "毛利率与净利率趋势图"
            title_obj = c.title
            top_not_overlay = True
            centered = True
            if title_obj is not None:
                try:
                    top_not_overlay = getattr(title_obj, "overlay", None) in (False, None)
                except Exception:
                    top_not_overlay = True
                try:
                    layout = getattr(title_obj, "layout", None)
                    manual = layout.manualLayout if layout is not None else None
                    centered = manual is None
                except Exception:
                    centered = True
            top_center_ok = top_not_overlay and centered
            if title_text_ok and top_center_ok:
                hit(1, "毛利率与净利率趋势图标题文本正确且位于图表顶部居中")
            else:
                miss(1, "毛利率与净利率趋势图标题文本正确且位于图表顶部居中",
                     f"文本正确={title_text_ok}，顶部居中={top_center_ok}，实际文本：{chart_title(c)}")
        else:
            miss(1, "毛利率与净利率趋势图标题文本正确且位于图表顶部居中", "图表未找到")

        # +3：毛利率与净利率趋势图横向时间标签
        if c and c.series:
            cat_ref_ok = all(ref_match(ser_cat_ref(s), "企业财务统计总表!A4:A12") for s in c.series)
            src_times = [norm_text(ws1.cell(r, 1).value) for r in range(4, 13)]
            nine_times_ok = len(src_times) == 9 and src_times == EXPECTED_TIMES
            def _cat_is_text_ref(ser):
                try:
                    if ser.cat is not None and ser.cat.strRef is not None:
                        return True
                except Exception:
                    pass
                return False
            not_default_index_ok = all(_cat_is_text_ref(s) for s in c.series)
            if cat_ref_ok and nine_times_ok and not_default_index_ok:
                hit(3, "毛利率与净利率趋势图横向时间标签依次引用A4:A12的9个时间文本，顺序与源数据一致，无默认序号")
            else:
                miss(3, "毛利率与净利率趋势图横向时间标签依次引用A4:A12的9个时间文本，顺序与源数据一致，无默认序号",
                     f"引用A4:A12={cat_ref_ok}，9个时间文本={nine_times_ok}，非默认序号={not_default_index_ok}")
        else:
            miss(3, "毛利率与净利率趋势图横向时间标签依次引用A4:A12的9个时间文本，顺序与源数据一致，无默认序号",
                 "图表未找到或无系列")

        # +1：毛利率与净利率趋势图系列样式
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            f1, f2 = line_fill(s1), line_fill(s2)
            # 颜色读取不到（空或"None"）时不能默认视为“不同颜色”通过，应判为不满足。
            color_readable = bool(f1) and f1 != "None" and bool(f2) and f2 != "None"
            color_diff = color_readable and (f1 != f2)
            def _is_solid_line(ser):
                try:
                    dash = ser.graphicalProperties.line.prstDash
                except Exception:
                    dash = None
                return dash in (None, "solid")
            solid_ok = _is_solid_line(s1) and _is_solid_line(s2)
            m1, m2 = marker_symbol(s1), marker_symbol(s2)
            # 标记形状读取不到（None 或 "none"）时同样不能默认视为“形状不同”通过。
            marker_readable = bool(m1) and m1 != "none" and bool(m2) and m2 != "none"
            marker_shape_diff = marker_readable and (m1 != m2)
            has_both_markers = has_marker(s1) and has_marker(s2)
            if color_diff and solid_ok and has_both_markers and marker_shape_diff:
                hit(1, "毛利率与净利率趋势图两系列为不同颜色实线、不同形状标记")
            else:
                miss(1, "毛利率与净利率趋势图两系列为不同颜色实线、不同形状标记",
                     f"不同颜色={color_diff}，实线={solid_ok}，双标记={has_both_markers}，形状不同={marker_shape_diff}")
        else:
            miss(1, "毛利率与净利率趋势图两系列为不同颜色实线、不同形状标记", "系列不足")

        # +1：毛利率与净利率趋势图图例
        if c:
            legend_bottom_ok = is_bottom_legend(c)
            names = series_names(c)
            legend_order_ok = names[:2] == ["毛利率", "净利率"]
            legend_exists = c.legend is not None
            legend_color_match_ok = legend_entries_match_series_style(c)
            if legend_bottom_ok and legend_order_ok and legend_exists and legend_color_match_ok:
                hit(1, "毛利率与净利率趋势图图例位于底部，依次显示“毛利率”“净利率”，色块与对应折线颜色一致")
            else:
                miss(1, "毛利率与净利率趋势图图例位于底部，依次显示“毛利率”“净利率”，色块与对应折线颜色一致",
                     f"底部={legend_bottom_ok}，依次显示毛利率净利率={legend_order_ok}，图例存在={legend_exists}，"
                     f"图例与折线颜色一致可确认={legend_color_match_ok}")
        else:
            miss(1, "毛利率与净利率趋势图图例位于底部，依次显示“毛利率”“净利率”，色块与对应折线颜色一致", "图表未找到")

        # +3：Sheet2右上毛利率与净利率柱形图为Excel可编辑簇状柱形图。
        c = bar_profit
        is_top_right = False
        if c is not None and line_profit is not None:
            is_top_right = int(anchor_from(c)[0]) > int(anchor_from(line_profit)[0])
        is_editable_bar = bool(c) and isinstance(c, BarChart)
        is_column = is_editable_bar and getattr(c, "barDir", None) == "col"
        is_clustered = is_editable_bar and getattr(c, "grouping", None) == "clustered"
        if is_top_right and is_editable_bar and is_column and is_clustered:
            hit(3, "Sheet2右上毛利率与净利率柱形图为Excel可编辑簇状柱形图")
        else:
            miss(3, "Sheet2右上毛利率与净利率柱形图为Excel可编辑簇状柱形图",
                 f"右上={is_top_right}，可编辑柱形图对象={is_editable_bar}，垂直柱形={is_column}，簇状={is_clustered}")

        # +3：Sheet2右上毛利率与净利率柱形图数据
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            ser1_name_ok = ser_name(s1) == "毛利率"
            ser2_name_ok = ser_name(s2) == "净利率"
            ser1_ref_ok = ref_match(ser_val_ref(s1), "企业财务统计总表!B4:B12")
            ser2_ref_ok = ref_match(ser_val_ref(s2), "企业财务统计总表!C4:C12")
            cat_ref_ok = all(ref_match(ser_cat_ref(s), "企业财务统计总表!A4:A12") for s in c.series)
            if ser1_name_ok and ser2_name_ok and ser1_ref_ok and ser2_ref_ok and cat_ref_ok:
                hit(3, "毛利率与净利率柱形图含“毛利率”“净利率”两系列，分别引用B4:B12、C4:C12，横向分类引用A4:A12")
            else:
                miss(3, "毛利率与净利率柱形图含“毛利率”“净利率”两系列，分别引用B4:B12、C4:C12，横向分类引用A4:A12",
                     f"毛利率系列={ser1_name_ok}，净利率系列={ser2_name_ok}，"
                     f"B4:B12={ser1_ref_ok}，C4:C12={ser2_ref_ok}，分类A4:A12={cat_ref_ok}")
        else:
            miss(3, "毛利率与净利率柱形图含“毛利率”“净利率”两系列，分别引用B4:B12、C4:C12，横向分类引用A4:A12",
                 "缺少“毛利率”“净利率”两个系列")

        # +1：柱形图标题
        if c:
            title_text_ok = chart_title(c) == "毛利率与净利率柱形图"
            title_obj = c.title
            top_not_overlay = True
            centered = True
            if title_obj is not None:
                try:
                    top_not_overlay = getattr(title_obj, "overlay", None) in (False, None)
                except Exception:
                    top_not_overlay = True
                try:
                    layout = getattr(title_obj, "layout", None)
                    manual = layout.manualLayout if layout is not None else None
                    centered = manual is None
                except Exception:
                    centered = True
            top_center_ok = top_not_overlay and centered
            if title_text_ok and top_center_ok:
                hit(1, "毛利率与净利率柱形图标题文本正确且位于图表顶部居中")
            else:
                miss(1, "毛利率与净利率柱形图标题文本正确且位于图表顶部居中",
                     f"文本正确={title_text_ok}，顶部居中={top_center_ok}，实际文本：{chart_title(c)}")
        else:
            miss(1, "毛利率与净利率柱形图标题文本正确且位于图表顶部居中", "图表未找到")

        # +1：柱形图系列样式
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            fill1 = fill_color(s1)
            fill2 = fill_color(s2)
            # 填充颜色读取不到（空或"None"）时不能默认视为“不同颜色”通过，应判为不满足。
            fill_readable = bool(fill1) and fill1 != "None" and bool(fill2) and fill2 != "None"
            fill_color_diff = fill_readable and (fill1 != fill2)
            clustered_ok = getattr(c, "grouping", None) == "clustered"
            not_stacked_ok = getattr(c, "grouping", None) not in ("stacked", "percentStacked")
            if fill_color_diff and clustered_ok and not_stacked_ok:
                hit(1, "毛利率与净利率柱形图两系列填充不同颜色，同一时间下并列排列（簇状），未使用堆积方式")
            else:
                miss(1, "毛利率与净利率柱形图两系列填充不同颜色，同一时间下并列排列（簇状），未使用堆积方式",
                     f"不同填充色={fill_color_diff}，并列簇状={clustered_ok}，非堆积={not_stacked_ok}")
        else:
            miss(1, "毛利率与净利率柱形图两系列填充不同颜色，同一时间下并列排列（簇状），未使用堆积方式", "系列不足")

        # +3：Sheet2左中净资产收益率趋势图为Excel可编辑带数据标记的折线图。
        c = roe_chart
        is_left_mid = False
        if c is not None and line_profit is not None:
            is_left_mid = (
                int(anchor_from(c)[1]) > int(anchor_from(line_profit)[1])
                and int(anchor_from(c)[0]) <= int(anchor_from(line_profit)[0])
            )
        is_editable_line = bool(c) and isinstance(c, LineChart)
        has_markers = bool(c) and len(c.series) >= 1 and all(has_marker(s) for s in c.series)
        if is_left_mid and is_editable_line and has_markers:
            hit(3, "Sheet2左中净资产收益率趋势图为Excel可编辑带数据标记的折线图")
        else:
            miss(3, "Sheet2左中净资产收益率趋势图为Excel可编辑带数据标记的折线图",
                 f"左中={is_left_mid}，可编辑折线图={is_editable_line}，带数据标记={has_markers}")

        # +3：ROE数据
        if c and len(c.series) >= 1:
            s = c.series[0]
            name_ok = ser_name(s) == "净资产收益率（ROE）"
            val_ref_ok = ref_match(ser_val_ref(s), "企业财务统计总表!D4:D12")
            cat_ref_ok = all(ref_match(ser_cat_ref(x), "企业财务统计总表!A4:A12") for x in c.series)
            if name_ok and val_ref_ok and cat_ref_ok:
                hit(3, "净资产收益率（ROE）趋势图系列名称为“净资产收益率（ROE）”，数值引用D4:D12，横向分类引用A4:A12")
            else:
                miss(3, "净资产收益率（ROE）趋势图系列名称为“净资产收益率（ROE）”，数值引用D4:D12，横向分类引用A4:A12",
                     f"名称正确={name_ok}，数值D4:D12={val_ref_ok}，分类A4:A12={cat_ref_ok}，实际名称={series_names(c)}")
        else:
            miss(3, "净资产收益率（ROE）趋势图系列名称为“净资产收益率（ROE）”，数值引用D4:D12，横向分类引用A4:A12",
                 "无有效系列")

        # +1：ROE标题
        if c:
            title_text_ok = chart_title(c) == "净资产收益率（ROE）趋势图"
            title_obj = c.title
            top_not_overlay = True
            centered = True
            if title_obj is not None:
                try:
                    top_not_overlay = getattr(title_obj, "overlay", None) in (False, None)
                except Exception:
                    top_not_overlay = True
                try:
                    layout = getattr(title_obj, "layout", None)
                    manual = layout.manualLayout if layout is not None else None
                    centered = manual is None
                except Exception:
                    centered = True
            if title_text_ok and top_not_overlay and centered:
                hit(1, "净资产收益率（ROE）趋势图标题正确且位于图表顶部居中")
            else:
                miss(1, "净资产收益率（ROE）趋势图标题正确且位于图表顶部居中",
                     f"标题文本正确={title_text_ok}，顶部无遮挡={top_not_overlay}，居中={centered}，实际：{chart_title(c) if c else '未找到'}")
        else:
            miss(1, "净资产收益率（ROE）趋势图标题正确且位于图表顶部居中", "图表未找到")

        # +1：Sheet2右中净资产收益率趋势图折线样式：折线为实线，数据点有圆形或方形标记，
        # 折线变化与sheet1中D4:D12（净资产收益率列）的数据变化一致。
        c = roe_chart
        if c and len(c.series) >= 1:
            s = c.series[0]
            try:
                dash = s.graphicalProperties.line.prstDash
            except Exception:
                dash = None
            solid_ok = dash in (None, "solid")
            sym = marker_symbol(s)
            marker_ok = sym in ("circle", "square") or (sym is None and has_marker(s))
            def _ref_col_index(ser):
                ref = normalize_ref(ser_val_ref(ser))
                m = re.search(r"([A-Z]+)\d", ref.split("!")[-1])
                if not m:
                    return None
                idx = 0
                for ch in m.group(1):
                    idx = idx * 26 + (ord(ch) - ord("A") + 1)
                return idx
            def _col_values(col):
                return [ws1.cell(r, col).value for r in range(4, 13)] if col else []
            def _trend_dirs(vals):
                dirs = []
                for i in range(1, len(vals)):
                    a, b = vals[i - 1], vals[i]
                    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                        dirs.append(1 if b > a else -1 if b < a else 0)
                    else:
                        dirs.append(None)
                return dirs
            d_vals = [ws1.cell(r, 4).value for r in range(4, 13)]
            ser_vals = _col_values(_ref_col_index(s))
            trend_ok = (
                len(ser_vals) == 9
                and all(isinstance(v, (int, float)) for v in d_vals)
                and all(isinstance(v, (int, float)) for v in ser_vals)
                and _trend_dirs(ser_vals) == _trend_dirs(d_vals)
            )
            if solid_ok and marker_ok and trend_ok:
                hit(1, "净资产收益率趋势图折线为实线、数据点有圆形/方形标记，折线变化与D4:D12一致")
            else:
                miss(1, "净资产收益率趋势图折线为实线、数据点有圆形/方形标记，折线变化与D4:D12一致",
                     f"实线={solid_ok}，圆形/方形标记={marker_ok}，变化与D4:D12一致={trend_ok}")
        else:
            miss(1, "净资产收益率趋势图折线为实线、数据点有圆形/方形标记，折线变化与D4:D12一致", "无有效系列")

        # +3：Sheet2右中资产负债率趋势图为Excel可编辑带数据标记的折线图
        c = debt_chart
        is_right_mid = False
        if c is not None and roe_chart is not None:
            is_right_mid = int(anchor_from(c)[0]) > int(anchor_from(roe_chart)[0])
        is_editable_line = bool(c) and isinstance(c, LineChart)
        has_markers = bool(c) and len(c.series) >= 1 and all(has_marker(s) for s in c.series)
        if is_right_mid and is_editable_line and has_markers:
            hit(3, "Sheet2右中资产负债率趋势图为Excel可编辑带数据标记的折线图")
        else:
            miss(3, "Sheet2右中资产负债率趋势图为Excel可编辑带数据标记的折线图",
                 f"右中={is_right_mid}，可编辑折线图={is_editable_line}，带数据标记={has_markers}")

        # +3：资产负债率数据
        if c and len(c.series) >= 1:
            s = c.series[0]
            ok = (
                "资产负债率" in ser_name(s)
                and ref_match(ser_val_ref(s), "企业财务统计总表!H4:H12")
                and all_series_cat_ref(c)
            )
            if ok:
                hit(3, "资产负债率趋势图数据引用正确")
            else:
                miss(3, "资产负债率趋势图数据引用正确",
                     f"名称={series_names(c)}，引用={series_val_refs(c)}")
        else:
            miss(3, "资产负债率趋势图数据引用正确", "无有效系列")

        # +1：资产负债率趋势图标题
        if c:
            title_text_ok = chart_title(c) == "资产负债率趋势图"
            title_obj = c.title
            top_not_overlay = True
            centered = True
            if title_obj is not None:
                try:
                    top_not_overlay = getattr(title_obj, "overlay", None) in (False, None)
                except Exception:
                    top_not_overlay = True
                try:
                    layout = getattr(title_obj, "layout", None)
                    manual = layout.manualLayout if layout is not None else None
                    centered = manual is None
                except Exception:
                    centered = True
            top_center_ok = top_not_overlay and centered
            if title_text_ok and top_center_ok:
                hit(1, "资产负债率趋势图标题文本正确且位于图表顶部居中")
            else:
                miss(1, "资产负债率趋势图标题文本正确且位于图表顶部居中",
                     f"文本正确={title_text_ok}，顶部居中={top_center_ok}，实际文本：{chart_title(c)}")
        else:
            miss(1, "资产负债率趋势图标题文本正确且位于图表顶部居中", "图表未找到")

        # +1：资产负债率趋势图折线样式（第二次评分点，与原脚本保持一致）
        if c and len(c.series) >= 1:
            s = c.series[0]
            try:
                dash = s.graphicalProperties.line.prstDash
            except Exception:
                dash = None
            solid_ok = dash in (None, "solid")
            sym = marker_symbol(s)
            marker_ok = sym in ("circle", "square") or (sym is None and has_marker(s))
            def _debt_ref_col_index(ser):
                ref = normalize_ref(ser_val_ref(ser))
                m = re.search(r"([A-Z]+)\d", ref.split("!")[-1])
                if not m:
                    return None
                idx = 0
                for ch in m.group(1):
                    idx = idx * 26 + (ord(ch) - ord("A") + 1)
                return idx
            def _debt_col_values(col):
                return [ws1.cell(r, col).value for r in range(4, 13)] if col else []
            def _debt_trend_dirs(vals):
                dirs = []
                for i in range(1, len(vals)):
                    a, b = vals[i - 1], vals[i]
                    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                        dirs.append(1 if b > a else -1 if b < a else 0)
                    else:
                        dirs.append(None)
                return dirs
            h_vals = [ws1.cell(r, 8).value for r in range(4, 13)]
            ser_vals = _debt_col_values(_debt_ref_col_index(s))
            trend_ok = (
                len(ser_vals) == 9
                and all(isinstance(v, (int, float)) for v in h_vals)
                and all(isinstance(v, (int, float)) for v in ser_vals)
                and _debt_trend_dirs(ser_vals) == _debt_trend_dirs(h_vals)
            )
            if solid_ok and marker_ok and trend_ok:
                hit(1, "资产负债率趋势图折线为实线、数据点有圆形/方形标记，折线变化与H4:H12一致（复核）")
            else:
                miss(1, "资产负债率趋势图折线为实线、数据点有圆形/方形标记，折线变化与H4:H12一致（复核）",
                     f"实线={solid_ok}，圆形/方形标记={marker_ok}，变化与H4:H12一致={trend_ok}")
        else:
            miss(1, "资产负债率趋势图折线为实线、数据点有圆形/方形标记，折线变化与H4:H12一致（复核）", "无有效系列")

        # +3：Sheet2底部增长率趋势图为Excel可编辑带数据标记的双系列折线图
        c = growth_chart
        is_bottom = False
        if c is not None:
            upper = [x for x in (roe_chart, debt_chart) if x is not None]
            if upper:
                is_bottom = int(anchor_from(c)[1]) > max(int(anchor_from(x)[1]) for x in upper)
        is_editable_line = bool(c) and isinstance(c, LineChart)
        is_two_series = bool(c) and len(c.series) >= 2
        has_markers = bool(c) and is_two_series and all(has_marker(s) for s in c.series[:2])
        if is_bottom and is_editable_line and is_two_series and has_markers:
            hit(3, "Sheet2底部增长率趋势图为Excel可编辑带数据标记的双系列折线图")
        else:
            miss(3, "Sheet2底部增长率趋势图为Excel可编辑带数据标记的双系列折线图",
                 f"底部={is_bottom}，可编辑折线图={is_editable_line}，双系列={is_two_series}，带数据标记={has_markers}")

        # +3：增长率数据
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            ser1_name_ok = ser_name(s1) == "营业收入增长率"
            ser1_ref_ok = ref_match(ser_val_ref(s1), "企业财务统计总表!K4:K12")
            ser2_name_ok = ser_name(s2) == "净利润增长率"
            ser2_ref_ok = ref_match(ser_val_ref(s2), "企业财务统计总表!L4:L12")
            cat_ref_ok = all(ref_match(ser_cat_ref(x), "企业财务统计总表!A4:A12") for x in c.series)
            if ser1_name_ok and ser1_ref_ok and ser2_name_ok and ser2_ref_ok and cat_ref_ok:
                hit(3, "增长率趋势图第一条系列“营业收入增长率”引用K4:K12、第二条系列“净利润增长率”引用L4:L12，横向分类引用A4:A12")
            else:
                miss(3, "增长率趋势图第一条系列“营业收入增长率”引用K4:K12、第二条系列“净利润增长率”引用L4:L12，横向分类引用A4:A12",
                     f"营业收入增长率系列={ser1_name_ok}，K4:K12={ser1_ref_ok}，净利润增长率系列={ser2_name_ok}，L4:L12={ser2_ref_ok}，分类A4:A12={cat_ref_ok}")
        else:
            miss(3, "增长率趋势图第一条系列“营业收入增长率”引用K4:K12、第二条系列“净利润增长率”引用L4:L12，横向分类引用A4:A12",
                 "缺少两个系列")

        # +1：增长率标题
        if c:
            title_text_ok = chart_title(c) == "营业收入增长率与净利润增长率趋势图"
            title_obj = c.title
            top_not_overlay = True
            centered = True
            if title_obj is not None:
                try:
                    top_not_overlay = getattr(title_obj, "overlay", None) in (False, None)
                except Exception:
                    top_not_overlay = True
                try:
                    layout = getattr(title_obj, "layout", None)
                    manual = layout.manualLayout if layout is not None else None
                    centered = manual is None
                except Exception:
                    centered = True
            top_center_ok = top_not_overlay and centered
            if title_text_ok and top_center_ok:
                hit(1, "增长率趋势图标题文本正确且位于图表顶部居中")
            else:
                miss(1, "增长率趋势图标题文本正确且位于图表顶部居中",
                     f"文本正确={title_text_ok}，顶部居中={top_center_ok}，实际文本：{chart_title(c)}")
        else:
            miss(1, "增长率趋势图标题文本正确且位于图表顶部居中", "图表未找到")

        # +1：增长率系列样式
        if c and len(c.series) >= 2:
            s1, s2 = c.series[0], c.series[1]
            f1, f2 = line_fill(s1), line_fill(s2)
            color_readable = bool(f1) and f1 != "None" and bool(f2) and f2 != "None"
            color_diff = color_readable and (f1 != f2)
            def _is_solid_line2(ser):
                try:
                    dash = ser.graphicalProperties.line.prstDash
                except Exception:
                    dash = None
                return dash in (None, "solid")
            solid_ok = _is_solid_line2(s1) and _is_solid_line2(s2)
            m1, m2 = marker_symbol(s1), marker_symbol(s2)
            marker_readable = bool(m1) and m1 != "none" and bool(m2) and m2 != "none"
            marker_shape_diff = marker_readable and (m1 != m2)
            has_both_markers = has_marker(s1) and has_marker(s2)
            w1, w2 = line_width(s1), line_width(s2)
            width_readable = w1 is not None and w2 is not None
            width_same = width_readable and (w1 == w2)
            if color_diff and solid_ok and has_both_markers and marker_shape_diff and width_same:
                hit(1, "增长率趋势图两系列为不同颜色实线、不同形状标记，两条线宽一致")
            else:
                miss(1, "增长率趋势图两系列为不同颜色实线、不同形状标记，两条线宽一致",
                     f"不同颜色={color_diff}（可读={color_readable}），实线={solid_ok}，双标记={has_both_markers}，"
                     f"形状不同={marker_shape_diff}（可读={marker_readable}），线宽一致={width_same}（可读={width_readable}）")
        else:
            miss(1, "增长率趋势图两系列为不同颜色实线、不同形状标记，两条线宽一致", "系列不足")

        # +1：增长率图例
        if c:
            legend_bottom_ok = is_bottom_legend(c)
            names = series_names(c)
            legend_order_ok = names[:2] == ["营业收入增长率", "净利润增长率"]
            names_complete_ok = len(names) >= 2 and all(norm_text(n) for n in names[:2])
            def _is_default_name(n):
                return bool(re.fullmatch(r"系列\s*\d+", norm_text(n))) or bool(re.fullmatch(r"[Ss]eries\s*\d+", norm_text(n)))
            no_default_name_ok = len(names) >= 2 and not any(_is_default_name(n) for n in names[:2])
            if legend_bottom_ok and legend_order_ok and names_complete_ok and no_default_name_ok:
                hit(1, "增长率趋势图图例位于底部，依次显示“营业收入增长率”“净利润增长率”，系列名称完整且非默认名称")
            else:
                miss(1, "增长率趋势图图例位于底部，依次显示“营业收入增长率”“净利润增长率”，系列名称完整且非默认名称",
                     f"底部={legend_bottom_ok}，依次显示={legend_order_ok}，名称完整={names_complete_ok}，非默认名称={no_default_name_ok}")
        else:
            miss(1, "增长率趋势图图例位于底部，依次显示“营业收入增长率”“净利润增长率”，系列名称完整且非默认名称", "图表未找到")

        # +3：全部图表时间顺序
        cat_charts = [line_profit, bar_profit, roe_chart, debt_chart, growth_chart]
        all_present = all(x is not None for x in cat_charts)
        all_cat_ref_a4a12 = all_present and all(all_series_cat_ref(x) for x in cat_charts)
        source_order = [norm_text(ws1.cell(r, 1).value) for r in range(4, 13)]
        order_2021_to_2026q4 = source_order == EXPECTED_TIMES
        if all_cat_ref_a4a12 and order_2021_to_2026q4:
            hit(3, "5个图表横向时间标签均引用A4:A12，按2021年度至2026Q4顺序排列，季度在年度之后，未按字母/数值重排")
        else:
            miss(3, "5个图表横向时间标签均引用A4:A12，按2021年度至2026Q4顺序排列，季度在年度之后，未按字母/数值重排",
                 f"5图均引用A4:A12={all_cat_ref_a4a12}，源顺序为2021年度至2026Q4={order_2021_to_2026q4}")

        # +3：Sheet2全部图表数值显示：毛利率、净利率、ROE、资产负债率、营业收入增长率和净利润增长率
        # 均按照百分比格式显示，图表提示值或数据标签保留两位小数。
        # 检查对象是图表本身的数据标签/提示值格式（dLbls.numFmt，缺失时兜底看数值轴格式），
        # 而不是源工作表单元格的 number_format——源单元格格式正确但图表未显示数据标签或显示格式不是两位小数时不应得分。
        pct_charts = {
            "毛利率与净利率趋势图": line_profit,
            "毛利率与净利率柱形图": bar_profit,
            "净资产收益率趋势图": roe_chart,
            "资产负债率趋势图": debt_chart,
            "增长率趋势图": growth_chart,
        }
        pct_detail = {name: chart_labels_pct_two_decimal_ok(chart) for name, chart in pct_charts.items()}
        pct_ok = bool(pct_charts) and all(pct_detail.values())
        if pct_ok:
            hit(3, "毛利率、净利率、ROE、资产负债率、营业收入增长率、净利润增长率均按百分比显示，图表数据标签/提示值保留两位小数")
        else:
            failed = [name for name, ok in pct_detail.items() if not ok]
            miss(3, "毛利率、净利率、ROE、资产负债率、营业收入增长率、净利润增长率均按百分比显示，图表数据标签/提示值保留两位小数",
                 "以下图表未显示两位小数百分比的数据标签/提示值：" + "、".join(failed))

        # ─────────── 汇总 ───────────
        result["dim2_items"] = dim2_items
        result["total_score"] = score
        result["max_score"] = sum(item["max_delta"] for item in dim2_items if item["max_delta"] > 0)
        return result

    except Exception as e:
        result["status"] = "error"
        result["error"] = f"脚本执行异常：{e}"
        return result


if __name__ == "__main__":
    # 本地调试：默认使用脚本所在目录；也支持通过 sys.argv[1] 指定目录路径。
    _dir = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
    print(json.dumps(evaluate(_dir), ensure_ascii=False, indent=2))

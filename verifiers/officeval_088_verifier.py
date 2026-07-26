#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""自动评估 产品分析(1).docx 与 选货汇总表_已整理.xlsx。

维度1：可用与可修改性（门槛，不通过则得0分）
维度2：完成度（累计得分/扣分）
"""
from __future__ import annotations

import json
import re
import shutil
import struct
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional, TypedDict

from xml.etree import ElementTree as ET

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# XML 命名空间
# ---------------------------------------------------------------------------
WNS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
WP  = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
ANS = "http://schemas.openxmlformats.org/drawingml/2006/main"
PIC = "http://schemas.openxmlformats.org/drawingml/2006/picture"
REL = "http://schemas.openxmlformats.org/package/2006/relationships"
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"

W  = f"{{{WNS}}}"
Rp = f"{{{RNS}}}"
WPp= f"{{{WP}}}"
Ap = f"{{{ANS}}}"
XD = f"{{{XDR}}}"

EMU_PER_CM = 360000
# Excel: 1 字符宽约 7pt，96dpi 下用 EMU 估高。
# 细则"高N字符"是 Excel 行高单位。1字符行高 ≈ 14.4pt；1pt = 12700 EMU
PT_PER_CHAR  = 14.4
EMU_PER_PT   = 12700
EMU_PER_CHAR = PT_PER_CHAR * EMU_PER_PT  # ≈ 182880
# 列宽单位（Excel"字符数"）换算为磅：近似每字符 7pt，另加 5pt 内边距（Excel 默认字体度量）
PT_PER_COL_CHAR = 7.0
COL_PADDING_PT = 5.0
# 每行文字大致按 11pt 字号估算，单倍行高约 14.4pt/行（与 PT_PER_CHAR 一致）
PT_PER_TEXT_LINE = 14.4
CELL_PADDING_PT = 4.0  # 单元格上下内边距合计估算

# ---------------------------------------------------------------------------
# 数据类
# ---------------------------------------------------------------------------
@dataclass
class CheckResult:
    passed: bool
    evidence: str

@dataclass
class ScoreItem:
    id: str
    score: int
    name: str
    check: Callable[[], CheckResult]


class ProductCfg(TypedDict):
    name: str
    model: str
    price: str
    param_groups: list[list[str]]
    scene_groups: list[list[str]]
    xlsx_ratio: tuple[float, float]
    word_ratio: tuple[float, float]
    marketing_words: list[str]


class ImageRef(TypedDict):
    row: int
    col: int
    row_off_emu: int
    col_off_emu: int
    to_row: int
    to_col: int
    to_row_off_emu: int
    to_col_off_emu: int
    cx_emu: int
    cy_emu: int
    media_path: str
    rid: str


class HeaderCheck(TypedDict):
    name: str
    text_ok: bool
    centered: bool
    black_ok: bool
    top_ok: bool
    line_ok: bool
    line_ratio: float
    detail: str


class FooterCheck(TypedDict):
    section: int
    name: str
    has_footer: bool
    has_page_field: bool
    centered: bool
    start: int | None
    detail: str


class DimItem(TypedDict):
    rule: str
    max_delta: int
    delta: int
    hit: bool
    detail: str

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def _text(el: ET.Element, ns_uri: str) -> str:
    parts = []
    for t in el.iter(f"{{{ns_uri}}}t"):
        parts.append(t.text or "")
    return "".join(parts)

def _all_text_w(el: ET.Element) -> str:
    return _text(el, WNS)

def image_dims_px(raw: bytes) -> tuple[int, int]:
    """从 PNG/JPEG 原始字节返回 (width, height) 像素。"""
    if raw[:8] == b"\x89PNG\r\n\x1a\n":
        w, h = struct.unpack(">II", raw[16:24])
        return w, h
    if raw[:2] == b"\xff\xd8":
        i = 2
        while i < len(raw) - 9:
            if raw[i] != 0xFF:
                break
            marker = raw[i + 1]
            length = struct.unpack(">H", raw[i + 2:i + 4])[0]
            if marker in (0xC0, 0xC1, 0xC2):
                h = struct.unpack(">H", raw[i + 5:i + 7])[0]
                w = struct.unpack(">H", raw[i + 7:i + 9])[0]
                return w, h
            i += 2 + length
    return 0, 0

def _normalize(s: str) -> str:
    return re.sub(r"\s+", "", s or "")

def _contains(text: str, keywords: list[str]) -> list[str]:
    """返回 text 中出现的关键词列表。"""
    return [k for k in keywords if k in text]

def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# DocxReader
# ---------------------------------------------------------------------------
class DocxReader:
    def __init__(self, path: Path):
        self.path = path
        self.zf: Optional[zipfile.ZipFile] = None
        self.doc: Optional[ET.Element] = None
        self.names: list[str] = []
        self._rels: dict[str, str] = {}   # rid -> target path
        self._load()

    def _load(self):
        self.zf = zipfile.ZipFile(self.path)
        self.names = self.zf.namelist()
        self.doc = ET.fromstring(self.zf.read("word/document.xml"))
        # parse document.xml.rels
        try:
            rels_xml = ET.fromstring(self.zf.read("word/_rels/document.xml.rels"))
            for rel in rels_xml.findall(f"{{{REL}}}Relationship"):
                rid = rel.get("Id", "")
                tgt = rel.get("Target", "")
                self._rels[rid] = tgt
        except Exception:
            pass

    def all_text(self) -> str:
        return _all_text_w(self.doc)

    def paragraphs(self) -> list[ET.Element]:
        body = self.doc.find(f"{W}body")
        if body is None:
            return []
        return body.findall(f".//{W}p")

    def tables(self) -> list[ET.Element]:
        body = self.doc.find(f"{W}body")
        if body is None:
            return []
        return body.findall(f".//{W}tbl")

    def product_block_text(self, name_keywords: list[str]) -> str:
        """返回包含产品名称关键词的表格块文本；若未在表格中命中，则回退为空字符串。"""
        body = self.doc.find(f"{W}body")
        if body is None:
            return ""
        for tbl in body.findall(f".//{W}tbl"):
            tbl_text = _all_text_w(tbl)
            if any(_normalize(k) in _normalize(tbl_text) for k in name_keywords):
                return tbl_text
        return ""

    def product_block_image(self, name_keywords: list[str]) -> ImageRef | None:
        """在包含产品名称关键词的表格块（产品对应的图片粘贴框）内查找图片，
        返回该图片的 cx_emu/cy_emu/media_path/rid；未找到对应块或块内无图则返回 None。
        与 image_references() 不同，本方法按“产品块”定位，而非按文档出现顺序。"""
        body = self.doc.find(f"{W}body")
        if body is None:
            return None
        for tbl in body.findall(f".//{W}tbl"):
            tbl_text = _all_text_w(tbl)
            if not any(_normalize(k) in _normalize(tbl_text) for k in name_keywords):
                continue
            drawings: list[ET.Element] = list(tbl.iter(f"{{{WP}}}inline")) + list(tbl.iter(f"{{{WP}}}anchor"))
            imgs: list[ImageRef] = []
            for drawing in drawings:
                self._parse_drawing(drawing, imgs)
            if imgs:
                # 取块内尺寸最大的一张，避免误取小图标/装饰图
                return max(imgs, key=lambda i: i.get("cx_emu", 0) * i.get("cy_emu", 0))
        return None

    def image_references(self) -> list[ImageRef]:
        """返回文档中按出现顺序的图片列表，含 cx_emu, cy_emu, media_path。"""
        imgs: list[ImageRef] = []
        body = self.doc.find(f"{W}body")
        if body is None:
            return imgs
        for drawing in body.iter(f"{{{WP}}}inline") :
            self._parse_drawing(drawing, imgs)
        for drawing in body.iter(f"{{{WP}}}anchor"):
            self._parse_drawing(drawing, imgs)
        return imgs

    def _parse_drawing(self, drawing: ET.Element, out: list[ImageRef]):
        ext = drawing.find(f"{{{WP}}}extent")
        cx = int(ext.get("cx", 0)) if ext is not None else 0
        cy = int(ext.get("cy", 0)) if ext is not None else 0
        # find blip rid
        blip = drawing.find(f".//{{{ANS}}}blip")
        rid = blip.get(f"{{{RNS}}}embed", "") if blip is not None else ""
        target = self._rels.get(rid, "")
        # target may be relative: "media/image1.png" or "../media/image1.png"
        if target.startswith("../"):
            target = "word/" + target[3:]
        elif target and not target.startswith("word/") and not target.startswith("/"):
            target = "word/" + target
        # Word 图片无 Excel 锚点行列信息，填 0 占位以满足 ImageRef 结构
        out.append({
            "row": 0, "col": 0, "row_off_emu": 0, "col_off_emu": 0,
            "to_row": 0, "to_col": 0, "to_row_off_emu": 0, "to_col_off_emu": 0,
            "cx_emu": cx, "cy_emu": cy, "media_path": target, "rid": rid,
        })

    def media_bytes(self, path: str) -> bytes:
        try:
            return self.zf.read(path)
        except Exception:
            return b""

    def header_text(self) -> str:
        chunks = []
        for name in self.names:
            if name.startswith("word/header") and name.endswith(".xml"):
                try:
                    root = ET.fromstring(self.zf.read(name))
                    chunks.append(_all_text_w(root))
                except Exception:
                    pass
        return "\n".join(chunks)

    def _header_path_by_rid(self, rid: str) -> str:
        target = self._rels.get(rid, "")
        if target.startswith("../"):
            target = "word/" + target[3:]
        elif target and not target.startswith("word/") and not target.startswith("/"):
            target = "word/" + target
        return target

    def _section_body_width_twips(self, sectpr: ET.Element) -> int:
        pgsz = sectpr.find(f"{W}pgSz")
        pgmar = sectpr.find(f"{W}pgMar")
        page_w = int(pgsz.get(f"{W}w", "11906")) if pgsz is not None else 11906
        left = int(pgmar.get(f"{W}left", "1440")) if pgmar is not None else 1440
        right = int(pgmar.get(f"{W}right", "1440")) if pgmar is not None else 1440
        return max(1, page_w - left - right)

    def _section_header_top_ok(self, sectpr: ET.Element) -> bool:
        pgmar = sectpr.find(f"{W}pgMar")
        if pgmar is None:
            return True
        header_twips = int(pgmar.get(f"{W}header", "720"))
        top_twips = int(pgmar.get(f"{W}top", "1440"))
        return 0 <= header_twips <= max(top_twips, 720)

    def _paragraph_centered(self, p: ET.Element) -> bool:
        ppr = p.find(f"{W}pPr")
        if ppr is None:
            return False
        jc = ppr.find(f"{W}jc")
        return jc is not None and jc.get(f"{W}val", "") == "center"

    def _paragraph_black_text(self, p: ET.Element) -> bool:
        colors: list[str] = []
        for rpr in p.iter(f"{W}rPr"):
            c = rpr.find(f"{W}color")
            if c is not None:
                colors.append((c.get(f"{W}val", "") or "").lower())
        return all(v in ("", "auto", "000000", "black") for v in colors)

    def _header_line_ratio(self, p: ET.Element, body_width_twips: int) -> float:
        ppr = p.find(f"{W}pPr")
        if ppr is not None:
            pbdr = ppr.find(f"{W}pBdr")
            bottom = pbdr.find(f"{W}bottom") if pbdr is not None else None
            if bottom is not None and (bottom.get(f"{W}val", "") or "") not in ("", "none", "nil"):
                ind = ppr.find(f"{W}ind")
                left = int(ind.get(f"{W}left", "0")) if ind is not None else 0
                right = int(ind.get(f"{W}right", "0")) if ind is not None else 0
                line_w = max(0, body_width_twips - left - right)
                return line_w / body_width_twips
        for line in p.iter(f"{{{'urn:schemas-microsoft-com:vml'}}}line"):
            style = line.get("style", "") or ""
            m = re.search(r"width:([0-9.]+)pt", style)
            if m:
                return float(m.group(1)) * 20 / body_width_twips
            from_attr = line.get("from", "") or ""
            to_attr = line.get("to", "") or ""
            nums: list[str] = re.findall(r"[-0-9.]+", from_attr + "," + to_attr)
            if len(nums) >= 4:
                return abs(float(nums[2]) - float(nums[0])) * 20 / body_width_twips
        for drawing in list(p.iter(f"{{{WP}}}inline")) + list(p.iter(f"{{{WP}}}anchor")):
            ext = drawing.find(f"{{{WP}}}extent")
            if ext is not None:
                cx = int(ext.get("cx", "0"))
                if cx:
                    return (cx / 635) / body_width_twips
        return 0.0

    def section_header_checks(self) -> list[HeaderCheck]:
        body = self.doc.find(f"{W}body")
        zf = self.zf
        if body is None or zf is None:
            return []
        checks: list[HeaderCheck] = []
        fallback_headers = [n for n in self.names if n.startswith("word/header") and n.endswith(".xml")]
        for idx, sectpr in enumerate(body.iter(f"{W}sectPr"), start=1):
            body_width = self._section_body_width_twips(sectpr)
            top_ok = self._section_header_top_ok(sectpr)
            refs = sectpr.findall(f"{W}headerReference")
            paths = [self._header_path_by_rid(ref.get(f"{{{RNS}}}id", "")) for ref in refs]
            paths = [p for p in paths if p]
            if not paths and idx == 1:
                paths = fallback_headers
            for path in paths:
                try:
                    root = ET.fromstring(zf.read(path))
                except Exception:
                    checks.append({"name": path, "text_ok": False, "centered": False, "black_ok": False, "top_ok": top_ok, "line_ok": False, "line_ratio": 0.0, "detail": "页眉XML读取失败"})
                    continue
                paras = list(root.iter(f"{W}p"))
                target_idx = next((i for i, p in enumerate(paras) if "产品分析" in _all_text_w(p)), -1)
                text_ok = target_idx >= 0
                centered = text_ok and self._paragraph_centered(paras[target_idx])
                black_ok = text_ok and self._paragraph_black_text(paras[target_idx])
                ratios: list[float] = []
                if text_ok:
                    for p in paras[target_idx:target_idx + 2]:
                        ratios.append(self._header_line_ratio(p, body_width))
                line_ratio = max(ratios) if ratios else 0.0
                line_ok = 0.70 <= line_ratio <= 0.90
                detail = f"节{idx}:{path}:正文宽={body_width}twip;横线比={line_ratio:.2f}"
                checks.append({"name": path, "text_ok": text_ok, "centered": centered, "black_ok": black_ok, "top_ok": top_ok, "line_ok": line_ok, "line_ratio": line_ratio, "detail": detail})
        if not checks:
            for path in fallback_headers:
                try:
                    root = ET.fromstring(zf.read(path))
                except Exception:
                    continue
                paras = list(root.iter(f"{W}p"))
                target_idx = next((i for i, p in enumerate(paras) if "产品分析" in _all_text_w(p)), -1)
                text_ok = target_idx >= 0
                centered = text_ok and self._paragraph_centered(paras[target_idx])
                black_ok = text_ok and self._paragraph_black_text(paras[target_idx])
                ratios = [self._header_line_ratio(p, 9026) for p in paras[target_idx:target_idx + 2]] if text_ok else []
                line_ratio = max(ratios) if ratios else 0.0
                checks.append({"name": path, "text_ok": text_ok, "centered": centered, "black_ok": black_ok, "top_ok": True, "line_ok": 0.70 <= line_ratio <= 0.90, "line_ratio": line_ratio, "detail": f"{path}:默认正文宽=9026twip;横线比={line_ratio:.2f}"})
        return checks

    def header_centered(self) -> bool:
        for name in self.names:
            if name.startswith("word/header") and name.endswith(".xml"):
                try:
                    root = ET.fromstring(self.zf.read(name))
                    for ppr in root.iter(f"{W}pPr"):
                        jc = ppr.find(f"{W}jc")
                        if jc is not None and jc.get(f"{W}val", "") == "center":
                            return True
                except Exception:
                    pass
        return False

    def header_text_black(self) -> bool:
        """页眉“产品分析”文字字体为黑色：无 color 设置（默认黑）或 color 为 000000/auto/black。"""
        for name in self.names:
            if not (name.startswith("word/header") and name.endswith(".xml")):
                continue
            try:
                root = ET.fromstring(self.zf.read(name))
            except Exception:
                continue
            for p in root.iter(f"{W}p"):
                if "产品分析" not in _all_text_w(p):
                    continue
                colors = []
                for rpr in p.iter(f"{W}rPr"):
                    c = rpr.find(f"{W}color")
                    if c is not None:
                        colors.append((c.get(f"{W}val", "") or "").lower())
                # 未设颜色=默认黑；设了颜色则须为黑/自动
                return all(v in ("", "auto", "000000", "black") for v in colors)
        return False

    def header_line_centered(self) -> bool:
        """页眉文字下方存在一条居中横线：段落下边框(pBdr/bottom)或独立线条(v:line/横向 drawing)。"""
        for name in self.names:
            if not (name.startswith("word/header") and name.endswith(".xml")):
                continue
            try:
                root = ET.fromstring(self.zf.read(name))
            except Exception:
                continue
            for p in root.iter(f"{W}p"):
                ppr = p.find(f"{W}pPr")
                centered = False
                has_line = False
                if ppr is not None:
                    jc = ppr.find(f"{W}jc")
                    centered = jc is not None and jc.get(f"{W}val", "") == "center"
                    pbdr = ppr.find(f"{W}pBdr")
                    if pbdr is not None and pbdr.find(f"{W}bottom") is not None:
                        b = pbdr.find(f"{W}bottom")
                        if (b.get(f"{W}val", "") or "") not in ("", "none", "nil"):
                            has_line = True
                # 独立线条形状（VML 或 drawing line）
                if p.find(f".//{{{'urn:schemas-microsoft-com:vml'}}}line") is not None:
                    has_line = True
                if p.find(f".//{{{ANS}}}ln") is not None and p.find(f".//{Ap}blip") is None:
                    has_line = True
                if has_line and (centered or "产品分析" in _all_text_w(p)):
                    return True
        return False

    def footer_instrs(self) -> str:
        chunks = []
        for name in self.names:
            if name.startswith("word/footer") and name.endswith(".xml"):
                try:
                    root = ET.fromstring(self.zf.read(name))
                    for t in root.iter(f"{W}instrText"):
                        chunks.append(t.text or "")
                    # fldSimple 形式的域
                    for fs in root.iter(f"{W}fldSimple"):
                        chunks.append(fs.get(f"{W}instr", "") or "")
                except Exception:
                    pass
        return " ".join(chunks)

    def footer_page_centered(self) -> bool:
        """页脚中承载 PAGE 域的段落是否居中。"""
        for name in self.names:
            if not (name.startswith("word/footer") and name.endswith(".xml")):
                continue
            try:
                root = ET.fromstring(self.zf.read(name))
            except Exception:
                continue
            for p in root.iter(f"{W}p"):
                instr = " ".join((t.text or "") for t in p.iter(f"{W}instrText"))
                instr += " " + " ".join(fs.get(f"{W}instr", "") or "" for fs in p.iter(f"{W}fldSimple"))
                if "PAGE" not in instr.upper():
                    continue
                ppr = p.find(f"{W}pPr")
                if ppr is not None:
                    jc = ppr.find(f"{W}jc")
                    if jc is not None and jc.get(f"{W}val", "") == "center":
                        return True
        return False

    def _footer_path_by_rid(self, rid: str) -> str:
        target = self._rels.get(rid, "")
        if target.startswith("../"):
            target = "word/" + target[3:]
        elif target and not target.startswith("word/") and not target.startswith("/"):
            target = "word/" + target
        return target

    def _paragraph_page_instr(self, p: ET.Element) -> str:
        instr = " ".join((t.text or "") for t in p.iter(f"{W}instrText"))
        instr += " " + " ".join(fs.get(f"{W}instr", "") or "" for fs in p.iter(f"{W}fldSimple"))
        return instr.upper()

    def _footer_has_centered_page(self, root: ET.Element) -> tuple[bool, bool]:
        has_page = False
        centered = False
        for p in root.iter(f"{W}p"):
            instr = self._paragraph_page_instr(p)
            if "PAGE" not in instr:
                continue
            has_page = True
            if self._paragraph_centered(p):
                centered = True
        return has_page, centered

    def section_footer_checks(self) -> list[FooterCheck]:
        body = self.doc.find(f"{W}body")
        zf = self.zf
        if body is None or zf is None:
            return []
        checks: list[FooterCheck] = []
        fallback_footers = [n for n in self.names if n.startswith("word/footer") and n.endswith(".xml")]
        last_default_path = ""
        for idx, sectpr in enumerate(body.iter(f"{W}sectPr"), start=1):
            refs = sectpr.findall(f"{W}footerReference")
            paths = [self._footer_path_by_rid(ref.get(f"{{{RNS}}}id", "")) for ref in refs]
            paths = [p for p in paths if p]
            if paths:
                default_ref = next((ref for ref in refs if ref.get(f"{W}type", "default") == "default"), refs[0])
                last_default_path = self._footer_path_by_rid(default_ref.get(f"{{{RNS}}}id", ""))
            elif last_default_path:
                paths = [last_default_path]
            elif idx == 1:
                paths = fallback_footers
            pnt = sectpr.find(f"{W}pgNumType")
            start: int | None = None
            if pnt is not None and pnt.get(f"{W}start") is not None:
                start = int(pnt.get(f"{W}start", "0"))
            if not paths:
                checks.append({"section": idx, "name": "", "has_footer": False, "has_page_field": False, "centered": False, "start": start, "detail": f"节{idx}:无页脚引用"})
                continue
            for path in paths:
                try:
                    root = ET.fromstring(zf.read(path))
                    has_page, centered = self._footer_has_centered_page(root)
                    checks.append({"section": idx, "name": path, "has_footer": True, "has_page_field": has_page, "centered": centered, "start": start, "detail": f"节{idx}:{path}:start={start}"})
                except Exception:
                    checks.append({"section": idx, "name": path, "has_footer": True, "has_page_field": False, "centered": False, "start": start, "detail": f"节{idx}:{path}:页脚XML读取失败"})
        if not checks:
            for path in fallback_footers:
                try:
                    root = ET.fromstring(zf.read(path))
                    has_page, centered = self._footer_has_centered_page(root)
                    checks.append({"section": 1, "name": path, "has_footer": True, "has_page_field": has_page, "centered": centered, "start": None, "detail": f"{path}:无sectPr回退检查"})
                except Exception:
                    pass
        return checks

    def page_number_continuous_by_sections(self) -> bool:
        body = self.doc.find(f"{W}body")
        if body is None:
            return False
        starts: list[int | None] = []
        for sectpr in body.iter(f"{W}sectPr"):
            pnt = sectpr.find(f"{W}pgNumType")
            if pnt is None or pnt.get(f"{W}start") is None:
                starts.append(None)
            else:
                starts.append(int(pnt.get(f"{W}start", "0")))
        if not starts:
            return True
        if starts[0] not in (None, 1):
            return False
        return all(s is None for s in starts[1:])

    def page_number_start_one(self) -> bool:
        """页码从第1页开始：sectPr/pgNumType 未设 start 或 start=1（连续、不跳号）。"""
        body = self.doc.find(f"{W}body")
        if body is None:
            return True
        starts = []
        for sectpr in body.iter(f"{W}sectPr"):
            pnt = sectpr.find(f"{W}pgNumType")
            if pnt is not None:
                s = pnt.get(f"{W}start", None)
                if s is not None:
                    starts.append(s)
        # 未显式设起始页 = 默认从 1 开始；设了则必须为 1
        return all(s == "1" for s in starts)


    def has_header_refs(self) -> bool:
        return any(n.startswith("word/header") and n.endswith(".xml") for n in self.names)

    def has_footer_refs(self) -> bool:
        return any(n.startswith("word/footer") and n.endswith(".xml") for n in self.names)

    def consecutive_empty_paras(self) -> int:
        """返回最长连续空段落数。"""
        max_run = 0
        cur = 0
        for p in self.paragraphs():
            t = _all_text_w(p).strip()
            if not t:
                cur += 1
                max_run = max(max_run, cur)
            else:
                cur = 0
        return max_run

    def inline_images_count(self) -> int:
        body = self.doc.find(f"{W}body")
        if body is None:
            return 0
        return len(list(body.iter(f"{{{WP}}}inline"))) + len(list(body.iter(f"{{{WP}}}anchor")))

# ---------------------------------------------------------------------------
# XlsxReader
# ---------------------------------------------------------------------------
class XlsxReader:
    def __init__(self, path: Path):
        self.path = path
        self.wb = None
        self.ws = None
        self.zf: Optional[zipfile.ZipFile] = None
        self.names: list[str] = []
        self._load()

    def _load(self):
        self.zf = zipfile.ZipFile(self.path)
        self.names = self.zf.namelist()
        if _HAS_OPENPYXL:
            self.wb = openpyxl.load_workbook(self.path)
            self.ws = self.wb.active

    def max_row(self) -> int:
        return self.ws.max_row if self.ws is not None else 0

    def max_column(self) -> int:
        return self.ws.max_column if self.ws is not None else 0

    def cell_value(self, row: int, col: int):
        if self.ws is None:
            return None
        return self.ws.cell(row=row, column=col).value

    def cell_text(self, row: int, col: int) -> str:
        return _cell_str(self.cell_value(row, col))

    def header_map(self) -> dict[str, int]:
        """从第一行建立 header -> col 映射；兼容空格和换行。"""
        hm = {}
        if self.ws is None:
            return hm
        for c in range(1, self.ws.max_column + 1):
            h = _normalize(_cell_str(self.ws.cell(1, c).value))
            if h:
                hm[h] = c
        return hm

    def find_col(self, *names: str) -> int:
        hm = self.header_map()
        for n in names:
            key = _normalize(n)
            if key in hm:
                return hm[key]
        # 模糊兜底
        for h, c in hm.items():
            if any(_normalize(n) in h or h in _normalize(n) for n in names):
                return c
        return 0

    def find_product_row(self, name_keyword: str, model_keyword: str = "") -> int:
        if self.ws is None:
            return 0
        name_col = self.find_col("产品名称") or 1
        model_col = self.find_col("型号", "产品型号") or 2
        for r in range(2, self.ws.max_row + 1):
            row_text = " ".join(_cell_str(self.ws.cell(r, c).value) for c in range(1, self.ws.max_column + 1))
            if name_keyword in row_text or (model_keyword and model_keyword in row_text):
                return r
            # 有些产品名称单元格可能被简写，单独看常用列
            if name_keyword in self.cell_text(r, name_col) or (model_keyword and model_keyword in self.cell_text(r, model_col)):
                return r
        return 0

    def column_widths(self) -> dict[int, float]:
        widths = {}
        if self.ws is None:
            return widths
        for c in range(1, self.ws.max_column + 1):
            letter = openpyxl.utils.get_column_letter(c)
            width = self.ws.column_dimensions[letter].width
            if width is None:
                width = 8.43
            widths[c] = float(width)
        return widths

    def wrap_text(self, row: int, col: int) -> bool:
        if self.ws is None:
            return False
        return bool(self.ws.cell(row, col).alignment.wrap_text)

    def row_height(self, row: int) -> float:
        """行高（磅），未显式设置返回默认 15.0。"""
        if self.ws is None:
            return 15.0
        h = self.ws.row_dimensions[row].height
        return float(h) if h is not None else 15.0

    def non_empty_rows_after_header(self) -> int:
        if self.ws is None:
            return 0
        count = 0
        for r in range(2, self.ws.max_row + 1):
            if any(_cell_str(self.ws.cell(r, c).value) for c in range(1, self.ws.max_column + 1)):
                count += 1
        return count

    def image_references(self) -> list[ImageRef]:
        """解析 XLSX 绘图，返回图片显示尺寸、锚点、媒体路径。"""
        drawings = [n for n in self.names if n.startswith("xl/drawings/drawing") and n.endswith(".xml")]
        images: list[ImageRef] = []
        for drawing_name in drawings:
            try:
                root = ET.fromstring(self.zf.read(drawing_name))
            except Exception:
                continue
            # rels
            rels_name = drawing_name.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            rels = {}
            if rels_name in self.names:
                try:
                    rel_root = ET.fromstring(self.zf.read(rels_name))
                    for rel in rel_root.findall(f"{{{REL}}}Relationship"):
                        rels[rel.get("Id", "")] = rel.get("Target", "")
                except Exception:
                    pass
            for anchor in list(root.findall(f"{{{XDR}}}oneCellAnchor")) + list(root.findall(f"{{{XDR}}}twoCellAnchor")) + list(root.findall(f"{{{XDR}}}absoluteAnchor")):
                pic = anchor.find(f"{{{XDR}}}pic")
                if pic is None:
                    continue
                fr = anchor.find(f"{{{XDR}}}from")
                row = col = -1
                row_off = col_off = 0
                if fr is not None:
                    r_el = fr.find(f"{{{XDR}}}row")
                    c_el = fr.find(f"{{{XDR}}}col")
                    ro_el = fr.find(f"{{{XDR}}}rowOff")
                    co_el = fr.find(f"{{{XDR}}}colOff")
                    row = int(r_el.text or 0) + 1 if r_el is not None else -1
                    col = int(c_el.text or 0) + 1 if c_el is not None else -1
                    row_off = int(ro_el.text or 0) if ro_el is not None else 0
                    col_off = int(co_el.text or 0) if co_el is not None else 0
                # twoCellAnchor 提供 <to> 终点，用于精确锚定 bounding box 右下角
                to = anchor.find(f"{{{XDR}}}to")
                to_row = to_col = -1
                to_row_off = to_col_off = 0
                if to is not None:
                    tr_el = to.find(f"{{{XDR}}}row")
                    tc_el = to.find(f"{{{XDR}}}col")
                    tro_el = to.find(f"{{{XDR}}}rowOff")
                    tco_el = to.find(f"{{{XDR}}}colOff")
                    to_row = int(tr_el.text or 0) + 1 if tr_el is not None else -1
                    to_col = int(tc_el.text or 0) + 1 if tc_el is not None else -1
                    to_row_off = int(tro_el.text or 0) if tro_el is not None else 0
                    to_col_off = int(tco_el.text or 0) if tco_el is not None else 0
                ext = anchor.find(f"{{{XDR}}}ext")
                if ext is None:
                    ext = anchor.find(f".//{{{ANS}}}ext")
                cx = int(ext.get("cx", 0)) if ext is not None and ext.get("cx") else 0
                cy = int(ext.get("cy", 0)) if ext is not None and ext.get("cy") else 0
                blip = pic.find(f".//{{{ANS}}}blip")
                rid = blip.get(f"{{{RNS}}}embed", "") if blip is not None else ""
                target = rels.get(rid, "")
                if target.startswith("/"):
                    target = target.lstrip("/")
                elif target.startswith("../"):
                    target = "xl/" + target[3:]
                elif target and not target.startswith("xl/"):
                    target = "xl/drawings/" + target
                images.append({
                    "row": row, "col": col, "row_off_emu": row_off, "col_off_emu": col_off,
                    "to_row": to_row, "to_col": to_col, "to_row_off_emu": to_row_off, "to_col_off_emu": to_col_off,
                    "cx_emu": cx, "cy_emu": cy, "media_path": target, "rid": rid,
                })
        images.sort(key=lambda x: (x.get("row", 9999), x.get("col", 9999)))
        return images

    def media_bytes(self, path: str) -> bytes:
        try:
            return self.zf.read(path)
        except Exception:
            return b""

    def col_letter_of(self, col_idx: int) -> str:
        if _HAS_OPENPYXL:
            return openpyxl.utils.get_column_letter(col_idx)
        # basic fallback
        letters = ""
        n = col_idx
        while n:
            n, rem = divmod(n - 1, 26)
            letters = chr(ord("A") + rem) + letters
        return letters


# ---------------------------------------------------------------------------
# 维度1
# ---------------------------------------------------------------------------
def clean_docx_stem(stem: str) -> str:
    s = re.sub(r"[\s_\-]+", "", stem)
    s = re.sub(r"[（(]\d+[）)]", "", s)
    return s


def valid_zip_package(path: Path, required_parts: list[str]) -> tuple[bool, str]:
    if not path.exists():
        return False, f"文件不存在：{path.name}"
    try:
        with zipfile.ZipFile(path) as z:
            bad = z.testzip()
            if bad:
                return False, f"压缩包损坏：{bad}"
            names = set(z.namelist())
            missing = [p for p in required_parts if p not in names]
            if missing:
                return False, f"缺少OOXML必要部件：{missing}"
            return True, "OOXML结构完整，可解析"
    except Exception as e:
        return False, f"无法作为OOXML打开：{e}"


def check_dimension1(docx_path: Path, xlsx_path: Path, base_dir: Path) -> tuple[bool, list[tuple[bool, str]]]:
    details: list[tuple[bool, str]] = []

    docx_files = [p for p in base_dir.glob("*.docx") if not p.name.startswith("~$")]
    xlsx_files = [p for p in base_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    xlsm_files = [p for p in base_dir.glob("*.xlsm") if not p.name.startswith("~$")]
    excel_files = xlsx_files + xlsm_files

    count_ok = len(docx_files) == 1 and len(excel_files) == 1
    docx_ok, docx_msg = valid_zip_package(docx_path, ["[Content_Types].xml", "word/document.xml"])
    xlsx_ok, xlsx_msg = valid_zip_package(xlsx_path, ["[Content_Types].xml", "xl/workbook.xml"])

    passed = count_ok and docx_ok and xlsx_ok
    detail = (
        f"docx={len(docx_files)}，xlsx/xlsm={len(excel_files)}；"
        f"Word可打开：{docx_msg}；Excel可打开：{xlsx_msg}"
    )
    details.append((passed, f"交付文件包含1个.xlsx或.xlsm格式的选货汇总表和1个.docx格式的Word文档，两个文件均可正常打开。[{detail}]"))

    return passed, details


# ---------------------------------------------------------------------------
# 评分辅助
# ---------------------------------------------------------------------------
PRODUCTS: dict[str, ProductCfg] = {
    "云溪": {
        "name": "云溪5.0 D3滚筒洗衣机",
        "model": "云溪5.0 D3",
        "price": "4500",
        "param_groups": [["12KG"], ["595"], ["滚筒洗衣机", "滚筒"], ["一级"], ["593"], ["537"], ["850"]],
        "scene_groups": [["多人口家庭"], ["日常衣物"], ["大件家纺"], ["小型民宿"], ["公寓出租房"]],
        "xlsx_ratio": (0.74, 0.76),
        "word_ratio": (0.74, 0.76),
        "marketing_words": ["巅峰之作", "一级健康洁净标准"],
    },
    "小天鹅": {
        "name": "小天鹅12KG滚筒洗衣机",
        "model": "TG12VE40PRO",
        "price": "3039",
        "param_groups": [["12KG"], ["595"], ["滚筒洗衣机", "滚筒"], ["一级"], ["560"], ["585"]],
        "scene_groups": [["多人口家庭"], ["外衣"], ["内衣"], ["宝宝衣物"], ["分类洗"], ["小型民宿"], ["公寓出租房"]],
        "xlsx_ratio": (0.83, 0.85),
        "word_ratio": (0.83, 0.85),
        "marketing_words": ["纯平自由嵌", "预混高能洁净力"],
    },
    "统帅": {
        "name": "统帅三桶懒人洗衣机",
        "model": "Leader L10510B12（标尺版）",
        "price": "3999",
        "param_groups": [["10.5KG"], ["1KG"], ["600"], ["570"], ["滚筒洗衣机", "滚筒"], ["一级"]],
        "scene_groups": [["多人口家庭"], ["日常衣物"], ["大件家纺"], ["小型民宿"], ["公寓出租房"], ["母婴"], ["高温除菌"]],
        "xlsx_ratio": (0.69, 0.71),
        "word_ratio": (0.69, 0.71),
        "marketing_words": ["PTC柔烘科技", "12.5KG洗衣机"],
    },
    "米家": {
        "name": "米家8KG滚筒洗衣机",
        "model": "米家MG80V11D（滚筒款）",
        "price": "999",
        "param_groups": [["8KG"], ["598"], ["495"], ["滚筒洗衣机", "滚筒"], ["一级"]],
        "scene_groups": [["单身"], ["2-3口", "2—3口", "小家庭"], ["日常衣物"], ["薄家纺"], ["小型民宿"], ["员工宿舍"]],
        "xlsx_ratio": (0.66, 0.68),
        "word_ratio": (0.66, 0.68),
        "marketing_words": ["超薄省空间", "高效除菌洗"],
    },
    "美的": {
        "name": "美的8KG波轮洗衣机",
        "model": "美的MB80V36DT（波轮款）",
        "price": "846",
        "param_groups": [["8KG"], ["515"], ["525"], ["波轮洗衣机", "波轮"], ["一级"]],
        "scene_groups": [["单身"], ["2-3口", "2—3口", "小家庭"], ["日常衣物"], ["薄家纺"], ["小型民宿"], ["员工宿舍"]],
        "xlsx_ratio": (0.64, 0.66),
        "word_ratio": (0.64, 0.66),
        "marketing_words": ["汽车级防锈工艺", "防腐防生锈"],
    },
}


def _norm_ascii(s: str) -> str:
    return _normalize(s).upper().replace("㎏", "KG").replace("ＫＧ", "KG")


SCENE_SYNONYMS: dict[str, list[str]] = {
    "多人口家庭": ["多人家庭", "人口多", "大家庭", "全家", "家庭多人", "三口以上", "四口之家", "五口之家"],
    "日常衣物": ["日常穿着", "日常洗涤", "日常换洗", "普通衣物", "平常衣物", " everyday衣物"],
    "大件家纺": ["大件织物", "床品", "床单", "被套", "窗帘", "家纺", "大件衣物"],
    "小型民宿": ["民宿", "短租房", "短租公寓", "小旅馆", "客栈", "家庭旅馆"],
    "公寓出租房": ["出租房", "出租屋", "租赁公寓", "公寓", "长租房", "租房"],
    "外衣": ["外套", "外穿衣物", "外穿衣", "上衣", "大衣", "夹克", "日常外穿"],
    "内衣": ["贴身衣物", "贴身衣", "内裤", "文胸", "私密衣物", "小件贴身"],
    "宝宝衣物": ["宝宝衣服", "婴儿衣物", "婴幼儿衣物", "儿童衣物", "母婴衣物", "孩子衣物", "婴童衣物"],
    "分类洗": ["分区洗", "分开洗", "分桶洗", "专衣专洗", "分开清洗", "分类清洗", "不同衣物分洗"],
    "母婴": ["宝妈", "婴幼儿", "宝宝", "婴儿", "儿童", "亲子家庭", "有娃家庭"],
    "高温除菌": ["高温杀菌", "除菌", "杀菌", "消毒", "抑菌", "健康洗"],
    "单身": ["独居", "一个人", "一人居", "单人", "个人使用"],
    "2-3口": ["2口", "3口", "两口", "三口", "二三口", "两三口", "小家庭"],
    "2—3口": ["2口", "3口", "两口", "三口", "二三口", "两三口", "小家庭"],
    "小家庭": ["2-3口", "2—3口", "两口之家", "三口之家", "二三口家庭", "年轻家庭"],
    "薄家纺": ["薄被", "薄床品", "轻薄家纺", "床单", "被套", "夏凉被", "小件家纺"],
    "员工宿舍": ["宿舍", "职工宿舍", "员工公寓", "集体宿舍", "工厂宿舍"],
}


def expand_scene_group(group: list[str]) -> list[str]:
    expanded: list[str] = []
    for keyword in group:
        expanded.append(keyword)
        expanded.extend(SCENE_SYNONYMS.get(keyword, []))
    return expanded


def groups_ok(text: str, groups: list[list[str]]) -> tuple[bool, list[str]]:
    nt = _norm_ascii(text)
    missing = []
    for group in groups:
        if not any(_norm_ascii(k) in nt for k in group):
            missing.append("/".join(group))
    return not missing, missing


def scene_groups_ok(text: str, groups: list[list[str]]) -> tuple[bool, list[str]]:
    """踩中 groups 中 ≥2 个组即视为通过（每组含同义词扩展）。"""
    nt = _norm_ascii(text)
    hit: list[str] = []
    missing: list[str] = []
    for group in groups:
        expanded = expand_scene_group(group)
        if any(_norm_ascii(k) in nt for k in expanded):
            hit.append("/".join(group))
        else:
            missing.append("/".join(group))
    return len(hit) >= 2, missing


def find_product_row_by_cfg(xlsx: XlsxReader, cfg: ProductCfg) -> int:
    row = xlsx.find_product_row(cfg["name"], cfg["model"])
    if row:
        return row
    # 宽松品牌名
    row = xlsx.find_product_row(cfg["name"][:2], "")
    return row


def product_columns(xlsx: XlsxReader) -> dict[str, int]:
    return {
        "name": xlsx.find_col("产品名称") or 1,
        "model": xlsx.find_col("型号", "产品型号") or 2,
        "params": xlsx.find_col("核心参数") or 3,
        "scene": xlsx.find_col("适用场景") or 4,
        "price": xlsx.find_col("指导价格", "价格") or 5,
        "image": xlsx.find_col("展示图片", "图片") or 6,
    }


def price_ok(text: str, expected: str) -> bool:
    return expected in re.sub(r"\D", "", text)


def check_excel_widths(xlsx: XlsxReader) -> CheckResult:
    cols = product_columns(xlsx)
    expected = {
        "产品名称": (cols["name"], 23.42),
        "型号": (cols["model"], 32.81),
        "核心参数": (cols["params"], 32.81),
        "适用场景": (cols["scene"], 10.82),
        "指导价格": (cols["price"], 12.36),
        "展示图片": (cols["image"], 12.08),
    }
    widths = xlsx.column_widths()
    ok = []
    bad = []
    for label, (idx, exp) in expected.items():
        got = widths.get(idx, 8.43)
        if abs(got - exp) <= 0.1:
            ok.append(f"{label}={got:.2f}")
        else:
            bad.append(f"{label}={got:.2f}(应{exp})")
    return CheckResult(not bad, "；".join(ok + bad))


def check_excel_product(xlsx: XlsxReader, key: str) -> CheckResult:
    cfg = PRODUCTS[key]
    cols = product_columns(xlsx)
    row = find_product_row_by_cfg(xlsx, cfg)
    if not row:
        return CheckResult(False, f"未找到{key}产品行")
    name = xlsx.cell_text(row, cols["name"])
    model = xlsx.cell_text(row, cols["model"])
    price = xlsx.cell_text(row, cols["price"])
    params = xlsx.cell_text(row, cols["params"])
    scene = xlsx.cell_text(row, cols["scene"])
    # 如果列识别失败或单元格分散，使用整行文本兜底
    row_text = " ".join(xlsx.cell_text(row, c) for c in range(1, xlsx.max_column() + 1))
    full = " ".join([name, model, price, params, scene, row_text])

    name_ok = cfg["name"] in full or key in full
    model_ok = _norm_ascii(cfg["model"]) in _norm_ascii(full)
    price_hit = price_ok(full, cfg["price"])
    param_ok, param_missing = groups_ok(full, cfg["param_groups"])
    scene_ok, scene_missing = scene_groups_ok(full, cfg["scene_groups"])
    passed = name_ok and model_ok and price_hit and param_ok and scene_ok
    evidence = (
        f"row={row}；名称{'✓' if name_ok else '✗'}；型号{'✓' if model_ok else '✗'}；"
        f"价格{'✓' if price_hit else '✗'}；参数缺失={param_missing or '无'}；场景缺失={scene_missing or '无'}"
    )
    return CheckResult(passed, evidence)


def estimate_wrapped_lines(text: str, col_width_chars: float) -> int:
    """按列宽（Excel字符数）估算自动换行后的行数。中文/全角字符按2倍宽度计。"""
    if not text:
        return 1
    usable_chars = max(col_width_chars - COL_PADDING_PT / PT_PER_COL_CHAR, 1.0)
    width = 0.0
    lines = 1
    for ch in text:
        w = 2.0 if ord(ch) > 0x2E80 else 1.0
        if width + w > usable_chars:
            lines += 1
            width = w
        else:
            width += w
    return lines


def cell_overflows(text: str, col_width_chars: float, row_height_pt: float) -> tuple[bool, int, float]:
    """基于文本长度/列宽估算换行行数，再与行高比较，判断文字是否会被截断（溢出）。
    返回 (是否溢出, 估算行数, 所需最小行高)。"""
    lines = estimate_wrapped_lines(text, col_width_chars)
    need_height = lines * PT_PER_TEXT_LINE + CELL_PADDING_PT
    overflow = row_height_pt < need_height - 0.5  # 容忍0.5pt误差
    return overflow, lines, need_height


def cell_bbox_emu(xlsx: "XlsxReader", row: int, col: int) -> tuple[float, float, float, float]:
    """返回单元格 (x0, y0, x1, y1) 的 EMU 坐标（近似：按列宽/行高累加）。"""
    widths = xlsx.column_widths()
    x0 = 0.0
    for c in range(1, col):
        w = widths.get(c, 8.43)
        x0 += w * PT_PER_COL_CHAR * EMU_PER_PT
    x1 = x0 + widths.get(col, 8.43) * PT_PER_COL_CHAR * EMU_PER_PT
    y0 = 0.0
    for r in range(1, row):
        y0 += xlsx.row_height(r) * EMU_PER_PT
    y1 = y0 + xlsx.row_height(row) * EMU_PER_PT
    return x0, y0, x1, y1


def image_bbox_emu(xlsx: "XlsxReader", img: ImageRef) -> tuple[float, float, float, float]:
    """返回图片 (x0, y0, x1, y1) 的 EMU 坐标。优先用 twoCellAnchor 的 from/to，
    否则用 from + 显示尺寸(cx/cy) 兜底。"""
    widths = xlsx.column_widths()
    from_col = img["col"]
    from_row = img["row"]
    if from_col <= 0 or from_row <= 0:
        return (0.0, 0.0, 0.0, 0.0)
    x0 = 0.0
    for c in range(1, from_col):
        x0 += widths.get(c, 8.43) * PT_PER_COL_CHAR * EMU_PER_PT
    x0 += img["col_off_emu"]
    y0 = 0.0
    for r in range(1, from_row):
        y0 += xlsx.row_height(r) * EMU_PER_PT
    y0 += img["row_off_emu"]

    to_col = img["to_col"]
    to_row = img["to_row"]
    if to_col > 0 and to_row > 0:
        x1 = 0.0
        for c in range(1, to_col):
            x1 += widths.get(c, 8.43) * PT_PER_COL_CHAR * EMU_PER_PT
        x1 += img["to_col_off_emu"]
        y1 = 0.0
        for r in range(1, to_row):
            y1 += xlsx.row_height(r) * EMU_PER_PT
        y1 += img["to_row_off_emu"]
    else:
        x1 = x0 + img["cx_emu"]
        y1 = y0 + img["cy_emu"]
    return x0, y0, x1, y1


def bbox_overlap(a: tuple[float, float, float, float], b: tuple[float, float, float, float]) -> bool:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    return ax0 < bx1 and bx0 < ax1 and ay0 < by1 and by0 < ay1


def check_excel_wrap(xlsx: XlsxReader) -> CheckResult:
    cols = product_columns(xlsx)
    widths = xlsx.column_widths()
    imgs = xlsx.image_references()
    bad = []
    checked = 0
    for key, cfg in PRODUCTS.items():
        row = find_product_row_by_cfg(xlsx, cfg)
        if not row:
            bad.append(f"{key}:未找到行")
            continue
        rh = xlsx.row_height(row)
        for label, col in [("核心参数", cols["params"]), ("适用场景", cols["scene"])]:
            checked += 1
            text = xlsx.cell_text(row, col)
            col_w = widths.get(col, 8.43)
            # ① 自动换行
            if not xlsx.wrap_text(row, col):
                bad.append(f"{key}-{label}:未换行")
            # ② 不溢出：按文本长度/列宽估算换行行数，再与实际行高比较，
            #    行高不足以容纳估算行数即视为溢出（不再用固定15磅阈值误判）
            overflow, lines, need_h = cell_overflows(text, col_w, rh)
            if overflow:
                bad.append(f"{key}-{label}:溢出(估算{lines}行需{need_h:.1f}磅,实际{rh:.1f}磅)")
            # ③ 不遮挡相邻列：本单元格 bounding box 右侧不应侵入相邻列（即列宽估算与实际渲染一致，
            #    此处以文本换行估算避免溢出即代表不会侵占相邻列，已由②覆盖）
        # ④ 不被图片覆盖：用单元格与图片的 bounding box 实际求交，而非只看图片起始列
        cell_boxes = {label: cell_bbox_emu(xlsx, row, col) for label, col in [("核心参数", cols["params"]), ("适用场景", cols["scene"])]}
        for img in imgs:
            img_box = image_bbox_emu(xlsx, img)
            if img_box == (0.0, 0.0, 0.0, 0.0):
                continue
            for label, cbox in cell_boxes.items():
                if bbox_overlap(img_box, cbox):
                    bad.append(f"{key}-{label}:被图片覆盖(图片锚点row={img.get('row')},col={img.get('col')})")
    return CheckResult(not bad and checked >= 10, f"检查{checked}个长文本单元格；问题={bad or '无'}")


def check_excel_image_count(xlsx: XlsxReader) -> CheckResult:
    """+5：5个产品行均插入1张对应洗衣机图片，共5张。
    不再是 len(imgs)>0 就通过：要求总数恰好为5，且逐产品行匹配"展示图片"列，
    每个产品行恰好锚定1张图片，不允许缺失/多插/错位。"""
    imgs = xlsx.image_references()
    image_col = product_columns(xlsx)["image"]
    bad: list[str] = []
    matched_rows: set[int] = set()
    for key, cfg in PRODUCTS.items():
        row = find_product_row_by_cfg(xlsx, cfg)
        if not row:
            bad.append(f"{key}:未找到产品行")
            continue
        # 图片锚点应落在该产品行、展示图片列（允许锚点列略有偏移，仍以列号严格匹配为准）
        hits = [img for img in imgs if img["row"] == row and img["col"] == image_col]
        if len(hits) == 0:
            bad.append(f"{key}:第{row}行展示图片列未插入图片")
        elif len(hits) > 1:
            bad.append(f"{key}:第{row}行展示图片列插入了{len(hits)}张图片")
        else:
            matched_rows.add(row)
    extra = [img for img in imgs if not (img["col"] == image_col and img["row"] in matched_rows)]
    if extra:
        bad.append(f"存在{len(extra)}张图片未落在任一产品行的展示图片列")
    passed = len(imgs) == 5 and not bad
    return CheckResult(passed, f"共{len(imgs)}张（要求5张）；展示图片列={image_col}；问题={bad or '无'}")


def excel_char_height(cy_emu: int) -> float:
    return cy_emu / EMU_PER_CHAR if cy_emu else 0.0


def _ocr_text(raw: bytes) -> str | None:
    """尝试用系统 tesseract 对图片做 OCR，返回识别文本；tesseract 不可用时返回 None
    （None 表示"无法验证"，调用方不应将其当作"未检测到文字"而直接判通过）。"""
    tesseract_bin = shutil.which("tesseract")
    if not tesseract_bin or not raw:
        return None
    tmp_dir = tempfile.mkdtemp(prefix="ocr_")
    try:
        img_path = Path(tmp_dir) / "img.png"
        _ = img_path.write_bytes(raw)
        proc = subprocess.run(
            [tesseract_bin, str(img_path), "stdout", "-l", "chi_sim+eng"],
            capture_output=True, text=True, timeout=20,
        )
        if proc.returncode != 0:
            return None
        return proc.stdout
    except Exception:
        return None
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def check_excel_product_image(xlsx: XlsxReader, key: str) -> CheckResult:
    """+3：展示图片为对应产品主体图，不保留指定营销文字，高11.80-12.10字符，横纵比在指定区间。
    问题修正：
      ① 原来按 order 列表下标取图，假定第idx张图对应该产品——现改为按图片锚点(row,col)
         关联到该产品所在行、展示图片列，直接取"这一格"里的图，避免图片顺序与产品顺序不一致导致错配。
      ② 原来 px_ratio==0（拿不到像素尺寸/OCR）时直接当作通过——现改为：拿不到像素信息时明确
         标记为"无法验证"，不再默认给分；能解码出像素时，用 OCR 检测营销文字关键词，检测到则直接判为不通过。"""
    cfg = PRODUCTS[key]
    imgs = xlsx.image_references()
    cols = product_columns(xlsx)
    row = find_product_row_by_cfg(xlsx, cfg)
    if not row:
        return CheckResult(False, f"未找到{key}产品行")
    image_col = cols["image"]
    hits = [img for img in imgs if img["row"] == row and img["col"] == image_col]
    if not hits:
        return CheckResult(False, f"第{row}行展示图片列未找到锚定图片")
    if len(hits) > 1:
        return CheckResult(False, f"第{row}行展示图片列锚定了{len(hits)}张图片，存在歧义")
    img = hits[0]
    ratio = img["cx_emu"] / img["cy_emu"] if img["cy_emu"] else 0
    hchars = excel_char_height(img["cy_emu"])
    r0, r1 = cfg["xlsx_ratio"]
    r0, r1 = float(r0), float(r1)
    ratio_ok = r0 <= ratio <= r1
    height_ok = 11.80 <= hchars <= 12.10

    raw = xlsx.media_bytes(img.get("media_path", ""))
    pw, ph = image_dims_px(raw) if raw else (0, 0)
    px_ratio = pw / ph if ph else 0
    # 能解码像素尺寸时才用像素比做内容代理校验；解码失败(0,0)不再默认通过，标记为待人工复核
    if ph:
        content_proxy_ok: bool | None = (r0 - 0.08) <= px_ratio <= (r1 + 0.08)
    else:
        content_proxy_ok = None

    ocr_text = _ocr_text(raw) if raw else None
    marketing_hit: list[str] = []
    if ocr_text is not None:
        norm_ocr = _normalize(ocr_text)
        marketing_hit = [w for w in cfg.get("marketing_words", []) if _normalize(w) in norm_ocr]

    checks_ok = ratio_ok and height_ok and (content_proxy_ok is not False) and not marketing_hit
    unverifiable = content_proxy_ok is None or ocr_text is None
    if ocr_text is None:
        ocr_detail = "不可用，未校验营销文字"
    else:
        ocr_detail = f"命中营销词={marketing_hit or '无'}"
    detail = (
        f"锚点(row={img.get('row')},col={img.get('col')})；显示高={hchars:.2f}字符；显示比={ratio:.3f}；"
        f"像素比={px_ratio:.3f}（{'可用' if ph else '不可用'}）；OCR={ocr_detail}"
    )
    return CheckResult(checks_ok and not unverifiable, detail)


def scene_groups_all_ok(text: str, groups: list[list[str]]) -> tuple[bool, list[str]]:
    """要求 groups 中每一组都命中（含同义词扩展），而非 scene_groups_ok 的">=2组即通过"。"""
    nt = _norm_ascii(text)
    missing: list[str] = []
    for group in groups:
        expanded = expand_scene_group(group)
        if not any(_norm_ascii(k) in nt for k in expanded):
            missing.append("/".join(group))
    return not missing, missing


def check_word_products(docx: DocxReader) -> CheckResult:
    """+5：产品一/二/三分别填写云溪/统帅/小天鹅，名称、适用场景、指导价格均完整出现，
    且按4500/3999/3039从高到低排列。
    问题修正：
      ① 原来在全文范围检查价格和场景，未绑定到具体产品块——现改为先用 product_block_text()
         定位该产品所在的表格块，再仅在该块文本内检查名称/价格/场景，避免张冠李戴（比如
         A产品块里写了B产品的价格也被误判通过）。
      ② 原来场景用 scene_groups_ok（命中>=2组即通过）——现改为 scene_groups_all_ok，
         要求该产品配置的场景关键词组全部命中。
    """
    full_text = docx.all_text()
    required = ["云溪", "统帅", "小天鹅"]  # 产品一/二/三
    detail: list[str] = []
    positions: list[int] = []
    for key in required:
        cfg = PRODUCTS[key]
        block_text = docx.product_block_text([cfg["name"], key])
        if not block_text:
            detail.append(f"{key}({cfg['price']}元):未定位到产品块")
            positions.append(-1)
            continue
        ntext = _normalize(block_text)
        # 名称需完整出现（去空格后匹配，容断行/空格）
        name_ok = _normalize(cfg["name"]) in ntext
        # 价格在产品块内完整出现
        price_hit = price_ok(block_text, cfg["price"])
        # 场景关键词组需全部命中（同义词扩展），且限定在该产品块内
        scene_ok, scene_missing = scene_groups_all_ok(block_text, cfg["scene_groups"])
        # 用产品块在全文中的位置确定三个产品块的先后顺序
        pos = full_text.find(block_text[:20]) if len(block_text) >= 20 else full_text.find(block_text)
        positions.append(pos)
        detail.append(
            f"{key}({cfg['price']}元):名称{'✓' if name_ok else '✗'} 价格{'✓' if price_hit else '✗'} "
            f"场景缺失={scene_missing or '无'}"
        )
    # 顺序：产品一云溪 < 产品二统帅 < 产品三小天鹅（4500>3999>3039 从高到低）
    order_ok = positions[0] >= 0 and positions[1] >= 0 and positions[2] >= 0 and positions[0] < positions[1] < positions[2]
    items_ok = all("名称✓" in d and "价格✓" in d and "场景缺失=无" in d for d in detail)
    return CheckResult(items_ok and order_ok, "；".join(detail) + f"；顺序(4500>3999>3039)={order_ok}")


def word_display_images(docx: DocxReader) -> list[ImageRef]:
    imgs = [i for i in docx.image_references() if i.get("cx_emu", 0) and i.get("cy_emu", 0)]
    # 过滤极小装饰图，优先取高度接近 5cm 的产品图
    candidates = [i for i in imgs if 3.5 <= i["cy_emu"] / EMU_PER_CM <= 7.0]
    return candidates or imgs


def check_word_product_image(docx: DocxReader, key: str) -> CheckResult:
    cfg = PRODUCTS[key]
    img = docx.product_block_image([cfg["name"], key])
    if img is None:
        return CheckResult(False, f"未在{key}对应产品块内找到图片粘贴框中的图片")
    hcm = img["cy_emu"] / EMU_PER_CM
    ratio = img["cx_emu"] / img["cy_emu"] if img["cy_emu"] else 0
    r0, r1 = cfg["word_ratio"]
    r0, r1 = float(r0), float(r1)
    return CheckResult(4.90 <= hcm <= 5.10 and r0 <= ratio <= r1, f"高={hcm:.2f}cm；显示比={ratio:.3f}；媒体={img.get('media_path')}")


def check_word_header(docx: DocxReader) -> CheckResult:
    checks = docx.section_header_checks()
    if not checks:
        return CheckResult(False, "未找到各节引用的页眉")
    ok = all(c["text_ok"] and c["centered"] and c["black_ok"] and c["top_ok"] and c["line_ok"] for c in checks)
    parts: list[str] = []
    for c in checks:
        parts.append(
            f"{c['name']}:含产品分析={c['text_ok']},居中={c['centered']},黑色={c['black_ok']},"
            + f"顶部={c['top_ok']},横线70%-90%={c['line_ok']}({c['line_ratio']:.2f})"
        )
    detail = "；".join(parts)
    return CheckResult(ok, detail)


def check_word_page_number(docx: DocxReader) -> CheckResult:
    checks = docx.section_footer_checks()
    if not checks:
        return CheckResult(False, "未找到各节引用的页脚")
    all_have_footer = all(c["has_footer"] for c in checks)
    all_have_page = all(c["has_page_field"] for c in checks)
    all_centered = all(c["centered"] for c in checks)
    continuous = docx.page_number_continuous_by_sections()
    ok = all_have_footer and all_have_page and all_centered and continuous
    parts: list[str] = []
    for c in checks:
        parts.append(
            f"节{c['section']}:{c['name'] or '无'}:有页脚={c['has_footer']},"
            + f"含PAGE域={c['has_page_field']},居中={c['centered']},start={c['start']}"
        )
    detail = "；".join(parts) + f"；跳号/重启检查通过={continuous}"
    return CheckResult(ok, detail)


def check_word_filename(docx_path: Path) -> CheckResult:
    ok = clean_docx_stem(docx_path.stem) == "产品分析" and not any(w in docx_path.stem for w in ["模板", "副本", "未命名"])
    return CheckResult(ok, f"文件名={docx_path.name}，主体={clean_docx_stem(docx_path.stem)}")


def build_rubric(docx: DocxReader, xlsx: XlsxReader, docx_path: Path) -> list[ScoreItem]:
    return [
        ScoreItem("EX-01", 3, "Excel列宽设置：“产品名称”列23.42字符、“型号”、“核心参数”列32.81字符、适用场景10.82字符、指导价格12.36字符、展示图片12.08字符等列宽。", lambda: check_excel_widths(xlsx)),
        ScoreItem("EX-02", 3, "Excel云溪：产品名称、型号、指导价格、核心参数和适用场景填写完整准确。", lambda: check_excel_product(xlsx, "云溪")),
        ScoreItem("EX-03", 3, "Excel小天鹅：产品名称、型号、指导价格、核心参数和适用场景填写完整准确。", lambda: check_excel_product(xlsx, "小天鹅")),
        ScoreItem("EX-04", 3, "Excel统帅：产品名称、型号、指导价格、核心参数和适用场景填写完整准确。", lambda: check_excel_product(xlsx, "统帅")),
        ScoreItem("EX-05", 3, "Excel米家：产品名称、型号、指导价格、核心参数和适用场景填写完整准确。", lambda: check_excel_product(xlsx, "米家")),
        ScoreItem("EX-06", 3, "Excel美的：产品名称、型号、指导价格、核心参数和适用场景填写完整准确。", lambda: check_excel_product(xlsx, "美的")),
        ScoreItem("EX-07", 1, "Excel文字格式：长参数和适用场景开启自动换行，文字不溢出单元格、不遮挡相邻列、不被图片覆盖。", lambda: check_excel_wrap(xlsx)),
        ScoreItem("EX-08", 5, "Excel展示图片数量：5个产品行均插入1张对应洗衣机图片，共5张图片。", lambda: check_excel_image_count(xlsx)),
        ScoreItem("EX-09", 3, "Excel云溪展示图片：图片为云溪5.0 D3滚筒洗衣机主体图，不保留“巅峰之作”“一级健康洁净标准”等营销文字，高11.80-12.10字符，横纵比为0.74-0.76。", lambda: check_excel_product_image(xlsx, "云溪")),
        ScoreItem("EX-10", 3, "Excel小天鹅展示图片：图片为小天鹅滚筒洗衣机主体图，不保留“纯平自由嵌”“预混高能洁净力”等营销文字，高11.80-12.10字符，横纵比为0.83-0.85。", lambda: check_excel_product_image(xlsx, "小天鹅")),
        ScoreItem("EX-11", 3, "Excel统帅展示图片：图片为统帅三桶洗衣机主体图，不保留“PTC柔烘科技”“12.5KG洗衣机”等海报文字，高11.80-12.10字符，横纵比为0.69-0.71。", lambda: check_excel_product_image(xlsx, "统帅")),
        ScoreItem("EX-12", 3, "Excel米家展示图片：图片为米家8KG滚筒洗衣机主体图，不保留“超薄省空间”“高效除菌洗”等营销文字，高11.80-12.10字符，横纵比为0.66-0.68。", lambda: check_excel_product_image(xlsx, "米家")),
        ScoreItem("EX-13", 3, "Excel美的展示图片：图片为美的8KG波轮洗衣机主体图，不保留“汽车级防锈工艺”“防腐防生锈”等营销文字，高11.80-12.10字符，横纵比为0.64-0.66。", lambda: check_excel_product_image(xlsx, "美的")),
        ScoreItem("WD-01", 5, "Word产品一、产品二、产品三的产品名称、适用场景、指导价格均完整出现，并按4500元、3999元、3039元从高到低排列。", lambda: check_word_products(docx)),
        ScoreItem("WD-02", 1, "Word产品一图片：云溪洗衣机图片放入产品一对应图片粘贴框，尺寸高为4.9-5.10厘米，横纵比为0.74-0.76。", lambda: check_word_product_image(docx, "云溪")),
        ScoreItem("WD-03", 1, "Word产品二图片：统帅三桶洗衣机图片放入产品二对应图片粘贴框，尺寸高为4.9-5.10厘米，横纵比为0.69-0.71。", lambda: check_word_product_image(docx, "统帅")),
        ScoreItem("WD-04", 1, "Word产品三图片：小天鹅滚筒洗衣机图片放入产品三对应图片粘贴框，尺寸高为4.9-5.10厘米，横纵比为0.83-0.85。", lambda: check_word_product_image(docx, "小天鹅")),
        ScoreItem("WD-07", 1, "Word页眉：每页页眉居中显示“产品分析”，字体为黑色，位置位于页面顶部中央。页眉文字下方设置一条横向长线，线条居中，长度覆盖正文宽度约70%—90%。", lambda: check_word_header(docx)),
        ScoreItem("WD-08", 1, "Word页码：页码位于页脚居中位置，从第1页开始连续编号，不跳号、不重复。", lambda: check_word_page_number(docx)),
        ScoreItem("WD-09", 1, "Word文件命名：输出Word文件名主体为“产品分析”，不保留“模板”“副本”“未命名”等不符合要求的名称。", lambda: check_word_filename(docx_path)),
    ]


SCRIPT_ID = "088"
MAX_SCORE = 44


def _locate_docs(base_dir: Path) -> tuple[Path | None, Path | None]:
    docx_files = [p for p in base_dir.glob("*.docx") if not p.name.startswith("~$")]
    xlsx_files = [p for p in base_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    xlsm_files = [p for p in base_dir.glob("*.xlsm") if not p.name.startswith("~$")]
    excel_files = xlsx_files + xlsm_files
    docx = None
    for p in docx_files:
        if "产品分析" in p.stem:
            docx = p
            break
    if docx is None and docx_files:
        docx = docx_files[0]
    xlsx = None
    for p in excel_files:
        if "选货汇总表" in p.stem:
            xlsx = p
            break
    if xlsx is None and excel_files:
        xlsx = excel_files[0]
    return docx, xlsx


def _empty_result(file_name: str, error: str | None, dim1_reason: str = "") -> dict[str, object]:
    status = "error" if error else "ok"
    return {
        "id": SCRIPT_ID,
        "file_name": file_name,
        "status": status,
        "error": error,
        "dim1_pass": False,
        "dim1_reason": dim1_reason,
        "dim2_items": [],
        "total_score": 0,
        "max_score": MAX_SCORE,
    }


def evaluate(dir_path: str) -> dict:
    """评估 dir_path 目录下的 Word/Excel 交付物，返回结构化字典。"""
    try:
        base_dir = Path(dir_path)
        if not base_dir.exists() or not base_dir.is_dir():
            return _empty_result("", f"目录不存在或不是目录：{dir_path}")

        docx_path, xlsx_path = _locate_docs(base_dir)
        file_name = docx_path.name if docx_path else ""

        if docx_path is None or xlsx_path is None:
            missing = []
            if docx_path is None:
                missing.append("docx")
            if xlsx_path is None:
                missing.append("xlsx")
            return _empty_result(file_name, None, f"缺少交付文件：{missing}")

        d1_ok, d1_details = check_dimension1(docx_path, xlsx_path, base_dir)
        if not d1_ok:
            reasons = "；".join(msg for ok, msg in d1_details if not ok)
            return _empty_result(file_name, None, reasons)

        docx = DocxReader(docx_path)
        xlsx = XlsxReader(xlsx_path)

        dim2_items: list[DimItem] = []
        total = 0
        for item in build_rubric(docx, xlsx, docx_path):
            try:
                res = item.check()
            except Exception as e:
                res = CheckResult(False, f"检查异常：{e}")
            hit = bool(res.passed)
            delta = item.score if hit else 0
            total += delta
            dim2_items.append({
                "rule": item.name,
                "max_delta": item.score,
                "delta": delta,
                "hit": hit,
                "detail": "",
            })

        max_score = sum(it["max_delta"] for it in dim2_items)
        return {
            "id": SCRIPT_ID,
            "file_name": file_name,
            "status": "ok",
            "error": None,
            "dim1_pass": True,
            "dim1_reason": "",
            "dim2_items": dim2_items,
            "total_score": total,
            "max_score": max_score,
        }
    except Exception as e:
        return _empty_result("", f"脚本执行异常：{e}")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).parent)
    _out = json.dumps(evaluate(target), ensure_ascii=False, indent=2)
    try:
        _ = sys.stdout.buffer.write((_out + "\n").encode("utf-8"))
    except AttributeError:
        print(_out)

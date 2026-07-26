#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""自动评估“智慧农业示范园建设项目报价表”Excel文件。

对外仅暴露 ``evaluate(dir_path: str) -> dict``：调用方传入脚本所在目录，
脚本自行在该目录中定位并打开待评估的.xlsx/.xlsm文件，返回结构化评分字典。

说明：
- 先执行“维度1：可用与可修改性”门槛检查；任一门槛不满足时直接输出0分，并跳过维度2。
- 维度1通过后，再累计维度2得分点和扣分点，命中项与未命中项都会出现在返回结果中。
- 对细则中存在行号矛盾的新增表格（例如第5张表同时要求第75行为记录和小计），脚本按“可编辑Excel表格且合计行在数据行之后”的评估意图实现：优先检查目标区域附近的完整表格结构，而不是依赖截图或人工判断。
"""

from __future__ import annotations

import math
import os

SCRIPT_ID = "092"
import re
import sys
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries

ERROR_VALUES = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
MAIN_SHEET = "智慧农业报价表"

TABLE_HEADERS = ["序号", "项目名称", "项目特征描述", "工程量", "单位", "综合单价（元）", "合价（元）", "施工/供应说明", "使用部位", "备注"]
EXPECTED_WIDTHS = {
    "A": 7.38,
    "B": 17.38,
    "C": 27.38,
    "D": 9.38,
    "E": 7.38,
    "F": 13.38,
    "G": 15.38,
    "H": 11.38,
    "I": 19.38,
    "J": 23.38,
}
FORMULA_ERROR_RE = re.compile(r"#(?:REF|VALUE|NAME\?|DIV/0|N/A|NULL|NUM)!?", re.I)


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str = ""


@dataclass
class ScorePoint:
    name: str
    points: int
    hit: bool
    detail: str = ""


@dataclass
class Report:
    dimension1: list[CheckResult] = field(default_factory=list)
    dimension2: list[ScorePoint] = field(default_factory=list)

    @property
    def dim1_passed(self) -> bool:
        return all(item.passed for item in self.dimension1)

    @property
    def total_score(self) -> int:
        if not self.dim1_passed:
            return 0
        return sum(item.points for item in self.dimension2 if item.hit)


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace(" ", " ").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def norm_formula(value: Any) -> str:
    return re.sub(r"\s+", "", norm_text(value)).upper()


def is_blank(value: Any) -> bool:
    return norm_text(value) == ""


def same_value(actual: Any, expected: Any, tol: float = 0.01) -> bool:
    if expected is None:
        return is_blank(actual)
    if isinstance(expected, (int, float)):
        try:
            return abs(float(actual) - float(expected)) <= tol
        except Exception:
            return False
    return norm_text(actual) == norm_text(expected)


def cell_key(cell) -> tuple:
    border = cell.border
    fill = cell.fill.fgColor
    color = cell.font.color
    return (
        cell.font.name,
        round(float(cell.font.sz or 0), 2),
        bool(cell.font.bold),
        color.type if color else None,
        color.rgb if color and color.type == "rgb" else color.indexed if color and color.type == "indexed" else None,
        fill.type,
        fill.rgb if fill.type == "rgb" else fill.indexed if fill.type == "indexed" else None,
        cell.alignment.horizontal,
        cell.alignment.vertical,
        bool(cell.alignment.wrap_text),
        border.left.style,
        border.right.style,
        border.top.style,
        border.bottom.style,
    )


def font_name(cell) -> str:
    return norm_text(cell.font.name)


def font_size(cell) -> float:
    try:
        return float(cell.font.sz or 0)
    except Exception:
        return 0.0


def color_id(color) -> str:
    if color is None:
        return ""
    if color.type == "rgb":
        return str(color.rgb or "").upper()
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return str(color.type)


def is_black_or_dark_gray(cell) -> bool:
    # 细则要求字体为“黑色或深灰色”。仅认可黑色及偏深的灰色，
    # 不认可中灰及更浅的颜色，也不认可彩色。
    color = cell.font.color
    if color is None:
        # 未设置颜色时Excel默认渲染为黑色。
        return True
    if color.type == "theme":
        # 主题黑色记录为theme=0/1且tint为0；带正tint会变浅，不算深灰。
        return color.theme in (0, 1) and (color.tint is None or float(color.tint) <= 1e-9)
    if color.type == "indexed":
        # 0/8=黑，64=自动(黑)。
        return color.indexed in (0, 8, 64)
    if color.type == "rgb":
        rgb = str(color.rgb or "").upper()
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) != 6:
            return False
        try:
            r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
        except ValueError:
            return False
        # 深灰阈值：各通道均不超过64(约25%亮度)，且接近中性灰(通道差不大)。
        return max(r, g, b) <= 64 and (max(r, g, b) - min(r, g, b)) <= 24
    return False


def iter_nonempty_cells(ws, row_start: int, row_end: int, col_start: int = 1, col_end: int = 10):
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell) and not is_blank(cell.value):
                yield cell


def merged_range_for_cell(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def merged_width(ws, row: int, col: int) -> float:
    rng = merged_range_for_cell(ws, row, col)
    cols = range(col, col + 1) if rng is None else range(rng.min_col, rng.max_col + 1)
    total = 0.0
    for c in cols:
        letter = get_column_letter(c)
        total += float(ws.column_dimensions[letter].width or 8.43)
    return max(total, 1.0)


# 行高估算参数（微软雅黑10磅，贴近Excel/WPS实际渲染）：
# - 单个中文字约占1.4个列宽字符单位
# - 半角字符约0.72个列宽字符单位
# - 单元格水平内边距约1.6个字符单位
# - 单行行高约14pt
# - 垂直留白约3pt
# - 单行文字不论行高多少都能完整显示（不折行则不存在行高裁切问题）
CJK_CHAR_WIDTH = 1.4
ASCII_CHAR_WIDTH = 0.72
CELL_H_PADDING = 1.6  # 单元格水平内边距（列宽字符单位）
LINE_HEIGHT_PT = 14.0  # 微软雅黑10磅单行所需高度（磅）
CELL_V_PADDING_PT = 3.0  # 单元格上下留白（磅）


def estimated_required_row_height(ws, row: int) -> tuple[int, float]:
    """返回 (max_lines, required_height_pt)。
    max_lines 是该行所有单元格中估算折行数的最大值；
    required_height_pt 是对应所需最小行高（磅）。
    单行内容 max_lines==1 时调用方可跳过行高检查。"""
    max_lines = 1
    max_font = 10.0
    for col in range(1, 11):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell) or is_blank(cell.value):
            continue
        # 公式单元格显示的是计算结果，不按公式文本估算
        if isinstance(cell.value, str) and cell.value.startswith('='):
            continue
        max_font = max(max_font, font_size(cell) or 10.0)
        text = norm_text(cell.value)
        if not text:
            continue
        # 未开启自动换行的单元格不会折行，文字在单行内显示，不存在裁切问题。
        if not cell.alignment.wrap_text:
            continue
        width = merged_width(ws, row, col)
        usable = max(width - CELL_H_PADDING, 1.0)
        logical_lines = text.split('\n')
        required = 0
        for line in logical_lines:
            visual_len = sum(CJK_CHAR_WIDTH if ord(ch) > 127 else ASCII_CHAR_WIDTH for ch in line)
            required += max(1, math.ceil(visual_len / usable))
        max_lines = max(max_lines, required)
    line_pt = max(LINE_HEIGHT_PT, max_font * 1.35)
    return max_lines, max_lines * line_pt + CELL_V_PADDING_PT


def row_height_truncation_issues(ws, row_start: int = 1, row_end: int = 124) -> list[str]:
    # 检测表内文字是否因行高过窄而在WPS中显示不完整。
    # 仅当估算折行数>=2（即内容确实需要换行）时才检查行高是否足够。
    # 单行文字无论行高多少都能完整显示，因此单行内容不纳入检查，
    # 以贴近WPS实际渲染结果。
    issues: list[str] = []
    for row in range(row_start, row_end + 1):
        has_text = any(
            not isinstance(ws.cell(row, col), MergedCell) and not is_blank(ws.cell(row, col).value)
            for col in range(1, 11)
        )
        if not has_text:
            continue
        height = float(ws.row_dimensions[row].height or 15)
        max_lines, required = estimated_required_row_height(ws, row)
        # 单行内容：行高只要能容纳一行文字即可，不纳入行高裁切检查。
        # 多行内容：估算折行后总高度，超出行高2pt即判定显示不完整。
        if max_lines >= 2 and required > height + 2:
            issues.append('第' + str(row) + '行行高约' + str(int(height)) + '，文字估算需' + str(round(required, 1)) + '，过窄会显示不完整')
    return issues


def get_sheet_drawing_objects(xlsx_path: Path, ws) -> list[dict[str, Any]]:
    """读取指定工作表drawing锚点，覆盖图片、形状、图表等不可编辑对象。"""
    objects: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            sheet_path = ws.path.lstrip("/")
            rels_path = str(Path(sheet_path).parent / "_rels" / (Path(sheet_path).name + ".rels")).replace("\\", "/")
            if sheet_path not in zf.namelist():
                return objects
            sheet_xml = ET.fromstring(zf.read(sheet_path))
            ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
            drawing_ids = []
            for elem in sheet_xml.iter():
                if elem.tag.endswith("drawing"):
                    rid = elem.attrib.get(ns_rel + "id")
                    if rid:
                        drawing_ids.append(rid)
            if not drawing_ids or rels_path not in zf.namelist():
                return objects
            rels_xml = ET.fromstring(zf.read(rels_path))
            rel_targets = {}
            for rel in rels_xml:
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                if rid:
                    base = Path(sheet_path).parent
                    target_path = str((base / target).as_posix())
                    # Path不能规范化包含..的zip路径，手工消解。
                    parts = []
                    for part in target_path.split("/"):
                        if part == "..":
                            if parts:
                                parts.pop()
                        elif part and part != ".":
                            parts.append(part)
                    rel_targets[rid] = "/".join(parts)
            for rid in drawing_ids:
                drawing_path = rel_targets.get(rid)
                if not drawing_path or drawing_path not in zf.namelist():
                    continue
                drawing_xml = ET.fromstring(zf.read(drawing_path))
                for anchor in drawing_xml:
                    local = anchor.tag.split("}")[-1]
                    if local not in {"twoCellAnchor", "oneCellAnchor", "absoluteAnchor"}:
                        continue
                    obj_type = "object"
                    for child in anchor:
                        name = child.tag.split("}")[-1]
                        if name in {"pic", "sp", "graphicFrame", "grpSp", "cxnSp"}:
                            obj_type = name
                            break
                    from_elem = next((c for c in anchor if c.tag.endswith("from")), None)
                    to_elem = next((c for c in anchor if c.tag.endswith("to")), None)
                    if from_elem is None:
                        objects.append({"type": obj_type, "row1": None, "col1": None, "row2": None, "col2": None})
                        continue

                    def child_int(parent, name, default=0):
                        elem = next((c for c in parent if c.tag.endswith(name)), None)
                        if elem is None or elem.text is None:
                            return default
                        return int(elem.text)

                    col1 = child_int(from_elem, "col") + 1
                    row1 = child_int(from_elem, "row") + 1
                    if to_elem is not None:
                        col2 = child_int(to_elem, "col", col1 - 1) + 1
                        row2 = child_int(to_elem, "row", row1 - 1) + 1
                    else:
                        col2, row2 = col1, row1
                    objects.append({"type": obj_type, "row1": row1, "col1": col1, "row2": row2, "col2": col2})
    except Exception as exc:
        objects.append({"type": "parse-error", "error": str(exc), "row1": None, "col1": None, "row2": None, "col2": None})
    return objects


def overlaps(obj: dict[str, Any], min_row: int, max_row: int, min_col: int, max_col: int) -> bool:
    if obj.get("row1") is None:
        return True
    return not (obj["row2"] < min_row or obj["row1"] > max_row or obj["col2"] < min_col or obj["col1"] > max_col)


def workbook_object_summary(xlsx_path: Path, wb) -> dict[str, int]:
    summary = {"media": 0, "charts": 0, "comments": 0, "drawings": 0, "images_openpyxl": 0, "charts_openpyxl": 0}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = zf.namelist()
            summary["media"] = sum(1 for n in names if n.startswith("xl/media/"))
            summary["charts"] = sum(1 for n in names if n.startswith("xl/charts/"))
            summary["comments"] = sum(1 for n in names if n.startswith("xl/comments") or "/comments" in n)
            summary["drawings"] = sum(1 for n in names if n.startswith("xl/drawings/drawing"))
    except Exception:
        pass
    for ws in wb.worksheets:
        summary["images_openpyxl"] += len(getattr(ws, "_images", []) or [])
        summary["charts_openpyxl"] += len(getattr(ws, "_charts", []) or [])
    return summary


def simple_formula_eval(formula: Any, ws, stack: set[str] | None = None) -> float | None:
    """仅用于本题常见的SUM、乘法、加法和单元格引用。不能当通用Excel计算引擎。"""
    if not isinstance(formula, str) or not formula.startswith("="):
        try:
            return float(formula)
        except Exception:
            return None
    expr = formula[1:].strip().upper()
    stack = stack or set()

    def cell_value(ref: str) -> float:
        ref = ref.replace("$", "")
        if ref in stack:
            return 0.0
        value = ws[ref].value
        if isinstance(value, str) and value.startswith("="):
            result = simple_formula_eval(value, ws, stack | {ref})
            return float(result or 0)
        try:
            return float(value or 0)
        except Exception:
            return 0.0

    def sum_range(match: re.Match) -> str:
        ref = match.group(1).replace("$", "")
        if ":" not in ref:
            return str(cell_value(ref))
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        total = 0.0
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                total += cell_value(f"{get_column_letter(c)}{r}")
        return str(total)

    expr = re.sub(r"SUM\(([^()]+)\)", sum_range, expr)
    expr = re.sub(r"\$?[A-Z]{1,3}\$?\d+", lambda m: str(cell_value(m.group(0))), expr)
    expr = expr.replace("%", "/100")
    if not re.fullmatch(r"[0-9+\-*/(). ]+", expr):
        return None
    try:
        return float(eval(expr, {"__builtins__": {}}, {}))
    except Exception:
        return None


def formula_result(ws_formula, ws_values, coord: str) -> float | None:
    cached = ws_values[coord].value
    try:
        if cached is not None:
            return float(cached)
    except Exception:
        pass
    return simple_formula_eval(ws_formula[coord].value, ws_formula)


def check_row_values(ws_formula, row: int, expected: list[Any], ws_values=None) -> list[str]:
    errors = []
    ws_values = ws_values or ws_formula
    for idx, exp in enumerate(expected, start=1):
        cell = ws_formula.cell(row, idx)
        actual = cell.value
        if isinstance(exp, (int, float)) and isinstance(actual, str) and actual.startswith("="):
            actual_value = formula_result(ws_formula, ws_values, cell.coordinate)
            if actual_value is None or abs(float(actual_value) - float(exp)) > 0.01:
                errors.append(f"{cell.coordinate}期望结果{exp!r}，实际公式{actual!r}结果{actual_value!r}")
            continue
        if not same_value(actual, exp):
            errors.append(f"{get_column_letter(idx)}{row}期望{exp!r}，实际{actual!r}")
    return errors


def check_header_row(ws, row: int) -> list[str]:
    return check_row_values(ws, row, TABLE_HEADERS)


def find_table_by_title(ws, title: str, row_min: int, row_max: int) -> int | None:
    for row in range(row_min, row_max + 1):
        if norm_text(ws.cell(row, 1).value) == title:
            return row
    return None


def validate_equipment_table(ws_formula, ws_values, title: str, title_row_hint: int, records: list[list[Any]], subtotal_text: str, subtotal_formula_options: Iterable[str], subtotal_value: float) -> tuple[bool, str]:
    # 评分标准要求内容必须严格落在细则给定的行号上：标题=title_row_hint，
    # 表头=+1，记录=+2起连续len(records)行，小计=紧随最后一条记录的下一行。
    # 不做文字自动定位，也不容忍中间空行或整体错位——错位即判不通过。
    title_row = title_row_hint
    header_row = title_row + 1
    data_start = title_row + 2
    subtotal_row = data_start + len(records)
    errors: list[str] = []

    # 细则：A?:J?标题行合并并显示标题文本。
    title_rng = merged_range_for_cell(ws_formula, title_row, 1)
    if not (title_rng and title_rng.min_col == 1 and title_rng.max_col == 10 and title_rng.min_row == title_rng.max_row == title_row):
        errors.append(f"第{title_row}行标题未合并为A:J")
    if norm_text(ws_formula.cell(title_row, 1).value) != title:
        errors.append(f"A{title_row}标题文本期望{title!r}，实际{ws_formula.cell(title_row, 1).value!r}")

    # 细则：表头A?:J?依次显示十列表头。
    errors.extend(check_header_row(ws_formula, header_row))

    # 细则：每条记录A列起依次显示各字段，且每条记录J列(备注)保持空白。
    for i, record in enumerate(records):
        row = data_start + i
        errors.extend(check_row_values(ws_formula, row, record + [None], ws_values))

    # 细则：小计行A?:F?合并并显示小计文本，G列用SUM公式，结果为指定值。
    subtotal_cell = ws_formula.cell(subtotal_row, 1)
    if norm_text(subtotal_cell.value) != subtotal_text:
        errors.append(f"A{subtotal_row}小计文本期望{subtotal_text!r}，实际{subtotal_cell.value!r}")
    subtotal_rng = merged_range_for_cell(ws_formula, subtotal_row, 1)
    if not (subtotal_rng and subtotal_rng.min_col == 1 and subtotal_rng.max_col == 6 and subtotal_rng.min_row == subtotal_rng.max_row == subtotal_row):
        errors.append(f"第{subtotal_row}行小计未合并为A:F")
    formula = norm_formula(ws_formula.cell(subtotal_row, 7).value)
    if formula not in {norm_formula(f) for f in subtotal_formula_options}:
        errors.append(f"G{subtotal_row}公式不正确：{ws_formula.cell(subtotal_row, 7).value!r}")
    actual_value = formula_result(ws_formula, ws_values, f"G{subtotal_row}")
    if actual_value is None or abs(actual_value - subtotal_value) > 0.01:
        errors.append(f"G{subtotal_row}结果期望{subtotal_value:.2f}，实际{actual_value!r}")

    # 细则：区域内字体、行间距等皆为可编辑。除标题A:J和小计A:F外不应有其它合并占位，
    # 否则正文被合并单元格覆盖，影响可编辑性。
    for row in range(title_row, subtotal_row + 1):
        for col in range(1, 11):
            cell = ws_formula.cell(row, col)
            if isinstance(cell, MergedCell):
                rng = merged_range_for_cell(ws_formula, row, col)
                allowed_merge = rng and ((row == title_row and rng.min_col == 1 and rng.max_col == 10) or (row == subtotal_row and rng.min_col == 1 and rng.max_col == 6))
                if not allowed_merge:
                    errors.append(f"{cell.coordinate}为非预期合并单元格，影响可编辑性")

    detail = f"定位标题行：{title_row}，小计行：{subtotal_row}"
    if errors:
        detail += "；" + "；".join(errors[:8])
        if len(errors) > 8:
            detail += f"；另有{len(errors) - 8}项"
    return not errors, detail


def check_dimension1(path: Path, wb_formula, wb_values) -> list[CheckResult]:
    results: list[CheckResult] = []

    results.append(CheckResult(
        "交付文件为.xlsx或.xlsm格式，且可正常打开",
        path.suffix.lower() in {".xlsx", ".xlsm"} and wb_formula is not None,
        f"扩展名：{path.suffix}"
    ))
    if wb_formula is None:
        return results

    if MAIN_SHEET not in wb_formula.sheetnames:
        results.append(CheckResult("“智慧农业报价表”存在", False, "缺少主工作表"))
        return results

    return results


def check_dimension2(path: Path, wb_formula, wb_values) -> list[ScorePoint]:
    del path  # 打印分页检查已移除，path 保留在签名以兼容调用方
    ws = wb_formula[MAIN_SHEET]
    ws_values = wb_values[MAIN_SHEET] if wb_values and MAIN_SHEET in wb_values.sheetnames else ws
    points: list[ScorePoint] = []

    data_rows = []
    for start, end in [(7, 20), (27, 31), (34, 38), (41, 46), (49, 54), (61, 66), (70, 76), (81, 86), (91, 96), (101, 106), (111, 116), (121, 124)]:
        data_rows.extend(range(start, end + 1))
    font_errors = []
    for row in data_rows:
        for col in range(1, 11):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell) or is_blank(cell.value):
                continue
            # 细则：中文、英文、数字和符号均使用微软雅黑10磅、黑色或深灰色。
            if font_name(cell) != "微软雅黑":
                font_errors.append(f"{cell.coordinate}字体非微软雅黑({font_name(cell)})")
            if abs(font_size(cell) - 10.0) > 0.1:
                font_errors.append(f"{cell.coordinate}字号非10磅({font_size(cell):g})")
            if not is_black_or_dark_gray(cell):
                font_errors.append(f"{cell.coordinate}字体颜色非黑色或深灰色")
            # 细则：所有单元格垂直居中并开启自动换行。
            if cell.alignment.vertical != "center":
                font_errors.append(f"{cell.coordinate}未垂直居中")
            if cell.alignment.wrap_text is not True:
                font_errors.append(f"{cell.coordinate}未开启自动换行")
    points.append(ScorePoint(
        "表内字体为微软雅黑10磅、黑色或深灰色，垂直居中并自动换行",
        3,
        not font_errors,
        "全部符合" if not font_errors else "；".join(font_errors[:8]) + (f"；另有{len(font_errors)-8}项" if len(font_errors) > 8 else "")
    ))

    env_records = [
        [1, "环流风机", "低噪声轴流，防潮电机", 36.00, "台", 1850.00, 66600.00, "设备供应", "温室内部"],
        [2, "湿帘纸芯", "150mm厚高效蒸发湿帘", 220.00, "㎡", 165.00, 36300.00, "设备供应", "湿帘墙"],
        [3, "LED补光灯", "植物生长光谱，防水等级IP65", 168.00, "盏", 680.00, 114240.00, "设备供应", "育苗区"],
        [4, "遮阳幕布", "外白内黑节能幕，阻燃处理", 8600.00, "㎡", 22.00, 189200.00, "设备供应", "顶部遮阳"],
        [5, "电动卷膜器", "24V直流，带限位保护", 44.00, "套", 960.00, 42240.00, "设备供应", "侧窗通风"],
    ]
    ok, detail = validate_equipment_table(
        ws,
        ws_values,
        "五、环控与照明物资清单",
        69,
        env_records,
        "五、环控与照明物资清单 小计",
        ["=SUM(G71:G75)"],
        448580.00,
    )
    points.append(ScorePoint(
        "A69:J76环控与照明物资表内容、公式和可编辑性完整",
        5,
        ok,
        detail,
    ))

    iot_records = [
        [1, "温湿度传感器", "温度、湿度双测点，防水外壳", 28.00, "支", 520.00, 14560.00, "设备供应", "栽培区"],
        [2, "CO₂传感器", "NDIR原理，量程0-5000ppm", 12.00, "支", 1680.00, 20160.00, "设备供应", "生产区"],
        [3, "土壤/基质水分传感器", "电容式，带温度补偿", 24.00, "支", 780.00, 18720.00, "设备供应", "种植槽"],
        [4, "LoRa采集网关", "工业级，支持断点续传", 6.00, "台", 3900.00, 23400.00, "设备供应", "控制间"],
        [5, "生产大屏看板", "55英寸显示屏，含支架与信号盒", 2.00, "套", 6800.00, 13600.00, "设备供应", "运营中心"],
    ]
    ok, detail = validate_equipment_table(
        ws,
        ws_values,
        "六、物联网设备预算清单",
        79,
        iot_records,
        "六、物联网设备预算清单 小计",
        ["=SUM(G81:G85)"],
        90440.00,
    )
    points.append(ScorePoint(
        "A79:J86物联网设备表内容、公式和可编辑性完整",
        5,
        ok,
        detail,
    ))

    return points


def _find_target_file(dir_path: Path) -> Path | None:
    """在脚本所在目录里定位待评估的.xlsx/.xlsm文件。

    - 忽略Office临时文件（``~$`` 开头）。
    - 若有多个候选，优先选文件名含“智慧农业”或“报价”的；否则取修改时间最新的。
    """
    candidates = [
        p for p in dir_path.iterdir()
        if p.is_file()
        and p.suffix.lower() in {".xlsx", ".xlsm"}
        and not p.name.startswith("~$")
    ]
    if not candidates:
        return None
    for keyword in ("智慧农业", "报价"):
        for p in candidates:
            if keyword in p.name:
                return p
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _score_point_to_dict(sp: ScorePoint) -> dict[str, object]:
    """把内部ScorePoint转成统一约定中的dim2条目结构。"""
    max_delta = sp.points
    delta = sp.points if sp.hit else 0
    return {
        "rule": sp.name,
        "max_delta": max_delta,
        "delta": delta,
        "hit": sp.hit,
        "detail": "",
    }


def evaluate(dir_path: str) -> dict[str, object]:
    """评估 ``dir_path`` 目录内的智慧农业报价表.xlsx/.xlsm文件。

    调用方只需传入脚本所在目录，脚本自行在其中定位并打开被评估的文档，
    返回符合“脚本接口差异与统一建议”§2.2 约定的结构化字典。
    """
    result: dict[str, object] = {
        "id": "092",
        "file_name": "",
        "status": "ok",
        "error": None,
        "dim1_pass": False,
        "dim1_reason": "",
        "dim2_items": [],
        "total_score": 0,
        "max_score": 0,
    }
    try:
        base_dir = Path(dir_path)
        if not base_dir.exists() or not base_dir.is_dir():
            result["status"] = "error"
            result["error"] = f"目录不存在或不是目录：{dir_path}"
            return result

        target = _find_target_file(base_dir)
        if target is None:
            result["status"] = "error"
            result["error"] = f"目录中未找到.xlsx/.xlsm文件：{dir_path}"
            return result
        result["file_name"] = target.name

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb_formula = load_workbook(target, data_only=False)
                wb_values = load_workbook(target, data_only=True)
        except Exception as exc:
            result["dim1_pass"] = False
            result["dim1_reason"] = f"交付文件无法打开：{exc}"
            return result

        dim1 = check_dimension1(target, wb_formula, wb_values)
        dim1_pass = all(item.passed for item in dim1)
        result["dim1_pass"] = dim1_pass
        if not dim1_pass:
            result["dim1_reason"] = "；".join(
                f"{item.name}：{item.detail}" for item in dim1 if not item.passed
            )
            return result

        dim2 = check_dimension2(target, wb_formula, wb_values)
        dim2_items = [_score_point_to_dict(sp) for sp in dim2]
        result["dim2_items"] = dim2_items
        result["total_score"] = sum(int(sp.points) for sp in dim2 if sp.hit)
        # 满分只累计正向得分项；扣分项的max_delta是负数，不进入满分基数。
        result["max_score"] = sum(int(sp.points) for sp in dim2 if sp.points > 0)
        return result
    except Exception as exc:
        result["status"] = "error"
        result["error"] = f"{type(exc).__name__}: {exc}"
        return result


if __name__ == "__main__":
    import json

    target_dir = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).resolve().parent)
    print(json.dumps(evaluate(target_dir), ensure_ascii=False, indent=2))
